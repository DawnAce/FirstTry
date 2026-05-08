"""Seed shipping details for issue 2649 from ZTO Express Excel file."""
import os
from datetime import datetime
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from app.models.shipping_detail import ShippingDetail

EXCEL_PATH = (
    r"C:\Users\luyal\xwechat_files\wxid_np9el7e4tpoq22_8fa5"
    r"\msg\file\2026-05"
    r"\2026年4月27日《中国经营报》中通快递发货明细（2649）.xlsx"
)
ISSUE_NUMBER = 2649


def _str(val: Any) -> str | None:
    """Convert a cell value to a stripped string, or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _int(val: Any) -> int:
    """Coerce a cell value to int, defaulting to 0."""
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _deadline_str(val: Any) -> str | None:
    """Convert deadline to string. Handles datetime objects and plain strings."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s if s else None


def _is_summary_row(phone_val: Any) -> bool:
    """Check if the row is a summary/total row (phone column == '合计')."""
    return _str(phone_val) == "合计"


def _classify_weekly_corporate(name: str | None, address: str | None, notes: str | None) -> tuple[str, str]:
    """Return (channel, transport) for 每周（对公） rows."""
    if name == "马飞":
        return "库房留存", "库房留存"
    if not name or name == "(未填写)":
        if address and "中通库房" in address:
            return "库房留存", "库房留存"
    if notes:
        if "广州日报" in notes or "杂志铺" in notes:
            return "渠道订阅", "中通物流"
        if "社用报" in notes:
            return "报社留存", "包车运输"
    if name in ("李广", "纪玉文", "党鹏", "王金龙"):
        return "记者站", "中通物流"
    # Fallback based on name
    if name == "叶剑":
        return "渠道订阅", "中通物流"
    if name == "肖波":
        return "渠道订阅", "中通物流"
    if name == "程先生":
        return "报社留存", "包车运输"
    return "对公订阅", "中通物流"


def _parse_weekly_corporate(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Parse 每周（对公） sheet: rows 3-11, 6 columns."""
    records = []
    for row in ws.iter_rows(min_row=3, max_row=11, max_col=6, values_only=True):
        name_val, address, phone, qty, publication, notes = row
        if _is_summary_row(phone):
            continue
        if not any(row):
            continue
        name = _str(name_val) or "(未填写)"
        notes_s = _str(notes)
        address_s = _str(address)
        channel, transport = _classify_weekly_corporate(name, address_s, notes_s)
        records.append({
            "sheet_name": "每周（对公）",
            "channel": channel,
            "transport": transport,
            "frequency": "周",
            "status": "正常",
            "name": name,
            "address": address_s,
            "phone": _str(phone),
            "quantity": _int(qty),
            "notes": notes_s,
        })
    return records


def _parse_weekly_reader(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Parse 每周（读者） sheet: rows 3-31, 8 columns.
    First 5 data rows (rows 3-7) are 监管赠阅, rest are 个人订户.
    """
    records = []
    data_index = 0
    for row in ws.iter_rows(min_row=3, max_row=31, max_col=8, values_only=True):
        name, address, phone, qty, publication, deadline, notes, extra = row
        if _is_summary_row(phone):
            continue
        if not any(row):
            continue
        channel = "监管赠阅" if data_index < 5 else "个人订户"
        data_index += 1
        records.append({
            "sheet_name": "每周（读者）",
            "channel": channel,
            "transport": "中通物流",
            "frequency": "周",
            "status": "正常",
            "name": _str(name) or "(未填写)",
            "address": _str(address),
            "phone": _str(phone),
            "quantity": _int(qty),
            "deadline": _deadline_str(deadline),
            "notes": _str(notes),
            "extra_info": _str(extra),
        })
    return records


def _parse_high_speed_rail(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Parse 高铁展示 sheet: rows 4-26, 10 columns. Carry forward city name."""
    records = []
    current_city = None
    for row in ws.iter_rows(min_row=4, max_row=26, max_col=10, values_only=True):
        city, seq, station, hall, contact, phone, address, qty, confirm, extra = row
        # Skip rows where both seq and station are empty (after data ends)
        if seq is None and station is None:
            continue
        if _str(city):
            current_city = _str(city)
        records.append({
            "sheet_name": "高铁展示",
            "channel": "对公订阅",
            "transport": "中通物流",
            "frequency": "周",
            "status": "正常",
            "name": _str(contact) or "(未填写)",
            "city": current_city,
            "seq_number": _int(seq),
            "station_name": _str(station),
            "station_hall": _str(hall),
            "contact_person": _str(contact),
            "phone": _str(phone),
            "address": _str(address),
            "quantity": _int(qty),
            "confirmation": _str(confirm),
            "extra_info": _str(extra),
        })
    return records


def _parse_shangyou(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Parse 上犹 sheet: rows 3-5, 6 columns."""
    records = []
    for row in ws.iter_rows(min_row=3, max_row=5, max_col=6, values_only=True):
        name, address, phone, qty, publication, notes = row
        if _is_summary_row(phone):
            continue
        if not any(row):
            continue
        notes_s = _str(notes)
        transport = "邮政物流" if notes_s and "邮政" in notes_s else "中通物流"
        records.append({
            "sheet_name": "上犹",
            "channel": "政府赠阅",
            "transport": transport,
            "frequency": "周",
            "status": "正常",
            "name": _str(name) or "(未填写)",
            "address": _str(address),
            "phone": _str(phone),
            "quantity": _int(qty),
            "notes": notes_s,
        })
    return records


def _parse_biweekly_suspended(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Parse 停发-双周（读者） sheet: rows 3-4, 8 columns."""
    records = []
    for row in ws.iter_rows(min_row=3, max_row=4, max_col=8, values_only=True):
        name, address, phone, period, qty, publication, deadline, notes = row
        if _is_summary_row(phone):
            continue
        if not any(row):
            continue
        records.append({
            "sheet_name": "停发-双周（读者）",
            "channel": "个人订户",
            "transport": "中通物流",
            "frequency": "半月",
            "status": "停发",
            "name": _str(name) or "(未填写)",
            "address": _str(address),
            "phone": _str(phone),
            "period_count": _int(period),
            "quantity": _int(qty),
            "deadline": _deadline_str(deadline),
            "notes": _str(notes),
        })
    return records


def _parse_monthly(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Parse 月底-整月 sheet: rows 3-22, 9 columns.
    First 4 data rows are 监管赠阅, row with 国图贸 in notes is 渠道订阅, rest are 个人订户.
    """
    records = []
    data_index = 0
    for row in ws.iter_rows(min_row=3, max_row=22, max_col=9, values_only=True):
        name, address, phone, period, qty, publication, deadline, notes, extra = row
        if _is_summary_row(phone):
            continue
        if not any(row):
            continue
        notes_s = _str(notes)
        if data_index < 4:
            channel = "监管赠阅"
        elif notes_s and "国图贸" in notes_s:
            channel = "渠道订阅"
        else:
            channel = "个人订户"
        data_index += 1
        records.append({
            "sheet_name": "月底-整月",
            "channel": channel,
            "transport": "中通物流",
            "frequency": "月",
            "status": "正常",
            "name": _str(name) or "(未填写)",
            "address": _str(address),
            "phone": _str(phone),
            "period_count": _int(period),
            "quantity": _int(qty),
            "deadline": _deadline_str(deadline),
            "notes": notes_s,
            "extra_info": _str(extra),
        })
    return records


# Map sheet names to their parser functions
SHEET_PARSERS = {
    "每周（对公）": _parse_weekly_corporate,
    "每周（读者）": _parse_weekly_reader,
    "高铁展示": _parse_high_speed_rail,
    "上犹": _parse_shangyou,
    "停发-双周（读者）": _parse_biweekly_suspended,
    "月底-整月": _parse_monthly,
}


def seed_shipping_details_2649(db: Session, excel_path: str = EXCEL_PATH) -> int:
    """Seed shipping details for issue 2649. Returns number of records inserted."""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Delete existing records for this issue
    deleted = db.query(ShippingDetail).filter(
        ShippingDetail.issue_number == ISSUE_NUMBER
    ).delete()
    if deleted:
        print(f"Deleted {deleted} existing records for issue {ISSUE_NUMBER}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    all_records: list[dict] = []
    for sheet_name, parser in SHEET_PARSERS.items():
        if sheet_name not in wb.sheetnames:
            print(f"Warning: sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        records = parser(ws)
        print(f"  {sheet_name}: {len(records)} records")
        all_records.extend(records)

    wb.close()

    # Insert all records
    for rec in all_records:
        detail = ShippingDetail(issue_number=ISSUE_NUMBER, **rec)
        db.add(detail)

    db.commit()
    print(f"Total: {len(all_records)} records inserted for issue {ISSUE_NUMBER}")
    return len(all_records)
