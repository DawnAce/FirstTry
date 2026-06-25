"""Parse a 淘宝/天猫 order-export workbook into the SAME structured rows the CBJ
parser produces, so the whole downstream (product resolver → status map → coverage
→ dedup → order create) is reused unchanged.

The Taobao export differs from CBJ in three ways that this module bridges:

* **Different columns.** 21 inline-string columns (订单编号 / 商品标题 / 买家实付金额 /
  收货地址 / 商品属性SKU / 商家备注 …) instead of CBJ's 订单号 / 产品名称 / 地址.
* **Delivery + issue live in the SKU column** (``商品属性SKU`` = ``分册名:全年-邮局-周投``
  / ``分册名:2026年5月刊``), not in the title. We fold the 分册名 value into the
  product-line name so the existing alias matching and 商学院 month-label
  auto-detection (``normalize_business_school_issue_label``) work as-is.
* **Recipient is desensitized** (``收货地址`` = ``江苏省 南京市 … 街道****``; no name /
  phone / detail, no separate 收货人 column). So recipient is left blank — the order
  imports as a record and the operator tops up recipient + coverage per order later
  (mirroring the 历史归档 workflow). 收货地址 prefix is kept as a hint only.

Pricing maps as: paid = 买家实付金额 (F, incl. postage); original = 总金额 (E, goods
list price) + 买家应付邮费 (D) — the pre-discount equivalent of the same basket, so
``折扣 = original − paid`` stays consistent with the CBJ ``original_amount`` column.

Multi-product orders (商品标题 comma-joined) only carry a TOTAL qty/amount — no
per-line split — so we even-split the paid amount across titles as a default and
flag the order for manual verification (see ``parse_taobao_orders`` return + the
``multi_product`` marker the dry-run / import surfaces).
"""

import io
import re
from datetime import datetime
from decimal import Decimal
from typing import List, Optional

import openpyxl

from app.services.cbj_order_import_parser import (
    ParsedOrder,
    ProductLine,
    _dec,
    _to_datetime,
)


_HEADER_MAP = {
    "订单编号": "external_order_no",
    "买家应付货款": "goods_payable",   # C — discounted goods (excl postage)
    "买家应付邮费": "postage",          # D — postage
    "总金额": "list_total",            # E — goods list price (pre-discount, excl postage)
    "买家实付金额": "paid_amount",      # F — actually paid (= C + D)
    "订单状态": "status",
    "收货地址": "address",             # desensitized: 省 市 区 街道****  (no name/phone/detail)
    "订单创建时间": "order_time",
    "商品标题": "product",             # comma-joined titles for multi-product orders
    "宝贝总数量": "total_qty",
    "开票信息": "invoice",
    "商品属性SKU": "sku",              # 分册名:… → delivery + 期次
    "商家备注": "merchant_note",        # 起止期 / 电话 often hand-typed here (free text)
    "物流单号": "tracking",
    "物流公司": "logistics",
    "发货时间": "ship_time",
}
# Address is NOT required (it's desensitized); identity + product + amount + status are.
_REQUIRED = {"external_order_no", "product", "paid_amount", "status"}

_SKU_PREFIX_RE = re.compile(r"^\s*分册名\s*[:：]\s*")


def _split_titles(value) -> List[str]:
    """淘宝 商品标题 joins multiple products with a (half/full-width) comma."""
    return [s.strip() for s in re.split(r"[,，]", str(value or "")) if s.strip()]


def _sku_value(value) -> str:
    """``分册名:全年-邮局-周投[（…）]`` → ``全年-邮局-周投[（…）]`` (drop the 分册名 label)."""
    return _SKU_PREFIX_RE.sub("", str(value or "").strip()).strip()


def _int(value, default: int = 1) -> int:
    try:
        n = int(float(str(value).strip()))
        return n if n > 0 else default
    except (TypeError, ValueError):
        return default


def _build_product_lines(title_cell, sku_cell, total_qty, paid) -> List[ProductLine]:
    """Turn 商品标题 (+ 分册名 SKU + total qty/amount) into product lines.

    The 分册名 value is appended to each title so downstream alias matching and the
    商学院 ``YYYY年M月刊`` label detector can read delivery/期次 that Taobao keeps in a
    separate column. Single-product orders are exact; multi-product orders even-split
    the paid amount (Taobao gives no per-line price) and should be operator-verified.
    """
    titles = _split_titles(title_cell)
    sku = _sku_value(sku_cell)
    qty_total = _int(total_qty, default=1)

    def _name_with_sku(title: str) -> str:
        return f"{title} {sku}".strip() if sku else title

    if len(titles) <= 1:
        name = _name_with_sku(titles[0] if titles else "")
        return [
            ProductLine(
                raw=name,
                name=name,
                quantity=qty_total,
                unit_price=Decimal("0"),  # single line → service assigns (paid − shipping)
                is_shipping=False,
                mentions_zto=False,  # Taobao delivery comes from SKU/product, not a 转中通 remark
            )
        ]

    # Multi-product: even-split paid across titles (no per-line price in the export).
    n = len(titles)
    per = (Decimal(str(paid)) / Decimal(n)) if n else Decimal("0")
    lines: List[ProductLine] = []
    for title in titles:
        name = _name_with_sku(title)
        lines.append(
            ProductLine(
                raw=title,
                name=name,
                quantity=1,
                unit_price=per,
                is_shipping=False,
                mentions_zto=False,
            )
        )
    return lines


def _find_header(ws):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        cells = ["" if v is None else str(v).strip() for v in row]
        if "订单编号" in cells and "商品标题" in cells:
            index = {}
            for col_idx, cell in enumerate(cells):
                if cell in _HEADER_MAP:
                    index[_HEADER_MAP[cell]] = col_idx
            return row_idx, index
    return None, None


def is_taobao_export(file_bytes: bytes) -> bool:
    """Cheap header sniff so the import dispatcher can auto-route by platform."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return False
    try:
        ws = wb[wb.sheetnames[0]]
        header_row, _ = _find_header(ws)
        return header_row is not None
    finally:
        wb.close()


def parse_taobao_orders(file_bytes: bytes) -> List[ParsedOrder]:
    """Parse the 淘宝 export bytes into ``ParsedOrder`` rows.

    Raises ValueError if the file is not a recognizable Taobao order export.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("无法读取 Excel 文件") from exc

    try:
        ws = wb[wb.sheetnames[0]]
        header_row, index = _find_header(ws)
        if header_row is None:
            raise ValueError("未找到淘宝订单表头（需要「订单编号」「商品标题」等列）")
        missing = _REQUIRED - set(index)
        if missing:
            labels = {v: k for k, v in _HEADER_MAP.items()}
            raise ValueError("缺少必需列：" + "、".join(labels[m] for m in missing))

        def cell(row, key):
            col = index.get(key)
            return row[col] if col is not None and col < len(row) else None

        orders: List[ParsedOrder] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            external = cell(row, "external_order_no")
            if external is None or str(external).strip() == "":
                continue  # blank / trailing row

            paid = _dec(cell(row, "paid_amount"))
            original = _dec(cell(row, "list_total")) + _dec(cell(row, "postage"))
            created = _to_datetime(cell(row, "order_time"))

            # Preserve the off-platform-fulfilled tracking in notes (the model has no
            # tracking field) so it is not lost; 起止期 hints stay in 商家备注 verbatim.
            note = str(cell(row, "merchant_note") or "").strip()
            tracking = str(cell(row, "tracking") or "").strip()
            logistics = str(cell(row, "logistics") or "").strip()
            if tracking:
                tag = f"淘宝运单 {logistics}{tracking}".strip()
                note = f"{note}｜{tag}" if note else tag

            orders.append(
                ParsedOrder(
                    external_order_no=str(external).strip(),
                    status_raw=str(cell(row, "status") or "").strip(),
                    paid_amount=paid,
                    original_amount=original,
                    order_date=created.date() if created else None,
                    payment_time=created,  # no separate 付款时间 column; 创建时间 is the proxy
                    payment_method_raw="支付宝",  # Taobao → alipay (no column; platform default)
                    invoice_raw=str(cell(row, "invoice") or "").strip(),
                    recipient_name="",          # desensitized — operator tops up later
                    recipient_phone="",
                    recipient_address=str(cell(row, "address") or "").strip(),  # masked (hint)
                    recipient_postal_code=None,
                    notes=note,
                    product_lines=_build_product_lines(
                        cell(row, "product"),
                        cell(row, "sku"),
                        cell(row, "total_qty"),
                        paid,
                    ),
                )
            )
        return orders
    finally:
        wb.close()
