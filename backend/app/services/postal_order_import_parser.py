"""邮局读者明细 Excel 解析（纯提取 → ``ParsedPostalRow``）。

只做「按表头名取值」，不做业务映射（产品/覆盖期/投递单位/去重放在 postal_delivery_import_service）。
兼容两种表：完整库的「邮局读者明细」与单年度的「邮局读者总明细」——都靠表头名匹配、不认列序。
表头签名：同时含 编号 / 姓名 / 起月日 / 投递单位 的那张 sheet 即读者名册。
"""

import io
from dataclasses import dataclass
from typing import List, Optional

import openpyxl

# 读者名册的表头签名（子集命中即认）。
_SIGNATURE = {"编号", "姓名", "起月日", "投递单位"}

# 关心的列（按名取；缺列则该字段为空）。
_FIELDS = {
    "编号": "external_no_raw",
    "地区": "region",
    "新姓名": "new_name",
    "新电话": "new_phone",
    "新地址": "new_address",
    "姓名": "name",
    "联系电话": "phone",
    "省": "province",
    "市": "city",
    "区": "district",
    "详细地址": "detail_address",
    "邮编": "postal_code",
    "年度": "year_raw",
    "产品名称": "product_name",
    "起月日": "start_mmdd",
    "止月日": "end_mmdd",
    "份数": "copies_raw",
    "金额": "amount_raw",
    "渠道": "channel",
    "汇款名称": "remittance_name",
    "汇款日期": "remittance_date_raw",
    "投递单位": "distribution_unit_name",
    "赠阅/关联": "salesperson",
    "备注": "notes",
}


@dataclass
class ParsedPostalRow:
    row_no: int  # 1-based 数据行号（含表头则 +1），用于报错定位
    external_no_raw: str = ""
    region: str = ""
    new_name: str = ""
    new_phone: str = ""
    new_address: str = ""
    name: str = ""
    phone: str = ""
    province: str = ""
    city: str = ""
    district: str = ""
    detail_address: str = ""
    postal_code: str = ""
    year_raw: str = ""
    product_name: str = ""
    start_mmdd: str = ""
    end_mmdd: str = ""
    copies_raw: str = ""
    amount_raw: str = ""
    channel: str = ""
    remittance_name: str = ""
    remittance_date_raw: str = ""
    distribution_unit_name: str = ""
    salesperson: str = ""
    notes: str = ""


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _header_map(header_row) -> dict:
    return {
        _cell(v): idx
        for idx, v in enumerate(header_row)
        if _cell(v)
    }


def _find_reader_sheet(wb):
    """返回 (worksheet, header_index_map) —— 表头命中签名的那张 sheet。"""
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        hmap = _header_map(first)
        if _SIGNATURE.issubset(hmap.keys()):
            return ws, hmap
    return None, None


def is_postal_reader_export(file_bytes: bytes) -> bool:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return False
    ws, _ = _find_reader_sheet(wb)
    wb.close()
    return ws is not None


def parse_postal_readers(file_bytes: bytes) -> List[ParsedPostalRow]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws, hmap = _find_reader_sheet(wb)
    if ws is None:
        wb.close()
        raise ValueError(
            "无法识别的邮局读者明细：未找到含「编号/姓名/起月日/投递单位」表头的工作表"
        )

    col = {field: hmap[name] for name, field in _FIELDS.items() if name in hmap}
    out: List[ParsedPostalRow] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 整行空 → 跳过
        if not any(_cell(v) for v in row):
            continue

        def get(field: str) -> str:
            idx = col.get(field)
            if idx is None or idx >= len(row):
                return ""
            return _cell(row[idx])

        pr = ParsedPostalRow(row_no=i)
        for field in _FIELDS.values():
            setattr(pr, field, get(field))
        # 必要字段：编号 + 姓名 —— 都空视为无效行跳过
        if not pr.external_no_raw and not pr.name:
            continue
        out.append(pr)

    wb.close()
    return out
