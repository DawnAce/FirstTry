"""Parse uploaded history report/shipping workbooks and produce a preview + commit."""

import datetime
import io
from zipfile import BadZipFile

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.history_import_cache import save_history_import_session, pop_history_import_session
from app.models import Issue, IssueStatus, PublicationSchedule, ReportEntry, ReportItemTemplate, ShippingDetail, TempPrintDetail
from app.services.original_zto_shipping_import_service import (
    is_original_zto_shipping_workbook,
    read_original_zto_shipping_basic_info,
    read_original_zto_shipping_rows,
)
from app.services.raw_report_import_service import parse_raw_report_workbook
from app.services.report_destination_service import DESTINATION_ZTO, resolve_report_destination
from app.schemas.history_import import (
    CommitReadiness,
    HistoryImportCommitOut,
    HistoryImportRow,
    TempPrintDetailRow,
    ShippingImportRow,
    HistoryImportPreviewOut,
)

_SHIPPING_COLUMNS = [
    "工作表名称", "渠道", "子渠道", "运输方式", "频次", "状态",
    "姓名", "地址", "电话", "数量", "截止日期", "备注", "附加信息",
    "网点名称", "网点大厅", "联系人", "序号", "期数", "公司",
]


def _normalize_date(value: object) -> str:
    """Return an ISO YYYY-MM-DD string from an Excel cell value (datetime, date, or string)."""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if value is None:
        return ""
    return str(value).strip()


def _parse_issue_number(value: object) -> int | None:
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _is_iso_date(value: str) -> bool:
    if not value:
        return False
    try:
        datetime.date.fromisoformat(value)
        return True
    except ValueError:
        return False


def _read_basic_info(wb) -> dict:
    """Return a key→value dict from the 基本信息 sheet (skipping the header row)."""
    sheet = wb["基本信息"]
    result: dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key, value = (row[0] if len(row) > 0 else None), (row[1] if len(row) > 1 else None)
        if key:
            result[str(key)] = value
    return result


def _read_report_rows(wb) -> list[HistoryImportRow]:
    sheet = wb["报数项"]
    rows: list[HistoryImportRow] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        cat_code, _cat_name, item, dest, is_var, value = (
            row[0] or "", row[1] or "", row[2] or "", row[3] or "",
            row[4], row[5],
        )
        rows.append(HistoryImportRow(
            category=str(cat_code),
            display_name="",        # enriched later from template
            sub_category=str(item),
            destination=str(dest),
            is_variable=(str(is_var).strip() == "是"),
            value=int(value) if value is not None else 0,
        ))
    return rows


def _read_temp_rows(wb) -> list[TempPrintDetailRow]:
    sheet = wb["临时加印明细"]
    rows: list[TempPrintDetailRow] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        dept, custom, qty, self_qty = (
            row[0] or "", row[1] or "",
            row[2] if row[2] is not None else 0,
            row[3] if row[3] is not None else 0,
        )
        rows.append(TempPrintDetailRow(
            department=str(dept),
            custom_name=str(custom),
            quantity=int(qty),
            self_quantity=int(self_qty),
        ))
    return rows


def _is_template_report_workbook(wb) -> bool:
    return {"基本信息", "报数项", "临时加印明细"}.issubset(set(wb.sheetnames))


def _is_template_shipping_workbook(wb) -> bool:
    return {"基本信息", "发货明细"}.issubset(set(wb.sheetnames))


def _read_shipping_rows(wb) -> list[ShippingImportRow]:
    sheet = wb["发货明细"]
    rows: list[ShippingImportRow] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        def _str(v: object) -> str:
            return str(v) if v is not None else ""
        def _int_or_none(v: object) -> int | None:
            return int(v) if v is not None else None
        rows.append(ShippingImportRow(
            sheet_name=_str(row[0]),
            channel=_str(row[1]),
            sub_channel=_str(row[2]),
            transport=_str(row[3]),
            frequency=_str(row[4]),
            status=_str(row[5]),
            name=_str(row[6]),
            address=_str(row[7]),
            phone=_str(row[8]),
            quantity=int(row[9]) if row[9] is not None else 0,
            deadline=_str(row[10]),
            notes=_str(row[11]),
            extra_info=_str(row[12]),
            station_name=_str(row[13]),
            station_hall=_str(row[14]),
            contact_person=_str(row[15]),
            seq_number=_int_or_none(row[16]),
            period_count=_int_or_none(row[17]),
            company=_str(row[18]),
        ))
    return rows


def _validate_and_enrich_report_rows(
    db: Session,
    rows: list[HistoryImportRow],
) -> tuple[list[str], list[HistoryImportRow]]:
    """Validate rows against templates and enrich each with display_name from the template.

    Returns (errors, enriched_rows). When no templates are seeded, validation is skipped
    and rows are returned with empty display_name (preserves existing behavior).
    """
    templates = db.query(ReportItemTemplate).all()
    if not templates:
        return [], rows
    template_map = {(t.category, t.sub_category): t for t in templates}
    errors: list[str] = []
    enriched: list[HistoryImportRow] = []
    for row in rows:
        key = (row.category, row.sub_category)
        if key not in template_map:
            errors.append(
                f"报数项行不在模板定义中：分类编码={row.category!r}，项目名称={row.sub_category!r}"
            )
        else:
            enriched.append(row.model_copy(update={"display_name": template_map[key].display_name}))
    return errors, enriched


def _error_response(
    issue_number: int,
    publish_date: str,
    readiness: CommitReadiness,
) -> HistoryImportPreviewOut:
    return HistoryImportPreviewOut(
        issue_number=issue_number,
        publish_date=publish_date,
        report_entry_count=0,
        temp_detail_count=0,
        shipping_detail_count=0,
        can_commit=False,
        import_session_id="",
        errors=readiness.errors,
        readiness=readiness,
    )


def _raw_report_validation_result(raw_report) -> tuple[list[str], int, int]:
    errors: list[str] = []
    row_map = {(row.category, row.sub_category): row.value for row in raw_report.report_rows}
    temp_total = row_map.get(("social_use", "临时加印"), 0)
    temp_self = row_map.get(("social_use", "临时加印_自留"), 0)
    pending_temp = max(temp_total - temp_self, 0)
    if raw_report.source_total != raw_report.mapped_total:
        diff = abs(raw_report.source_total - raw_report.mapped_total)
        errors.append(f"原表总印数 {raw_report.source_total} 与映射后总数 {raw_report.mapped_total} 不一致，相差 {diff} 份")
    if raw_report.unmapped_items:
        errors.append(f"未命中映射项：{'；'.join(raw_report.unmapped_items)}")
    return errors, pending_temp, temp_self


def _zto_total_validation_errors(
    report_rows: list[HistoryImportRow],
    shipping_rows: list[ShippingImportRow],
) -> list[str]:
    report_total = sum(
        row.value or 0
        for row in report_rows
        if resolve_report_destination(row.category, row.sub_category, row.destination) == DESTINATION_ZTO
    )
    if report_total == 0:
        return []
    shipping_total = sum(row.quantity or 0 for row in shipping_rows)
    if report_total == shipping_total:
        return []
    diff = abs(report_total - shipping_total)
    return [
        f"中通物流份数不一致：报数合计 {report_total} 份，发货明细合计 {shipping_total} 份，相差 {diff} 份。请先核对导入文件后再提交。"
    ]


def preview_history_import(
    db: Session,
    report_bytes: bytes,
    shipping_bytes: bytes,
) -> HistoryImportPreviewOut:
    try:
        report_wb = load_workbook(io.BytesIO(report_bytes), data_only=True)
        shipping_wb = load_workbook(io.BytesIO(shipping_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise HTTPException(
            status_code=422,
            detail="无法解析上传的文件，请确保上传的是 .xlsx 格式",
        ) from exc

    uses_template_report = _is_template_report_workbook(report_wb)
    uses_template_shipping = _is_template_shipping_workbook(shipping_wb)
    uses_original_zto_shipping = is_original_zto_shipping_workbook(shipping_wb)
    raw_report = None if uses_template_report else parse_raw_report_workbook(report_wb)
    report_basic = _read_basic_info(report_wb) if uses_template_report else {
        "期号": raw_report.issue_number,
        "出版日期": raw_report.publish_date,
        "版数": raw_report.page_count,
        "备注": "",
    }
    if uses_template_shipping:
        shipping_basic = _read_basic_info(shipping_wb)
    elif uses_original_zto_shipping:
        shipping_basic = read_original_zto_shipping_basic_info(shipping_wb)
    else:
        report_issue_raw = report_basic.get("期号")
        report_issue_number = _parse_issue_number(report_issue_raw) or 0
        readiness = CommitReadiness(
            same_issue=False,
            issue_exists=False,
            can_commit=False,
            errors=["中通发货文件格式不支持，请上传系统发货明细模板或原始中通多工作表文件"],
        )
        return _error_response(
            report_issue_number,
            _normalize_date(report_basic.get("出版日期")),
            readiness,
        )

    report_issue_raw = report_basic.get("期号")
    shipping_issue_raw = shipping_basic.get("期号")
    report_issue_number = _parse_issue_number(report_issue_raw)
    shipping_issue_number = _parse_issue_number(shipping_issue_raw)
    publish_date = _normalize_date(report_basic.get("出版日期"))

    if report_issue_number is None:
        readiness = CommitReadiness(
            same_issue=False,
            issue_exists=False,
            can_commit=False,
            errors=["报数模板中的期号格式无效，请填写纯数字期号"],
        )
        return _error_response(0, publish_date, readiness)

    if shipping_issue_number is None:
        readiness = CommitReadiness(
            same_issue=False,
            issue_exists=False,
            can_commit=False,
            errors=["中通模板中的期号格式无效，请填写纯数字期号"],
        )
        return _error_response(report_issue_number, publish_date, readiness)

    if not _is_iso_date(publish_date):
        readiness = CommitReadiness(
            same_issue=True,
            issue_exists=False,
            can_commit=False,
            errors=["出版日期不能为空，且必须为 YYYY-MM-DD 格式"],
        )
        return _error_response(report_issue_number, publish_date, readiness)

    # Validate cross-issue upload
    if report_issue_number != shipping_issue_number:
        readiness = CommitReadiness(
            same_issue=False,
            issue_exists=False,
            can_commit=False,
            errors=[f"两份文件不是同一期：报数文件为 {report_issue_number} 期，发货文件为 {shipping_issue_number} 期"],
        )
        return _error_response(
            report_issue_number,
            publish_date,
            readiness,
        )

    issue_number = report_issue_number
    page_count = int(report_basic.get("版数", 24) or 24)
    notes = str(report_basic.get("备注", "") or "")

    # Block duplicate import
    existing = db.query(Issue).filter(Issue.issue_number == issue_number).first()
    if existing is not None:
        readiness = CommitReadiness(
            same_issue=True,
            issue_exists=True,
            can_commit=False,
            errors=[f"该期已存在：第 {issue_number} 期已录入系统，无法重复导入"],
        )
        return _error_response(issue_number, publish_date, readiness)

    if raw_report is None:
        report_rows = _read_report_rows(report_wb)
        temp_rows = _read_temp_rows(report_wb)
    else:
        report_rows = raw_report.report_rows
        temp_rows = []
    shipping_rows = (
        _read_shipping_rows(shipping_wb)
        if uses_template_shipping
        else read_original_zto_shipping_rows(shipping_wb)
    )

    manual_temp_print_required_quantity = 0
    manual_temp_print_self_quantity = 0
    manual_temp_rows: list[TempPrintDetailRow] = []
    if raw_report is not None:
        (
            validation_errors,
            manual_temp_print_required_quantity,
            manual_temp_print_self_quantity,
        ) = _raw_report_validation_result(raw_report)
        manual_temp_rows = raw_report.temp_print_rows
        if validation_errors:
            readiness = CommitReadiness(
                same_issue=True,
                issue_exists=False,
                can_commit=False,
                errors=validation_errors,
            )
            return _error_response(issue_number, publish_date, readiness)

    # Validate report rows against template structure and enrich with display_name
    validation_errors, enriched_report_rows = _validate_and_enrich_report_rows(db, report_rows)
    if validation_errors:
        readiness = CommitReadiness(
            same_issue=True,
            issue_exists=False,
            can_commit=False,
            errors=validation_errors,
        )
        return _error_response(issue_number, publish_date, readiness)

    total_validation_errors = _zto_total_validation_errors(enriched_report_rows, shipping_rows)
    if total_validation_errors:
        readiness = CommitReadiness(
            same_issue=True,
            issue_exists=False,
            can_commit=False,
            errors=total_validation_errors,
        )
        return _error_response(issue_number, publish_date, readiness)

    payload: dict = {
        "issue_number": issue_number,
        "publish_date": publish_date,
        "page_count": page_count,
        "notes": notes,
        "report_rows": [r.model_dump() for r in enriched_report_rows],
        "temp_rows": [r.model_dump() for r in temp_rows],
        "shipping_rows": [r.model_dump() for r in shipping_rows],
        "manual_temp_print_required_quantity": manual_temp_print_required_quantity,
        "manual_temp_print_self_quantity": manual_temp_print_self_quantity,
        "manual_temp_rows": [r.model_dump() for r in manual_temp_rows],
    }
    session_id = save_history_import_session(payload)

    # Compare against publication_schedule.page_count, warn on mismatch
    warnings: list[str] = []
    schedule_row = _find_schedule_row(db, issue_number, publish_date)
    if schedule_row is not None and schedule_row.page_count is not None and schedule_row.page_count != page_count:
        warnings.append(
            f"印数表版数为 {page_count} 版，与刊期表登记的 {schedule_row.page_count} 版不一致。"
            f"导入后将以印数表为准，自动把刊期表第 {issue_number} 期的版数更新为 {page_count} 版。"
        )

    readiness = CommitReadiness(
        same_issue=True,
        issue_exists=False,
        can_commit=True,
        errors=[],
    )
    return HistoryImportPreviewOut(
        issue_number=issue_number,
        publish_date=publish_date,
        page_count=page_count,
        report_entry_count=len(enriched_report_rows),
        temp_detail_count=len(temp_rows),
        shipping_detail_count=len(shipping_rows),
        can_commit=manual_temp_print_required_quantity == 0,
        import_session_id=session_id,
        errors=[],
        warnings=warnings,
        readiness=readiness,
        manual_temp_print_required_quantity=manual_temp_print_required_quantity,
        manual_temp_print_self_quantity=manual_temp_print_self_quantity,
        manual_temp_rows=manual_temp_rows,
    )


def _find_schedule_row(
    db: Session,
    issue_number: int,
    publish_date_str: str,
) -> PublicationSchedule | None:
    """Locate the schedule row for this import. Prefer issue_number, fall back to publish_date."""
    row = db.query(PublicationSchedule).filter(
        PublicationSchedule.issue_number == issue_number,
    ).first()
    if row is not None:
        return row
    try:
        from datetime import date as _date
        publish_date = _date.fromisoformat(publish_date_str)
    except ValueError:
        return None
    return db.query(PublicationSchedule).filter(
        PublicationSchedule.publish_date == publish_date,
    ).first()


def _validate_manual_temp_rows(
    required_quantity: int,
    rows: list[TempPrintDetailRow] | None,
) -> list[dict]:
    if required_quantity <= 0:
        return [] if rows is None else [row.model_dump() for row in rows]
    if not rows:
        raise HTTPException(
            status_code=422,
            detail=f"临时加印需要补充分配明细，合计应为 {required_quantity} 份",
        )
    total = sum(row.quantity for row in rows)
    if total != required_quantity:
        raise HTTPException(
            status_code=422,
            detail=f"临时加印明细合计 {total} 份，应等于待手工确认 {required_quantity} 份",
        )
    for row in rows:
        if row.quantity < 0 or row.self_quantity < 0:
            raise HTTPException(status_code=422, detail="临时加印明细数量不能为负数")
        if row.self_quantity > row.quantity:
            raise HTTPException(status_code=422, detail="临时加印明细自留分发数量不能大于加印数量")
    return [row.model_dump() for row in rows]


def commit_history_import(
    db: Session,
    import_session_id: str,
    manual_temp_rows: list[TempPrintDetailRow] | None = None,
) -> HistoryImportCommitOut:
    """Persist a previously previewed history import from cache to the database.

    Raises:
        HTTPException 400: session missing or expired.
        HTTPException 409: issue_number already exists in the database.
    """
    payload = pop_history_import_session(import_session_id)
    if payload is None:
        raise HTTPException(
            status_code=400,
            detail=f"导入会话不存在或已过期：{import_session_id}",
        )

    issue_number: int = payload["issue_number"]
    existing = db.query(Issue).filter(Issue.issue_number == issue_number).first()
    if existing is not None:
        raise HTTPException(
            status_code=409,
            detail=f"该期已存在：第 {issue_number} 期已录入系统，无法重复导入",
        )

    required_temp_quantity = int(payload.get("manual_temp_print_required_quantity", 0) or 0)
    if required_temp_quantity > 0 or manual_temp_rows is not None:
        payload["temp_rows"] = _validate_manual_temp_rows(required_temp_quantity, manual_temp_rows)

    from datetime import date as _date
    publish_date_str: str = payload["publish_date"]
    publish_date = _date.fromisoformat(publish_date_str)

    issue = Issue(
        issue_number=issue_number,
        publish_date=publish_date,
        page_count=payload.get("page_count", 24),
        notes=payload.get("notes", ""),
        status=IssueStatus.draft,
    )
    db.add(issue)
    db.flush()  # populate issue.id without committing

    for row in payload.get("report_rows", []):
        db.add(ReportEntry(
            issue_id=issue.id,
            category=row["category"],
            sub_category=row["sub_category"],
            destination=row.get("destination", ""),
            value=row.get("value", 0),
            is_variable=row.get("is_variable", False),
        ))

    for row in payload.get("temp_rows", []):
        db.add(TempPrintDetail(
            issue_id=issue.id,
            department=row.get("department", ""),
            custom_name=row.get("custom_name", ""),
            quantity=row.get("quantity", 0),
            self_quantity=row.get("self_quantity", 0),
        ))

    for row in payload.get("shipping_rows", []):
        db.add(ShippingDetail(
            issue_number=issue_number,
            sheet_name=row.get("sheet_name", ""),
            channel=row.get("channel", ""),
            sub_channel=row.get("sub_channel", ""),
            transport=row.get("transport", ""),
            frequency=row.get("frequency", ""),
            status=row.get("status", ""),
            name=row.get("name", ""),
            address=row.get("address", ""),
            phone=row.get("phone", ""),
            quantity=row.get("quantity", 0),
            deadline=row.get("deadline", ""),
            notes=row.get("notes", ""),
            extra_info=row.get("extra_info", ""),
            station_name=row.get("station_name", ""),
            station_hall=row.get("station_hall", ""),
            contact_person=row.get("contact_person", ""),
            seq_number=row.get("seq_number"),
            period_count=row.get("period_count"),
            confirmation=row.get("confirmation", ""),
            company=row.get("company", ""),
        ))

    db.commit()
    db.refresh(issue)

    # Sync publication_schedule.page_count to the actually-imported value
    schedule_page_count_updated = False
    previous_schedule_page_count: int | None = None
    schedule_row = _find_schedule_row(db, issue_number, publish_date_str)
    if schedule_row is not None and schedule_row.page_count != issue.page_count:
        previous_schedule_page_count = schedule_row.page_count
        schedule_row.page_count = issue.page_count
        if schedule_row.issue_number is None:
            schedule_row.issue_number = issue_number
        db.commit()
        schedule_page_count_updated = True

    return HistoryImportCommitOut(
        issue_id=issue.id,
        issue_number=issue.issue_number,
        report_entry_count=len(payload.get("report_rows", [])),
        temp_detail_count=len(payload.get("temp_rows", [])),
        shipping_detail_count=len(payload.get("shipping_rows", [])),
        schedule_page_count_updated=schedule_page_count_updated,
        previous_schedule_page_count=previous_schedule_page_count,
        new_page_count=issue.page_count,
    )
