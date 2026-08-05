"""北京报零结算单 Excel 识别。

解析以表头语义为主，不绑定固定行号。原文件始终作为附件保存；本模块只给出可由用户
复核、修改的预填结果，并保留校验告警供审计。
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PARSER_VERSION = "beijing-retail-v1"


def _text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def _decimal(value: Any) -> Decimal | None:
    if value in (None, "", "--"):
        return None
    try:
        return Decimal(str(value).replace(",", "").replace("¥", ""))
    except (InvalidOperation, ValueError):
        return None


def _metadata_value(rows: list[list[Any]], labels: tuple[str, ...]) -> str | None:
    for row in rows:
        for index, value in enumerate(row):
            normalized = _text(value).rstrip("：:")
            if not any(label in normalized for label in labels):
                continue
            for candidate in row[index + 1 :]:
                candidate_text = _text(candidate)
                if candidate_text:
                    return candidate_text
    return None


def _issue_date(value: Any, year: int) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    digits = re.sub(r"\D", "", _text(value))
    if len(digits) >= 8 and digits.startswith(str(year)):
        month_text, day_text = digits[4:6], digits[6:8]
    else:
        digits = digits.zfill(6)
        if len(digits) < 4:
            return None
        month_text, day_text = digits[:2], digits[2:4]
    try:
        return date(year, int(month_text), int(day_text))
    except ValueError:
        return None


def _find_column(headers: list[str], *candidates: str) -> int | None:
    for candidate in candidates:
        for index, header in enumerate(headers):
            if header == candidate:
                return index
    for candidate in candidates:
        for index, header in enumerate(headers):
            if candidate in header:
                return index
    return None


def _json_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, list):
        return [_json_value(item) for item in value]
    if isinstance(value, dict):
        return {key: _json_value(item) for key, item in value.items()}
    return value


def audit_result(result: dict[str, Any]) -> dict[str, Any]:
    """把识别结果转换成可写入 JSON 列的稳定快照。"""

    return _json_value(result)


def _unrecognized(filename: str, warning: str) -> dict[str, Any]:
    return {
        "recognized": False,
        "parser_version": PARSER_VERSION,
        "filename": filename,
        "return_deduction_amount": Decimal("0"),
        "detail_count": 0,
        "return_detail_count": 0,
        "warnings": [warning],
    }


def parse_settlement_excel(content: bytes, filename: str) -> dict[str, Any]:
    """识别结算 Excel，返回适合 API 预览和表单预填的结构。"""

    if Path(filename).suffix.lower() != ".xlsx":
        return _unrecognized(filename, "当前仅自动识别 XLSX；文件仍可作为原始附件归档")
    try:
        # 部分渠道文件缺少标准 worksheet dimension；只读模式会误判为 A1。
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:
        return _unrecognized(filename, f"工作簿无法读取：{exc}")

    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            header_index = None
            for index, row in enumerate(rows):
                values = [_text(value) for value in row]
                if "期数" in values and "结算数量" in values and any(
                    "结算款额" in value for value in values
                ):
                    header_index = index
                    break
            if header_index is None:
                continue

            headers = [_text(value) for value in rows[header_index]]
            issue_col = _find_column(headers, "期数")
            quantity_col = _find_column(headers, "结算数量")
            amount_col = _find_column(headers, "结算款额（含税）", "结算款额(含税)")
            unit_price_col = _find_column(headers, "结算单价")
            product_col = _find_column(headers, "产品名称")
            if issue_col is None or quantity_col is None or amount_col is None:
                continue

            year_text = _metadata_value(rows[:header_index], ("结算年度",))
            year_match = re.search(r"20\d{2}", year_text or "")
            year = int(year_match.group()) if year_match else date.today().year
            warnings: list[str] = []
            if not year_match:
                warnings.append(f"未识别结算年度，暂按 {year} 年解析期次")

            details: list[dict[str, Any]] = []
            total_quantity: Decimal | None = None
            source_total: Decimal | None = None
            for row in rows[header_index + 1 :]:
                first_cell = _text(row[0] if row else None)
                if first_cell == "合计":
                    total_quantity = _decimal(row[quantity_col] if quantity_col < len(row) else None)
                    source_total = _decimal(row[amount_col] if amount_col < len(row) else None)
                    break
                amount = _decimal(row[amount_col] if amount_col < len(row) else None)
                quantity = _decimal(row[quantity_col] if quantity_col < len(row) else None)
                if amount is None or quantity is None:
                    continue
                issue = _issue_date(row[issue_col] if issue_col < len(row) else None, year)
                details.append(
                    {
                        "issue_date": issue,
                        "quantity": quantity,
                        "amount": amount,
                        "unit_price": _decimal(
                            row[unit_price_col]
                            if unit_price_col is not None and unit_price_col < len(row)
                            else None
                        ),
                        "product": _text(
                            row[product_col]
                            if product_col is not None and product_col < len(row)
                            else None
                        )
                        or None,
                    }
                )
            if not details:
                continue

            normal = [item for item in details if item["amount"] >= 0]
            returns = [item for item in details if item["amount"] < 0]
            normal_dates = [item["issue_date"] for item in normal if item["issue_date"]]
            return_dates = [item["issue_date"] for item in returns if item["issue_date"]]
            gross = sum((item["amount"] for item in normal), Decimal("0")).quantize(Decimal("0.01"))
            deduction = abs(sum((item["amount"] for item in returns), Decimal("0"))).quantize(Decimal("0.01"))
            amount_due = (gross - deduction).quantize(Decimal("0.01"))
            if source_total is not None and source_total != amount_due:
                warnings.append(
                    f"明细净额 {amount_due:.2f} 与表格合计 {source_total:.2f} 不一致"
                )
            if any(item["issue_date"] is None for item in details):
                warnings.append("部分期次无法转换为日期，请人工核对周期")

            unit_prices = {
                abs(item["unit_price"])
                for item in details
                if item["unit_price"] not in (None, Decimal("0"))
            }
            invoice_unit_price = (
                next(iter(unit_prices)).quantize(Decimal("0.0001"))
                if len(unit_prices) == 1
                else None
            )
            if len(unit_prices) > 1:
                warnings.append("明细存在多个结算单价，未自动填写开票单价")
            products = [item["product"] for item in details if item["product"]]

            return {
                "recognized": True,
                "parser_version": PARSER_VERSION,
                "filename": filename,
                "supplier_name": _metadata_value(rows[:header_index], ("供应商名称",)),
                "external_no": _metadata_value(rows[:header_index], ("结算单号",)),
                "settlement_start_date": min(normal_dates) if normal_dates else None,
                "settlement_end_date": max(normal_dates) if normal_dates else None,
                "return_start_date": min(return_dates) if return_dates else None,
                "return_end_date": max(return_dates) if return_dates else None,
                "gross_amount": gross,
                "return_deduction_amount": deduction,
                "amount_due": amount_due,
                "invoice_item_name": products[0] if products else None,
                "invoice_quantity": (
                    abs(total_quantity).quantize(Decimal("0.01"))
                    if total_quantity is not None
                    else None
                ),
                "invoice_unit_price": invoice_unit_price,
                "invoice_amount": amount_due,
                "detail_count": len(details),
                "return_detail_count": len(returns),
                "warnings": warnings,
            }
    finally:
        workbook.close()

    return _unrecognized(filename, "未找到可识别的结算明细表头，请人工填写")
