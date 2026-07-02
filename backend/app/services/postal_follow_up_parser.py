"""回访解析：把读者明细里「按天开列」的回访列拍平成一行一条（``ParsedFollowUp``）。

读者明细里凡表头含「回访」的列（如 20240227回访 / 2025回访），每个非空单元格 → 一条回访。
"""

import io
from dataclasses import dataclass
from typing import List

import openpyxl

_READER_SIGNATURE = {"编号", "姓名", "起月日", "投递单位"}


@dataclass
class ParsedFollowUp:
    row_no: int
    year_raw: str
    external_no_raw: str
    name: str
    batch_label: str   # 列头，如 "20240227回访"
    result: str


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _find_reader_sheet(wb):
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        hmap = {_cell(v): i for i, v in enumerate(first) if _cell(v)}
        if _READER_SIGNATURE.issubset(hmap.keys()):
            follow_cols = [(i, h) for h, i in hmap.items() if "回访" in h]
            return ws, hmap, follow_cols
    return None, None, None


def is_postal_follow_up_export(file_bytes: bytes) -> bool:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return False
    ws, _, follow = _find_reader_sheet(wb)
    wb.close()
    return ws is not None and bool(follow)


def parse_postal_follow_ups(file_bytes: bytes) -> List[ParsedFollowUp]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws, hmap, follow_cols = _find_reader_sheet(wb)
    if ws is None or not follow_cols:
        wb.close()
        raise ValueError("未找到含「回访」列的邮局读者明细工作表")
    ci = {k: hmap.get(k) for k in ("编号", "姓名", "年度")}
    out: List[ParsedFollowUp] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_cell(v) for v in row):
            continue

        def g(idx):
            return _cell(row[idx]) if idx is not None and idx < len(row) else ""

        no = g(ci["编号"])
        name = g(ci["姓名"])
        year = g(ci["年度"])
        for col_idx, header in follow_cols:
            val = g(col_idx)
            if not val:
                continue
            out.append(ParsedFollowUp(
                row_no=i, year_raw=year, external_no_raw=no, name=name,
                batch_label=header, result=val,
            ))
    wb.close()
    return out
