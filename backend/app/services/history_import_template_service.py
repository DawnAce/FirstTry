import io

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models import ReportItemTemplate
from app.schemas.history_import import HistoryImportRow, TempPrintDetailRow

_BASIC_INFO_HEADERS = ["字段", "值"]
_SHIPPING_DETAIL_HEADERS = [
    "工作表名称",
    "渠道",
    "子渠道",
    "运输方式",
    "频次",
    "状态",
    "姓名",
    "地址",
    "电话",
    "数量",
    "截止日期",
    "备注",
    "附加信息",
    "城市",
    "网点名称",
    "网点大厅",
    "联系人",
    "序号",
    "期数",
    "公司",
]

_CATEGORY_LABELS = {
    "postal": "北京邮发",
    "retail": "北京报零",
    "guangzhou": "广州日报",
    "guotumao": "国图贸",
    "chengdu": "成都杂志铺",
    "binding": "合订本",
    "temp": "临时加印",
    "social_use": "营报传媒",
}


def _get_category_label(category: str, sub_category: str, display_name: str) -> str:
    if category == "other":
        if "合订本" in sub_category:
            return "合订本"
        if "国图贸" in sub_category:
            return "国图贸"
        if "杂志铺" in sub_category:
            return "成都杂志铺"
        if "上犹" in sub_category:
            return "报社订阅自投/展示"
        return display_name or sub_category or "其他"
    return _CATEGORY_LABELS.get(category, category)


def _to_bytes(workbook: Workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_report_import_template(db: Session) -> bytes:
    workbook = Workbook()
    basic_sheet = workbook.active
    basic_sheet.title = "基本信息"
    basic_sheet.append(_BASIC_INFO_HEADERS)
    basic_sheet.append(["期号", "填写历史期号"])
    basic_sheet.append(["出版日期", "填写为 YYYY-MM-DD"])
    basic_sheet.append(["版数", "可选，默认按 24 版处理"])
    basic_sheet.append(["备注", "可选"])

    report_sheet = workbook.create_sheet("报数项")
    report_sheet.append(["分类编码", "分类名称", "项目名称", "去向", "是否变动", "数值"])

    templates = (
        db.query(ReportItemTemplate)
        .order_by(ReportItemTemplate.sort_order, ReportItemTemplate.id)
        .all()
    )
    for template in templates:
        row = HistoryImportRow(
            category=template.category,
            category_name=_get_category_label(
                template.category,
                template.sub_category,
                template.display_name,
            ),
            sub_category=template.sub_category,
            destination=template.destination or "",
            is_variable=bool(template.is_variable),
            value=template.default_value or 0,
        )
        report_sheet.append(
            [
                row.category,
                row.category_name,
                row.sub_category,
                row.destination,
                "是" if row.is_variable else "否",
                row.value,
            ]
        )

    temp_sheet = workbook.create_sheet("临时加印明细")
    temp_sheet.append(["部门", "自定义名称", "数量", "自留分发数量"])
    temp_row = TempPrintDetailRow()
    temp_sheet.append(
        [
            temp_row.department,
            temp_row.custom_name,
            temp_row.quantity,
            temp_row.self_quantity,
        ]
    )

    return _to_bytes(workbook)


def build_shipping_import_template() -> bytes:
    workbook = Workbook()
    basic_sheet = workbook.active
    basic_sheet.title = "基本信息"
    basic_sheet.append(_BASIC_INFO_HEADERS)
    basic_sheet.append(["期号", "填写历史期号"])
    basic_sheet.append(["出版日期", "填写为 YYYY-MM-DD"])

    detail_sheet = workbook.create_sheet("发货明细")
    detail_sheet.append(_SHIPPING_DETAIL_HEADERS)

    return _to_bytes(workbook)
