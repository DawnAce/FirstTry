"""邮局订报数据生成模块 · 文件生成（严格对齐 7月/8月 黄金样本）。

基于批次**当前有效版本**的有效明细，生成三类产物并落盘、登记 artifacts：
1. 北京-{Y}年{M}月中国经营报订阅汇总+明细+申请.xlsx（北京-汇总 / 北京-明细 / 未到款申请）
2. 北京局订报汇总表.xlsx（代码|报刊名称|订期|省份|条数|份数|单价|款额）
3. 1-76《中国经营报》集订分送表~{Y}年{M}月起报~{地区}.xlsx（仅有订户地区；用邮局模板保留说明页/产品页）
并打包 北京{yy}年{M}月订报明细.zip（邮局汇总 + 全部地区明细）。

金额/汇总**复刻活公式**（明细 = 份数 × 完整订期单价、汇总 SUMIF）；批次未配置
完整订期单价时使用 N×20，N=13−起始月。
版式：宋体11、标题/表头/合计加粗、A4+打印区域；无签名图片（签名为文字）。
"""

import io
import os
import zipfile
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.properties import PageSetupProperties
from sqlalchemy.orm import Session

from app.models import (
    SubscriptionArtifactType,
    SubscriptionBatch,
    SubscriptionBatchStatus,
    SubscriptionGenerationRun,
    SubscriptionImportStatus,
    SubscriptionOutputArtifact,
    SubscriptionRecord,
    SubscriptionRunStatus,
)
from app.services import attachment_service
from app.services import subscription_calc_service as calc
from app.services import subscription_import_service as import_svc

BASE_FONT = "宋体"
FN = Font(name=BASE_FONT, size=11)
FB = Font(name=BASE_FONT, size=11, bold=True)
RULE_VERSION = "v1"
TEMPLATE_VERSION = "v1"
REGION_TEMPLATE = os.path.join(os.path.dirname(__file__), "..", "templates", "postal_region_template.xlsx")

# 样式常量（对齐黄金样本）。
_THIN = Side(style="thin")
BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CC = Alignment(horizontal="center", vertical="center")
LC = Alignment(horizontal="left", vertical="center")
RC = Alignment(horizontal="right", vertical="center")
VC = Alignment(vertical="center")
MC = Alignment(horizontal="center")               # 份数：水平居中、垂直默认（对齐样本）
WRAP = Alignment(vertical="center", wrap_text=True)
WRAPC = Alignment(horizontal="center", vertical="center", wrap_text=True)
NF_YEN0 = "\\¥#,##0"           # 明细/汇总 金额（¥、无小数）
NF_YEN_ACC = '"￥"#,##0.00;"￥"\\-#,##0.00'   # 明细 金额表头会计格式
NF_YEN2 = '"¥"#,##0.00'         # 汇总表 款额
NF_INT = "#,##0"               # 汇总表 单价
NF_TEXT = "@"                  # 文本（电话/编号/地址）
NF_COPIES = "0_ "             # 份数
CENTER = CC


def _cn(d: date) -> str:
    return f"{d.year}年{d.month}月{d.day}日"


def _box(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    """给矩形区域每个单元格加细边框（合并区先加边框再合并即为外框）。"""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BOX


def _cell(ws, r: int, c: int, value, *, font=FN, al=CC, nf=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = font
    cell.alignment = al
    if nf:
        cell.number_format = nf
    return cell


def _a4(ws, *, landscape: bool, print_area: str) -> None:
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = print_area


def _widths(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _region_order(records: List[SubscriptionRecord]) -> List[str]:
    """按明细首次出现顺序列出地区（与黄金样本一致）。"""
    order: List[str] = []
    for r in records:
        reg = r.region_name or "(未识别地区)"
        if reg not in order:
            order.append(reg)
    return order


# --- (a) 汇总 + 明细 + 申请 --------------------------------------------------

def _build_workbook(records: List[SubscriptionRecord], batch: SubscriptionBatch, N: int, make_date: date) -> bytes:
    count = len(records)
    total_amount = sum((r.amount or Decimal(0)) for r in records)
    last = count + 1  # 明细数据末行
    mm = f"{batch.start_month:02d}"
    regions = _region_order(records)
    complete_term_unit_price = calc.resolve_complete_term_unit_price(N, batch.unit_price)

    wb = Workbook()

    # --- 北京-明细（先建，供汇总 SUMIF 引用） ---
    det = wb.active
    det.title = "北京-明细"
    headers = ["地区", "姓名", "联系电话", "省", "市", "区", "详细地址", "邮编", "年度",
               "产品名称", "起月日", "止月日", "份数", "金额", "渠道", "汇款名称", "汇款日期"]
    for c, h in enumerate(headers, start=1):
        al = CC if c == 1 else (RC if c == 14 else LC)
        _cell(det, 1, c, h, font=FB, al=al)
    det.cell(row=1, column=13).number_format = NF_COPIES
    det.cell(row=1, column=3).number_format = NF_TEXT
    det.cell(row=1, column=14).number_format = NF_YEN_ACC
    for i, r in enumerate(records, start=2):
        _cell(det, i, 1, r.region_name or "", al=CC)
        _cell(det, i, 2, r.name, al=VC, nf=NF_TEXT)
        _cell(det, i, 3, r.phone or "", al=LC, nf=NF_TEXT)
        _cell(det, i, 4, r.province or "", al=LC)
        _cell(det, i, 5, r.city or "", al=LC)
        _cell(det, i, 6, r.district or "", al=LC)
        _cell(det, i, 7, r.address or "", al=LC)
        _cell(det, i, 8, r.postal_code or "", al=LC, nf=NF_TEXT)
        _cell(det, i, 9, f"{batch.year}年", al=LC)
        _cell(det, i, 10, "中国经营报", al=LC)
        _cell(det, i, 11, f"{mm}01", al=LC, nf=NF_TEXT)
        _cell(det, i, 12, "1231", al=RC, nf=NF_TEXT)
        _cell(det, i, 13, int(r.copies or 0), al=MC, nf=NF_COPIES)
        amount_formula = (
            f"=M{i}*{N}*20"
            if batch.unit_price is None
            else f"=M{i}*{complete_term_unit_price}"
        )
        _cell(det, i, 14, amount_formula, al=RC, nf=NF_YEN0)
        _cell(det, i, 15, r.source_channel or "", al=LC)
        _cell(det, i, 16, r.remittance_name or "", al=LC)
        _cell(det, i, 17, r.remittance_date or "", al=LC)
    _box(det, 1, last, 1, 17)
    # 合计行（份数/金额求和）+ 制表人/时间 页脚（对齐黄金样本）。
    tot = last + 1
    mtot = _cell(det, tot, 13, f"=SUM(M2:M{last})", al=MC, nf=NF_COPIES)
    ntot = _cell(det, tot, 14, f"=SUM(N2:N{last})", al=RC, nf=NF_YEN0)
    mtot.border = BOX
    ntot.border = BOX
    _cell(det, last + 4, 14, "制表人：", al=VC)
    _cell(det, last + 5, 14, "时间：", al=VC)
    dft = _cell(det, last + 5, 15, make_date, al=VC)
    dft.number_format = 'yyyy"年"m"月"d"日"'
    for rr in range(1, last + 6):
        det.row_dimensions[rr].height = 15
    _widths(det, {"A": 6, "B": 7, "C": 12, "D": 9, "E": 7, "G": 30, "H": 7.4, "I": 7,
                  "J": 9.5, "K": 6.2, "L": 6.2, "M": 4.9, "N": 10.9, "O": 15.8, "P": 17.7, "Q": 21.1, "R": 9.7})
    _a4(det, landscape=True, print_area=f"北京-明细!$A$1:$Q${max(last, 32)}")

    # --- 北京-汇总 ---
    su = wb.create_sheet("北京-汇总", 0)
    su.merge_cells("A1:D1")
    _cell(su, 1, 1, f"北京局-{batch.year}年{batch.start_month}月各地区订报汇总", font=FB, al=CC)
    su.row_dimensions[1].height = 28
    for c, h in enumerate(["序号", "地区", "金额", "数量"], start=1):
        _cell(su, 2, c, h, font=FB, al=CC)
    r = 3
    for idx, region in enumerate(regions, start=1):
        _cell(su, r, 1, idx, al=CC)
        _cell(su, r, 2, region, al=CC)
        _cell(su, r, 3, f"=SUMIF('北京-明细'!$A$2:$A${last},B{r},'北京-明细'!$N$2:$N${last})", al=CC, nf=NF_YEN0)
        _cell(su, r, 4, f"=SUMIF('北京-明细'!$A$2:$A${last},B{r},'北京-明细'!$M$2:$M${last})", al=CC)
        r += 1
    _cell(su, r, 1, "合计", font=FB, al=CC)
    _cell(su, r, 3, f"=SUM(C3:C{r-1})", font=FB, al=CC, nf=NF_YEN0)
    _cell(su, r, 4, f"=SUM(D3:D{r-1})", font=FB, al=CC)
    _box(su, 2, r, 1, 4)
    su.merge_cells(f"A{r}:B{r}")
    _cell(su, r + 2, 3, "制表人：", al=VC)
    _cell(su, r + 3, 3, "制表时间：", al=VC)
    dcell = _cell(su, r + 3, 4, make_date, al=VC)
    dcell.number_format = 'yyyy"年"m"月"d"日"'
    _widths(su, {"A": 16, "B": 24, "C": 28, "D": 24, "E": 30.7, "F": 9.8})
    _a4(su, landscape=False, print_area="北京-汇总!$A$1:$D$20")

    # --- 未到款申请（8月紧凑版；无边框、隐藏网格线、段落跨两行） ---
    ap = wb.create_sheet("未到款申请")
    ap.sheet_view.showGridLines = False
    for rng in ("A1:I1", "A2:I2", "A3:I4", "A5:I5", "A6:I6", "A7:I7", "A8:I8", "A9:I9", "F12:I12"):
        ap.merge_cells(rng)
    _cell(ap, 1, 1, "未到款提前订报的申请", font=FB, al=WRAPC)
    _cell(ap, 2, 1, "尊敬的领导：", font=FB, al=WRAP)
    _cell(ap, 3, 1,
          (f"         {batch.year}年{batch.start_month}月份《中国经营报》订阅共计{count}份。"
           f"付费读者款项正在流程中，为不影响读者按时收报，现申请先行支付订报款人民币"
           f"{total_amount:,.2f}元整。"), al=WRAP)
    _cell(ap, 5, 1, "汇款信息：", font=FB, al=WRAP)
    _cell(ap, 6, 1, "名称：中国邮政集团有限公司北京市报刊发行局", al=WRAP)
    _cell(ap, 7, 1, "账号：0200 0031 0905 4203 874", al=WRAP)
    _cell(ap, 8, 1, "开户行：中国工商银行股份有限公司北京珠市口支行", al=WRAP)
    _cell(ap, 9, 1, "妥否，请领导批示！（名单附后）", font=FB, al=WRAP)
    _cell(ap, 11, 6, "发行部：", al=CC)
    _cell(ap, 12, 6, _cn(make_date), al=CC)
    _apply_apply_row_heights(ap)
    _widths(ap, {"A": 6, "B": 9, "C": 7, "E": 14, "F": 10, "G": 19, "H": 7, "J": 9})
    _a4(ap, landscape=False, print_area="未到款申请!$A$1:$I$17")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _apply_apply_row_heights(ap) -> None:
    heights = {1: 24, 2: 24, 3: 28, 4: 28, 5: 24, 6: 24, 7: 24, 8: 24, 9: 26,
               10: 26, 11: 28, 12: 26, 13: 24, 14: 16.8, 15: 26, 16: 16.8, 17: 16.8}
    for rr, h in heights.items():
        ap.row_dimensions[rr].height = h


# --- (b) 北京局订报汇总表 ----------------------------------------------------

def _build_postal_summary(records: List[SubscriptionRecord], batch: SubscriptionBatch, N: int) -> bytes:
    mm = f"{batch.start_month:02d}"
    unit = calc.resolve_complete_term_unit_price(N, batch.unit_price)
    unit_number_format = NF_INT if batch.unit_price is None else NF_YEN2
    regions = _region_order(records)
    agg = {reg: {"count": 0, "copies": 0} for reg in regions}
    for r in records:
        reg = r.region_name or "(未识别地区)"
        agg[reg]["count"] += 1
        agg[reg]["copies"] += int(r.copies or 0)

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.row_dimensions[1].height = 24
    for c, h in enumerate(["代码", "报刊名称", "订期", "省份", "条数", "份数", "单价", "款额"], start=1):
        _cell(ws, 1, c, h, font=FB, al=CC)
    r = 2
    for region in regions:
        cnt = agg[region]["count"]
        cop = agg[region]["copies"]
        _cell(ws, r, 3, f"{mm}01-1231", al=CC)
        _cell(ws, r, 4, region, al=CC)
        _cell(ws, r, 5, cnt, al=CC)
        _cell(ws, r, 6, cop, al=CC)
        _cell(ws, r, 7, unit, al=CC, nf=unit_number_format)
        _cell(ws, r, 8, cop * unit, al=CC, nf=NF_YEN2)
        r += 1
    if regions:
        _cell(ws, 2, 1, "1-76", al=CC)
        _cell(ws, 2, 2, "中国经营报", al=CC)
    _cell(ws, r, 1, "合计", font=FB, al=CC)
    _cell(ws, r, 5, sum(a["count"] for a in agg.values()), font=FB, al=CC)
    _cell(ws, r, 6, sum(a["copies"] for a in agg.values()), font=FB, al=CC)
    _cell(ws, r, 7, None, font=FB, al=CC, nf=unit_number_format)
    _cell(ws, r, 8, sum(a["copies"] for a in agg.values()) * unit, font=FB, al=CC, nf=NF_YEN2)
    _box(ws, 1, r, 1, 8)
    if regions:
        ws.merge_cells(f"A2:A{r-1}")
        ws.merge_cells(f"B2:B{r-1}")
    ws.merge_cells(f"A{r}:D{r}")
    _widths(ws, {"A": 12, "B": 18, "C": 15, "D": 12, "E": 10, "F": 10, "G": 10, "H": 14})
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# --- (c) 地区集订分送表（邮局模板 · 数据页填充） -----------------------------

def _build_region_detail(region: str, recs: List[SubscriptionRecord], batch: SubscriptionBatch) -> bytes:
    from copy import copy
    mm = f"{batch.start_month:02d}"
    wb = openpyxl.load_workbook(REGION_TEMPLATE)
    ws = wb["数据页"]
    proto = [ws.cell(row=2, column=c)._style for c in range(1, 16)]  # 模板第2行样式原型
    for i, r in enumerate(recs, start=1):
        row = i + 1
        vals = [str(i), str(batch.year), None, "1-76", "中国经营报", f"{mm}01", "1231",
                str(int(r.copies or 0)), r.name, (r.phone or ""), r.province or "",
                r.city or "", r.district or "", r.address or "", (r.postal_code or "")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell._style = copy(proto[c - 1])   # 逐格复刻模板样式（对齐/文本格式/边框）
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# --- 编排 -------------------------------------------------------------------

def _persist_failed_run(
    db: Session,
    *,
    batch_id: int,
    version_id: int,
    started_at: datetime,
    error: Exception,
) -> None:
    """在独立事务中尽力留下失败记录，不让审计写入影响原始异常。"""
    try:
        db.rollback()
        db.add(SubscriptionGenerationRun(
            batch_id=batch_id,
            version_id=version_id,
            rule_version=RULE_VERSION,
            template_version=TEMPLATE_VERSION,
            status=SubscriptionRunStatus.failed,
            started_at=started_at,
            ended_at=datetime.now(),
            error=str(error),
        ))
        db.commit()
    except Exception:  # noqa: BLE001
        db.rollback()


def _delete_stored_files(stored_paths: List[str]) -> None:
    for stored_path in stored_paths:
        attachment_service.delete_file(stored_path)


def _generation_error(exc: Exception):
    from fastapi import HTTPException

    return HTTPException(status_code=500, detail=f"生成失败：{exc}")


def generate(db: Session, batch: SubscriptionBatch, *, operator_id: Optional[int] = None) -> SubscriptionGenerationRun:
    from fastapi import HTTPException

    if batch.active_version_id is None:
        raise HTTPException(status_code=409, detail="批次尚无当前有效版本，请先上传并设为有效")
    version = next((v for v in batch.versions if v.id == batch.active_version_id), None)
    if version is None or version.status != SubscriptionImportStatus.active:
        raise HTTPException(status_code=409, detail="当前有效版本无效，请重新设为有效")

    records = [r for r in version.records if not r.excluded]
    if not records:
        raise HTTPException(status_code=409, detail="当前有效版本无有效明细，无法生成")

    make_date = batch.make_date or date.today()
    N = import_svc.months_for(batch.start_month)
    m = batch.start_month
    batch_id = batch.id
    version_id = version.id
    started_at = datetime.now()
    category = f"subscription/{batch.year}-{m:02d}/gen"
    artifact_specs = []

    try:
        wb_bytes = _build_workbook(records, batch, N, make_date)
        wb_name = f"北京-{batch.year}年{m}月中国经营报订阅汇总+明细+申请.xlsx"
        artifact_specs.append((SubscriptionArtifactType.workbook, wb_name, wb_bytes, None))

        ps_bytes = _build_postal_summary(records, batch, N)
        artifact_specs.append((
            SubscriptionArtifactType.postal_summary,
            "北京局订报汇总表.xlsx",
            ps_bytes,
            None,
        ))

        by_region: dict = {}
        for r in records:
            by_region.setdefault(r.region_name or "(未识别地区)", []).append(r)
        region_files = []
        for region in _region_order(records):
            rd = _build_region_detail(region, by_region[region], batch)
            fname = f"1-76《中国经营报》集订分送表~{batch.year}年{m}月起报~{region}.xlsx"
            artifact_specs.append((SubscriptionArtifactType.region_detail, fname, rd, region))
            region_files.append((fname, rd))

        # ZIP 完整还原「输出」文件夹结构：顶层汇总+明细+申请，子目录「北京邮局订报数据」。
        zbio = io.BytesIO()
        with zipfile.ZipFile(zbio, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(wb_name, wb_bytes)
            zf.writestr("北京邮局订报数据/北京局订报汇总表.xlsx", ps_bytes)
            for fname, content in region_files:
                zf.writestr(f"北京邮局订报数据/{fname}", content)
        artifact_specs.append((
            SubscriptionArtifactType.zip,
            f"北京-{batch.year}年{m}月中国经营报订报数据.zip",
            zbio.getvalue(),
            None,
        ))
    except Exception as exc:  # noqa: BLE001
        _persist_failed_run(
            db,
            batch_id=batch_id,
            version_id=version_id,
            started_at=started_at,
            error=exc,
        )
        raise _generation_error(exc) from exc

    # 所有字节先构建完成，再落盘；任何一次写文件失败都会清理本轮已写文件。
    stored_artifacts = []
    try:
        for atype, filename, content, region in artifact_specs:
            stored_path = attachment_service.store_file(category, filename, content)
            stored_artifacts.append((atype, filename, content, region, stored_path))
    except Exception as exc:  # noqa: BLE001
        _delete_stored_files([item[4] for item in stored_artifacts])
        _persist_failed_run(
            db,
            batch_id=batch_id,
            version_id=version_id,
            started_at=started_at,
            error=exc,
        )
        raise _generation_error(exc) from exc

    # 文件全部就绪后才在一个数据库事务里切换「当前产物」。提交失败则旧产物状态自动回滚。
    try:
        locked_batch = (
            db.query(SubscriptionBatch)
            .filter(SubscriptionBatch.id == batch_id)
            .with_for_update()
            .one()
        )
        if locked_batch.active_version_id != version_id:
            raise RuntimeError("生成期间当前有效版本已变化，请基于新版本重新生成")
        run = SubscriptionGenerationRun(
            batch_id=batch_id,
            version_id=version_id,
            rule_version=RULE_VERSION,
            template_version=TEMPLATE_VERSION,
            status=SubscriptionRunStatus.success,
            started_at=started_at,
            ended_at=datetime.now(),
        )
        db.add(run)
        db.flush()

        db.query(SubscriptionOutputArtifact).filter(
            SubscriptionOutputArtifact.batch_id == batch_id,
            SubscriptionOutputArtifact.is_historical == False,  # noqa: E712
        ).update({"is_historical": True}, synchronize_session=False)

        for atype, filename, content, region, stored_path in stored_artifacts:
            db.add(SubscriptionOutputArtifact(
                run_id=run.id,
                batch_id=batch_id,
                version_id=version_id,
                artifact_type=atype,
                region_name=region,
                filename=filename,
                stored_path=stored_path,
                sha256=attachment_service.sha256_hex(content),
            ))
        locked_batch.status = SubscriptionBatchStatus.generated
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        _delete_stored_files([item[4] for item in stored_artifacts])
        _persist_failed_run(
            db,
            batch_id=batch_id,
            version_id=version_id,
            started_at=started_at,
            error=exc,
        )
        raise _generation_error(exc) from exc

    db.refresh(run)
    return run
