"""Parse original multi-sheet print report workbooks into history import rows."""

from dataclasses import dataclass
import datetime
import re
from typing import Any

from openpyxl.workbook.workbook import Workbook

from app.schemas.history_import import HistoryImportRow


@dataclass
class RawReportParseResult:
    issue_number: int
    publish_date: str
    page_count: int
    report_rows: list[HistoryImportRow]
    source_total: int
    mapped_total: int
    unmapped_items: list[str]


def _norm(value: Any) -> str:
    return str(value or "").replace("\n", "").replace(" ", "").strip()


def _int(value: Any) -> int:
    if value is None or value == "":
        return 0
    return int(float(value))


def _row_values(row: tuple[Any, ...]) -> list[Any]:
    return list(row)


def _parse_metadata(workbook: Workbook) -> tuple[int, str, int]:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            cells = _row_values(row)
            joined = " ".join(str(cell) for cell in cells if cell is not None)
            issue_match = re.search(r"期数[:：]?\s*(\d+)", joined)
            page_match = re.search(r"版数[:：]?\s*(\d+)", joined)
            date_match = re.search(r"出版日期[:：]?\s*(\d{4})年(\d{1,2})月(\d{1,2})日", joined)
            if issue_match and page_match and date_match:
                publish_date = datetime.date(
                    int(date_match.group(1)),
                    int(date_match.group(2)),
                    int(date_match.group(3)),
                ).isoformat()
                return int(issue_match.group(1)), publish_date, int(page_match.group(1))

            for index, cell in enumerate(cells):
                if _norm(cell) == "期数：" and index + 1 < len(cells):
                    issue_number = _int(cells[index + 1])
                    page_count = 24
                    publish_date = ""
                    for label_index, label_cell in enumerate(cells):
                        label = _norm(label_cell)
                        if label == "版数：" and label_index + 1 < len(cells):
                            page_count = _int(cells[label_index + 1])
                        if label == "出版日期：" and label_index + 1 < len(cells):
                            raw_date = cells[label_index + 1]
                            if isinstance(raw_date, datetime.datetime):
                                publish_date = raw_date.date().isoformat()
                            elif isinstance(raw_date, datetime.date):
                                publish_date = raw_date.isoformat()
                    if publish_date:
                        return issue_number, publish_date, page_count

    raise ValueError("无法识别原始印数表中的期数、版数或出版日期")


def _find_current_col(sheet, header_name: str = "本期印数") -> int:
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        for index, cell in enumerate(row):
            if _norm(cell) == header_name:
                return index
    raise ValueError(f"无法在工作表 {sheet.title} 中找到 {header_name} 列")


def _find_optional_header_col(sheet, header_name: str) -> int | None:
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        for index, cell in enumerate(row):
            if _norm(cell) == header_name:
                return index
    return None


def _set(entries: dict[tuple[str, str], int], category: str, sub_category: str, value: Any) -> None:
    entries[(category, sub_category)] = _int(value)


def _last_number(cells: list[Any], max_columns: int = 8) -> int:
    for cell in reversed(cells[:max_columns]):
        if isinstance(cell, (int, float)):
            return _int(cell)
    return 0


def _parse_print_factory(sheet, entries: dict[tuple[str, str], int]) -> int:
    source_total = 0
    for row in sheet.iter_rows(values_only=True):
        cells = _row_values(row)
        normalized = [_norm(cell) for cell in cells]
        label = normalized[0] if len(normalized) > 0 else ""
        sub_label = normalized[1] if len(normalized) > 1 else ""
        value = _last_number(cells)
        if "北京报刊局" in label and sub_label == "本市":
            _set(entries, "postal", "本市", value)
        elif "北京报刊局" in "".join(normalized) and "本市" in normalized:
            _set(entries, "postal", "本市", value)
        elif "外埠" in normalized and "北京零售公司" not in label:
            _set(entries, "postal", "外埠", value)
        elif "合订本" in label:
            _set(entries, "binding", "合订本（印厂留存）", value)
        elif label == "合计":
            source_total = value
    return source_total


def _parse_retail(sheet, entries: dict[tuple[str, str], int]) -> None:
    current_col = _find_current_col(sheet)
    print_location_col = _find_optional_header_col(sheet, "印刷地点")
    last_label = ""
    for row in sheet.iter_rows(min_row=4, values_only=True):
        cells = _row_values(row)
        label = _norm(cells[0] if len(cells) > 0 else None) or last_label
        if label:
            last_label = label
        if label == "报数合计":
            break
        value = cells[current_col] if current_col < len(cells) else None
        print_location = _norm(cells[print_location_col] if print_location_col is not None and print_location_col < len(cells) else None)
        if "北京报零" in label and print_location in {"东部", "西部"}:
            _set(entries, "retail", print_location, value)
        elif "广州日报" in label and "零售" in label:
            _set(entries, "guangzhou", "零售", value)


def _parse_subscription(sheet, entries: dict[tuple[str, str], int]) -> None:
    current_col = _find_current_col(sheet)
    for row in sheet.iter_rows(min_row=4, values_only=True):
        cells = _row_values(row)
        label = _norm(cells[0] if len(cells) > 0 else None)
        if label == "报数合计":
            break
        value = cells[current_col] if current_col < len(cells) else None
        if label == "国图贸":
            _set(entries, "guotumao", "国图贸", value)
        elif "广州日报" in label and ("订户" in label or "订阅" in label):
            _set(entries, "guangzhou", "订阅", value)
        elif label in {"杂志铺", "成都杂志铺"}:
            _set(entries, "chengdu", "成都杂志铺", value)


_SOCIAL_ROW_MAP = {
    "中经传媒智库": "中经传媒智库",
    "新闻中心": "新闻中心",
    "行政": "行政",
    "财经中心": "财经中心",
    "产经中心": "产经中心",
    "出版中心": "出版中心",
    "品牌中心": "品牌中心",
    "经营网": "经营网",
    "法务": "法务",
    "社科院、工经所": "社科院、工经所",
    "财务": "财务",
    "库房": "库房",
    "上海站用": "上海站用",
    "广东站用": "广东站用",
    "成都站用": "成都站用",
    "西安站用": "西安站用",
}


def _parse_side_value(cells: list[Any], side_label: str) -> int | None:
    normalized = [_norm(cell) for cell in cells]
    for index in range(len(normalized) - 1, 0, -1):
        label = normalized[index]
        if label == side_label and index >= 7:
            return _int(cells[index - 1])
    return None


def _parse_social(sheet, entries: dict[tuple[str, str], int]) -> None:
    current_col = _find_current_col(sheet)
    for row in sheet.iter_rows(min_row=4, values_only=True):
        cells = _row_values(row)
        label = _norm(cells[0] if len(cells) > 0 else None)
        if label == "合计":
            break
        value = cells[current_col] if current_col < len(cells) else None
        if label in _SOCIAL_ROW_MAP:
            _set(entries, "social_use", _SOCIAL_ROW_MAP[label], value)
        elif label == "临时加印":
            _set(entries, "social_use", "临时加印", value)
        if (report_value := _parse_side_value(cells, "报社")) is not None:
            _set(entries, "social_use", "营报传媒_收发室", report_value)
        if (reader_value := _parse_side_value(cells, "读者")) is not None:
            _set(entries, "social_use", "营报传媒_读者", reader_value)
        if (spare_value := _parse_side_value(cells, "备用")) is not None:
            _set(entries, "social_use", "营报传媒_备用报", spare_value)
        if (shangyou_value := _parse_side_value(cells, "上犹")) is not None:
            _set(entries, "social_use", "营报传媒_上犹", shangyou_value)
        if (rail_value := _parse_side_value(cells, "高铁展示")) is not None:
            _set(entries, "social_use", "高铁展示", rail_value)


def _parse_distribution(sheet, entries: dict[tuple[str, str], int]) -> None:
    current_col = _find_current_col(sheet)
    for row in sheet.iter_rows(min_row=5, values_only=True):
        cells = _row_values(row)
        label = _norm(cells[0] if len(cells) > 0 else None)
        if label == "合计":
            break
        value = cells[current_col] if current_col < len(cells) else None
        if label == "临时加印（报社内存）":
            _set(entries, "social_use", "临时加印_自留", value)


def parse_raw_report_workbook(workbook: Workbook) -> RawReportParseResult:
    issue_number, publish_date, page_count = _parse_metadata(workbook)
    entries: dict[tuple[str, str], int] = {}
    source_total = 0

    if "北京印厂" in workbook.sheetnames:
        source_total = _parse_print_factory(workbook["北京印厂"], entries)
    if "零售渠道`" in workbook.sheetnames:
        _parse_retail(workbook["零售渠道`"], entries)
    if "订阅渠道`" in workbook.sheetnames:
        _parse_subscription(workbook["订阅渠道`"], entries)
    if "社用报`" in workbook.sheetnames:
        _parse_social(workbook["社用报`"], entries)
    if "收发室自留分发（需打印）" in workbook.sheetnames:
        _parse_distribution(workbook["收发室自留分发（需打印）"], entries)

    entries.setdefault(("social_use", "临时加印"), 0)
    entries.setdefault(("social_use", "临时加印_自留"), 0)

    report_rows = [
        HistoryImportRow(
            category=category,
            display_name="",
            sub_category=sub_category,
            destination="",
            is_variable=True,
            value=value,
        )
        for (category, sub_category), value in entries.items()
    ]
    mapped_total = sum(row.value for row in report_rows)

    return RawReportParseResult(
        issue_number=issue_number,
        publish_date=publish_date,
        page_count=page_count,
        report_rows=report_rows,
        source_total=source_total,
        mapped_total=mapped_total,
        unmapped_items=[],
    )
