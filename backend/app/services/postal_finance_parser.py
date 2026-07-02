"""提现发票合集 Excel 解析（纯提取 → ``ParsedFinance``）。

表头签名：含 发票信息 / 发票类型 / 到款金额。「订单号」列（将来补）按表头含「订单号/单号」自动识别。
发票抬头/税号 从「发票信息」文本里解析（放在 import_service，parser 保留原文）。
"""

import io
from dataclasses import dataclass
from typing import List

import openpyxl

_SIGNATURE = {"发票信息", "发票类型", "到款金额"}

_FIELDS = {
    "姓名": "payer_name",
    "商品名称": "product",
    "份数": "copies_raw",
    "金额": "amount_raw",
    "手续费": "fee_raw",
    "到款金额": "net_raw",
    "到款日期": "collected_raw",
    "开票金额": "invoiced_raw",
    "发票信息": "invoice_info",
    "发票接收手机/邮箱": "recipient",
    "发票类型": "tax_category",
    "订单平台": "platform",
}


@dataclass
class ParsedFinance:
    row_no: int
    external_no_raw: str = ""  # 原始平台订单号（将来补）
    payer_name: str = ""
    product: str = ""
    copies_raw: str = ""
    amount_raw: str = ""
    fee_raw: str = ""
    net_raw: str = ""
    collected_raw: str = ""
    invoiced_raw: str = ""
    invoice_info: str = ""
    recipient: str = ""
    tax_category: str = ""
    platform: str = ""


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _find_sheet(wb):
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        hmap = {_cell(v): i for i, v in enumerate(first) if _cell(v)}
        if _SIGNATURE.issubset(hmap.keys()):
            return ws, hmap
    return None, None


def is_postal_finance_export(file_bytes: bytes) -> bool:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return False
    ws, _ = _find_sheet(wb)
    wb.close()
    return ws is not None


def parse_postal_finance(file_bytes: bytes) -> List[ParsedFinance]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws, hmap = _find_sheet(wb)
    if ws is None:
        wb.close()
        raise ValueError("无法识别的提现发票表：未找到含「发票信息/发票类型/到款金额」表头的工作表")
    col = {field: hmap[name] for name, field in _FIELDS.items() if name in hmap}
    # 订单号列（将来补）：表头含「订单号/单号」
    order_no_idx = next((i for h, i in hmap.items() if "订单号" in h or "单号" in h), None)

    out: List[ParsedFinance] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_cell(v) for v in row):
            continue

        def get(idx):
            return _cell(row[idx]) if idx is not None and idx < len(row) else ""

        pf = ParsedFinance(row_no=i)
        for field in _FIELDS.values():
            pf.__setattr__(field, get(col.get(field)))
        pf.external_no_raw = get(order_no_idx)
        if not pf.payer_name and not pf.invoice_info:
            continue
        out.append(pf)
    wb.close()
    return out
