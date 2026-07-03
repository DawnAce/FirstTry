"""邮局年投诉 Excel 解析（纯提取 → ``ParsedComplaint``）。

只按表头名取值；业务映射（挂订单 / 处理情况归一 / 状态派生）放在 postal_complaint_import_service。
表头签名：同时含 投诉情况 / 处理情况 / 编号 / 接诉日期 的 sheet 即投诉表。
"""

import io
from dataclasses import dataclass
from typing import List

import openpyxl

_SIGNATURE = {"投诉情况", "处理情况", "编号", "接诉日期"}

_FIELDS = {
    "接诉日期": "complaint_date_raw",
    "姓名": "name",
    "联系电话": "phone",
    "省": "province",
    "市": "city",
    "区": "district",
    "详细地址": "detail_address",
    "邮编": "postal_code",
    "年度": "year_raw",
    "投诉情况": "missing_issues",
    "处理情况": "handling",
    "回访": "follow_up",
    "处理次数": "handling_count_raw",
    "编号": "external_no_raw",
    "第一接诉人": "first_handler",
    "投递渠道单位": "distribution_unit_name",
    "备注": "notes",
}


@dataclass
class ParsedComplaint:
    row_no: int
    complaint_date_raw: str = ""
    name: str = ""
    phone: str = ""
    province: str = ""
    city: str = ""
    district: str = ""
    detail_address: str = ""
    postal_code: str = ""
    year_raw: str = ""
    missing_issues: str = ""
    handling: str = ""
    follow_up: str = ""
    handling_count_raw: str = ""
    external_no_raw: str = ""
    first_handler: str = ""
    distribution_unit_name: str = ""
    notes: str = ""


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _find_sheet(wb):
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        hmap = {_cell(v): i for i, v in enumerate(first) if _cell(v)}
        if _SIGNATURE.issubset(hmap.keys()):
            return ws, hmap
    return None, None


def is_postal_complaint_export(file_bytes: bytes) -> bool:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return False
    ws, _ = _find_sheet(wb)
    wb.close()
    return ws is not None


def parse_postal_complaints(file_bytes: bytes) -> List[ParsedComplaint]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws, hmap = _find_sheet(wb)
    if ws is None:
        wb.close()
        raise ValueError(
            "无法识别的邮局投诉表：未找到含「投诉情况/处理情况/编号/接诉日期」表头的工作表"
        )
    col = {field: hmap[name] for name, field in _FIELDS.items() if name in hmap}
    out: List[ParsedComplaint] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_cell(v) for v in row):
            continue

        def get(field: str) -> str:
            idx = col.get(field)
            if idx is None or idx >= len(row):
                return ""
            return _cell(row[idx])

        pc = ParsedComplaint(row_no=i)
        for field in _FIELDS.values():
            setattr(pc, field, get(field))
        # 无编号且无投诉情况 → 无效行
        if not pc.external_no_raw and not pc.missing_issues:
            continue
        out.append(pc)
    wb.close()
    return out
