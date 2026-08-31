"""运单导入、逐包裹核销和期级待补统计。"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
import re
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, insert, or_, select
from sqlalchemy.orm import Session, joinedload, noload

from app.models import (
    Issue,
    IssueAuditSnapshot,
    PublicationSchedule,
    ShippingDetail,
    ShippingDetailSourceType,
    ShippingDeferral,
    ShippingFulfillmentAdjustment,
    ShippingPackage,
    ShippingPackageAllocation,
    ShippingWaybillImportBatch,
    ShippingWaybillImportDocument,
    ShippingWaybillImportRow,
    WaybillImportStatus,
    WaybillMatchStatus,
)
from app.models.user import User
from app.schemas.shipping_waybill import (
    FulfillmentAdjustmentAttributionIn,
    FulfillmentAdjustmentIn,
    FulfillmentSummaryOut,
    ConsolidatedPackageIn,
    ConsolidatedPackageOut,
    ShippingDeferralBulkIn,
    ShippingGapDetailOut,
    ShippingPlanTransferIn,
    ShippingPlanTransferOut,
    WaybillBulkMatchIn,
    WaybillImportRowCreate,
    WaybillImportRowUpdate,
)
from app.services.operation_log_service import record_operation
from app.services.original_zto_shipping_import_service import is_original_zto_shipping_workbook
from app.services.shipping_suspension_service import (
    MANUAL_ADJUSTMENT_SOURCE,
    PLAN_STATUS_ADJUSTMENT_SOURCE,
    sync_plan_status_adjustment,
)


WAREHOUSE_STOCK_IN = "warehouse_stock_in"
WAREHOUSE_STOCK_IN_REASON = "转库留存 · 当期报纸入马飞中通库房备货"
WAREHOUSE_STOCK_IN_IMPORT_REASON = "已转换：马飞—库房留存改按转库留存/库存入库核销"
TWICE_MONTHLY_CONSOLIDATION = "twice_monthly_consolidation"
MONTH_END_CONSOLIDATION = "month_end_consolidation"


@dataclass
class ParsedWaybillRow:
    source_sheet: str
    source_row: int
    carrier: str
    tracking_no: str | None
    recipient_name: str
    phone: str
    address: str
    quantity: int
    no_tracking_required: bool = False
    source_detail_id: int | None = None
    raw_values: list[str] | None = None
    parse_reason: str | None = None


@dataclass
class ParsedImportDocument:
    source_sheet: str
    document_type: str
    associated_name: str
    extracted_data: dict[str, Any]


@dataclass
class ParsedWaybillWorkbook:
    rows: list[ParsedWaybillRow]
    documents: list[ParsedImportDocument]


@dataclass
class OrphanedWaybillRestoreResult:
    restored_rows: int = 0
    restored_quantity: int = 0
    unresolved_rows: int = 0
    restored_adjustments: int = 0
    restored_deferrals: int = 0


@dataclass
class WaybillPreviewRepairResult:
    issue_number: int
    batch_id: int
    repaired_rows: int = 0
    repaired_quantity: int = 0
    row_ids: list[int] = field(default_factory=list)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _quantity(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _tracking(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value)).upper()


def _looks_like_tracking(value: Any) -> bool:
    value = _tracking(value)
    return bool(re.fullmatch(r"(?:[A-Z]{1,4})?\d{10,24}", value))


def _looks_like_registered_tracking(value: Any) -> bool:
    return bool(re.fullmatch(r"SC\d{10,24}", _tracking(value)))


def _carrier_for_tracking(tracking_no: str, fallback: str = "中通") -> str:
    if tracking_no.startswith("SF"):
        return "顺丰"
    if tracking_no.startswith("SC"):
        return "邮政挂号"
    return fallback


def _raw_values(row: tuple[Any, ...] | list[Any]) -> list[str]:
    values = [_text(value) for value in row]
    while values and not values[-1]:
        values.pop()
    return values


def _is_meaningful_candidate(row: tuple[Any, ...] | list[Any]) -> bool:
    values = [value for value in _raw_values(row) if value]
    if len(values) < 2:
        return False
    total_labels = {"合计", "总计", "小计", "共计"}
    if any(value.replace(" ", "") in total_labels for value in values):
        return False
    header_words = {"姓名", "收件人", "电话", "手机号", "地址", "份数", "数量", "运单号", "快递单号"}
    return len(set(values) & header_words) < 2


def _row_value(row: tuple[Any, ...] | list[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def _parse_postal_30_fields(
    row: tuple[Any, ...] | list[Any],
) -> tuple[str, str, str, int, str]:
    """Read both historical layouts of the postal-30 actual-waybill sheet.

    The compact layout used by the 2026-01-05 workbook has address/name/quantity
    in F/G/H. Later workbooks insert a blank F column and use G/H/I. The
    quantity cell identifies the layout without changing the separate original
    shipping-plan parser for the ``上犹`` worksheet.
    """
    compact_quantity = _quantity(_row_value(row, 7))
    spaced_quantity = _quantity(_row_value(row, 8))
    phone = _text(_row_value(row, 4))
    if compact_quantity > 0 and spaced_quantity <= 0:
        return (
            _text(_row_value(row, 6)),
            phone,
            _text(_row_value(row, 5)),
            compact_quantity,
            "compact_fgh",
        )
    return (
        _text(_row_value(row, 7)),
        phone,
        _text(_row_value(row, 6)),
        spaced_quantity,
        "spaced_ghi",
    )


def _parse_high_speed_rail_fields(
    row: tuple[Any, ...] | list[Any],
) -> tuple[str, str, str, int, str]:
    """Read both layouts used by high-speed-rail actual-waybill sheets.

    Some carrier workbooks use G/H for recipient/quantity.  Another variant
    inserts a display-name column at G and uses H/I for recipient/quantity.
    The positive quantity cell identifies the variant without treating a
    station/display label as the recipient.
    """
    compact_quantity = _quantity(_row_value(row, 7))
    display_name_quantity = _quantity(_row_value(row, 8))
    phone = _text(_row_value(row, 4))
    address = _text(_row_value(row, 5))
    if display_name_quantity > 0:
        return (
            _text(_row_value(row, 7)),
            phone,
            address,
            display_name_quantity,
            "display_name_ghi",
        )
    return (
        _text(_row_value(row, 6)),
        phone,
        address,
        compact_quantity,
        "compact_gh",
    )


def _parse_standard_sheet(ws) -> list[ParsedWaybillRow] | None:
    aliases = {
        "detail_id": {"发货明细id", "明细id", "shippingdetailid"},
        "carrier": {"快递公司", "承运公司", "物流公司"},
        "tracking": {"运单号", "快递单号", "物流单号"},
        "quantity": {"实发份数", "份数", "数量"},
        "name": {"姓名", "收件人", "收件人姓名"},
        "phone": {"电话", "手机号", "收件人电话"},
        "address": {"地址", "收件地址", "收件人地址"},
        "no_tracking": {"无需运单", "无需快递", "无需发货"},
    }
    for header_row in range(1, min(ws.max_row, 10) + 1):
        values = [_text(cell.value).lower().replace(" ", "") for cell in ws[header_row]]
        mapping: dict[str, int] = {}
        for key, names in aliases.items():
            for index, value in enumerate(values):
                if value in names:
                    mapping[key] = index
                    break
        if not {"quantity", "name"}.issubset(mapping) or not ({"tracking", "no_tracking"} & mapping.keys()):
            continue
        parsed: list[ParsedWaybillRow] = []
        for row_number, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            raw_values = _raw_values(row)
            name = _text(row[mapping["name"]])
            qty = _quantity(row[mapping["quantity"]])
            tracking = _tracking(row[mapping["tracking"]]) if "tracking" in mapping else ""
            no_tracking_value = _text(row[mapping["no_tracking"]]).lower() if "no_tracking" in mapping else ""
            no_tracking = no_tracking_value in {"是", "true", "1", "无需运单", "无需发货"}
            if not name and not tracking and qty == 0:
                if not _is_meaningful_candidate(row):
                    continue
                parsed.append(ParsedWaybillRow(
                    source_sheet=ws.title,
                    source_row=row_number,
                    carrier="中通",
                    tracking_no=None,
                    recipient_name="",
                    phone="",
                    address="",
                    quantity=0,
                    raw_values=raw_values,
                    parse_reason="未能按当前工作表格式识别，请人工补充",
                ))
                continue
            detail_id = _quantity(row[mapping["detail_id"]]) if "detail_id" in mapping else 0
            carrier = _text(row[mapping["carrier"]]) if "carrier" in mapping else ""
            parsed.append(ParsedWaybillRow(
                source_sheet=ws.title,
                source_row=row_number,
                carrier=carrier or ("无需运单" if no_tracking else _carrier_for_tracking(tracking)),
                tracking_no=tracking or None,
                recipient_name=name,
                phone=_text(row[mapping["phone"]]) if "phone" in mapping else "",
                address=_text(row[mapping["address"]]) if "address" in mapping else "",
                quantity=qty,
                no_tracking_required=no_tracking,
                source_detail_id=detail_id or None,
                raw_values=raw_values,
            ))
        return parsed
    return None


def _parse_known_sheet(ws) -> list[ParsedWaybillRow]:
    rows: list[ParsedWaybillRow] = []
    title = ws.title
    if "样报缴送清单" in title:
        return rows
    if "备用" in title or "社用" in title:
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = tuple(row)
            name, address, phone, qty = (_text(row[i]) if i < len(row) else "" for i in range(4))
            quantity = _quantity(qty)
            if name == "合计" or not _is_meaningful_candidate(row):
                continue
            rows.append(ParsedWaybillRow(
                source_sheet=title, source_row=row_number, carrier="无需运单", tracking_no=None,
                recipient_name=name, phone=phone, address=address, quantity=quantity,
                no_tracking_required=True,
                raw_values=_raw_values(row),
                parse_reason=None if name and quantity > 0 else "未能按当前工作表格式识别，请人工补充",
            ))
        return rows

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = tuple(row)
        tracking_c = _tracking(row[2]) if len(row) > 2 and _looks_like_tracking(row[2]) else ""
        tracking_d = (
            _tracking(row[3])
            if len(row) > 3 and _looks_like_registered_tracking(row[3])
            else ""
        )
        tracking_e = _tracking(row[4]) if len(row) > 4 and _looks_like_tracking(row[4]) else ""
        registered_name = _text(_row_value(row, 6))
        registered_address = _text(_row_value(row, 5))
        registered_quantity = _quantity(_row_value(row, 9))
        if (
            ("月底" in title or "整月" in title or "挂号" in title)
            and tracking_d
            and registered_name
            and registered_address
            and registered_quantity > 0
        ):
            name = registered_name
            phone = _text(_row_value(row, 7))
            address = registered_address
            qty = registered_quantity
            carrier, tracking = "邮政挂号", tracking_d
        elif "邮政30" in title and tracking_c:
            name, phone, address, qty, _layout = _parse_postal_30_fields(row)
            carrier, tracking = "邮政", tracking_c
        elif "高铁" in title and tracking_c:
            name, phone, address, qty, _layout = _parse_high_speed_rail_fields(row)
            carrier, tracking = _carrier_for_tracking(tracking_c), tracking_c
        elif ("挂号" in title or "整月" in title) and tracking_c:
            name, phone, address, qty = _text(row[6] or row[3]), _text(row[4]), _text(row[5]), _quantity(row[9])
            carrier, tracking = _carrier_for_tracking(tracking_c), tracking_c
        elif ("挂号" in title or "整月" in title) and tracking_e:
            name, phone, address, qty = _text(row[6]), _text(row[7]), _text(row[5]), _quantity(row[9])
            carrier, tracking = _carrier_for_tracking(tracking_e, "邮政挂号"), tracking_e
        elif tracking_c:
            name, phone, address, qty = _text(row[6]), _text(row[4]), _text(row[5]), _quantity(row[7])
            carrier, tracking = _carrier_for_tracking(tracking_c), tracking_c
        else:
            if _is_meaningful_candidate(row):
                rows.append(ParsedWaybillRow(
                    source_sheet=title,
                    source_row=row_number,
                    carrier="中通",
                    tracking_no=None,
                    recipient_name="",
                    phone="",
                    address="",
                    quantity=0,
                    raw_values=_raw_values(row),
                    parse_reason="未能按当前工作表格式识别，请人工补充",
                ))
            continue
        rows.append(ParsedWaybillRow(
            source_sheet=title, source_row=row_number, carrier=carrier, tracking_no=tracking,
            recipient_name=name, phone=phone, address=address, quantity=qty,
            raw_values=_raw_values(row),
            parse_reason=None if name and qty > 0 else "未能按当前工作表格式识别，请人工补充",
        ))
    return rows


_ORIGINAL_ZTO_DETAIL_LAYOUTS = {
    "每周（对公）": {"min_row": 3, "max_col": 13, "quantity": 3, "transport": 9},
    "每周（读者）": {"min_row": 3, "max_col": 14, "quantity": 3, "transport": 10},
    "高铁展示": {
        "min_row": 4, "max_col": 15, "name": 4, "phone": 5,
        "address": 6, "quantity": 7, "transport": 14,
    },
    "北京悦途出行（高铁）": {
        "min_row": 4, "max_col": 15, "name": 4, "phone": 5,
        "address": 6, "quantity": 7, "transport": 14,
    },
    "上犹": {"min_row": 3, "max_col": 13, "quantity": 3, "transport": 10},
    "停发-双周（读者）": {"min_row": 3, "max_col": 8, "quantity": 4},
    "月底-整月": {"min_row": 3, "max_col": 15, "quantity": 4, "transport": 11},
}


def _carrier_for_transport(value: Any) -> str:
    transport = _text(value)
    if "邮政" in transport:
        return "邮政"
    if "顺丰" in transport:
        return "顺丰"
    return "中通"


def _parse_original_zto_detail_workbook(workbook) -> list[ParsedWaybillRow]:
    """Read recipient rows from the original shipping-plan workbook.

    This workbook intentionally has no tracking-number column.  Keeping the
    recipient fields lets the preview explain that only the tracking number is
    missing instead of replacing the row with an unrecognized placeholder.
    """
    parsed: list[ParsedWaybillRow] = []
    for ws in workbook.worksheets:
        layout = _ORIGINAL_ZTO_DETAIL_LAYOUTS.get(ws.title)
        if layout is None:
            continue
        for row_number, row in enumerate(
            ws.iter_rows(
                min_row=layout["min_row"],
                max_col=layout["max_col"],
                values_only=True,
            ),
            start=layout["min_row"],
        ):
            name = _text(row[layout.get("name", 0)])
            address = _text(row[layout.get("address", 1)])
            phone = _text(row[layout.get("phone", 2)])
            quantity = _quantity(row[layout["quantity"]])
            if phone == "合计" or (not name and not address and not phone and quantity <= 0):
                continue
            transport_index = layout.get("transport")
            transport = row[transport_index] if transport_index is not None else "中通"
            parsed.append(ParsedWaybillRow(
                source_sheet=ws.title,
                source_row=row_number,
                carrier=_carrier_for_transport(transport),
                tracking_no=None,
                recipient_name=name,
                phone=phone,
                address=address,
                quantity=quantity,
                raw_values=_raw_values(row),
            ))
    return parsed


def _value_after_label(value: Any) -> str:
    text = _text(value)
    parts = re.split(r"[:：]", text, maxsplit=1)
    return parts[1].strip() if len(parts) == 2 else text


def _formatted_number(value: Any, suffix: str) -> int | None:
    match = re.search(rf"(\d+)\s*{re.escape(suffix)}", _text(value))
    if match:
        return int(match.group(1))
    number = _quantity(value)
    return number if number > 0 else None


def _parse_sample_submission_list(ws) -> ParsedImportDocument:
    associated_match = re.search(r"关联\s*(.+)$", ws.title)
    associated_name = associated_match.group(1).strip() if associated_match else ""
    extracted: dict[str, Any] = {"associated_name": associated_name}
    header: dict[str, int] = {}
    aliases = {
        "publication_name": {"报纸题名", "报刊名称", "报纸名称"},
        "cn_number": {"cn号", "国内统一连续出版物号"},
        "year": {"年份", "年度"},
        "month": {"期次/月", "期次/月份", "月份"},
        "quantity": {"散报总份数", "缴送份数", "份数"},
        "unit_price": {"每份定价/元", "定价/元", "定价"},
    }
    rows = list(ws.iter_rows(values_only=True))
    for row in rows:
        for value in row:
            text = _text(value)
            if text.startswith("接收单位"):
                extracted["receiving_unit"] = _value_after_label(text)
            elif text.startswith("缴送单位"):
                extracted["sender"] = _value_after_label(text)
            sequence_match = re.search(r"(\d{4})\s*年第\s*(\d+)\s*次寄送", text)
            if sequence_match:
                extracted["submission_year"] = int(sequence_match.group(1))
                extracted["submission_sequence"] = int(sequence_match.group(2))
        normalized = [_text(value).lower().replace(" ", "") for value in row]
        candidate: dict[str, int] = {}
        for key, names in aliases.items():
            for index, value in enumerate(normalized):
                if value in names:
                    candidate[key] = index
                    break
        if {"publication_name", "cn_number", "year", "month", "quantity"}.issubset(candidate):
            header = candidate
            continue
        if not header:
            continue
        publication = _text(_row_value(row, header["publication_name"]))
        if not publication:
            continue
        extracted["publication_name"] = publication
        extracted["cn_number"] = _text(_row_value(row, header["cn_number"]))
        extracted["year"] = _quantity(_row_value(row, header["year"])) or None
        month_value = _text(_row_value(row, header["month"]))
        month_match = re.search(r"(\d{1,2})", month_value)
        extracted["month"] = int(month_match.group(1)) if month_match else None
        extracted["quantity"] = _formatted_number(
            _row_value(row, header["quantity"]), "份"
        )
        if "unit_price" in header:
            extracted["unit_price"] = _formatted_number(
                _row_value(row, header["unit_price"]), "元"
            )
        break
    return ParsedImportDocument(
        source_sheet=ws.title,
        document_type="sample_submission_list",
        associated_name=associated_name,
        extracted_data=extracted,
    )


def _parse_waybill_workbook_payload(content: bytes) -> ParsedWaybillWorkbook:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="无法读取运单Excel文件") from exc
    documents = [
        _parse_sample_submission_list(ws)
        for ws in workbook.worksheets
        if "样报缴送清单" in ws.title
    ]
    if is_original_zto_shipping_workbook(workbook):
        original_rows = _parse_original_zto_detail_workbook(workbook)
        if original_rows:
            return ParsedWaybillWorkbook(rows=original_rows, documents=documents)
    parsed: list[ParsedWaybillRow] = []
    for ws in workbook.worksheets:
        if "样报缴送清单" in ws.title:
            continue
        standard = _parse_standard_sheet(ws)
        parsed.extend(standard if standard is not None else _parse_known_sheet(ws))
    if not parsed:
        raise HTTPException(status_code=400, detail="未在文件中识别到有效运单或无需运单记录")
    return ParsedWaybillWorkbook(rows=parsed, documents=documents)


def parse_waybill_workbook(content: bytes) -> list[ParsedWaybillRow]:
    return _parse_waybill_workbook_payload(content).rows


def _normalized(value: str) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]", "", (value or "").lower())


def _phone(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _match_key(name: str, phone: str, address: str) -> tuple[str, str, str]:
    return _normalized(name), _phone(phone), _normalized(address)


def _contact_key_matches(
    left: tuple[str, str, str],
    right: tuple[str, str, str],
) -> bool:
    return (
        left == right
        or (left[1] and left[2] and left[1:] == right[1:])
        or (left[0] and left[2] and (left[0], left[2]) == (right[0], right[2]))
    )


def _expected_quantity(db: Session, issue: Issue) -> int:
    snapshot = (
        db.query(IssueAuditSnapshot)
        .filter(IssueAuditSnapshot.issue_id == issue.id, IssueAuditSnapshot.snapshot_type == "confirm")
        .order_by(IssueAuditSnapshot.created_at.desc(), IssueAuditSnapshot.id.desc())
        .first()
    )
    if snapshot:
        return snapshot.report_total
    return int(db.query(func.coalesce(func.sum(ShippingDetail.quantity), 0)).filter(
        ShippingDetail.issue_number == issue.issue_number,
        ShippingDetail.source_type != ShippingDetailSourceType.complaint_makeup,
    ).scalar() or 0)


def _details_for_issue(db: Session, issue_number: int) -> list[ShippingDetail]:
    return db.query(ShippingDetail).filter(
        ShippingDetail.issue_number == issue_number,
        ShippingDetail.source_type != ShippingDetailSourceType.complaint_makeup,
    ).order_by(ShippingDetail.id).all()


def _is_last_issue_of_month(db: Session, issue: Issue) -> bool:
    month_start, next_month = _month_bounds(issue.publish_date)
    last_publish_date = db.query(func.max(PublicationSchedule.publish_date)).filter(
        PublicationSchedule.publish_date >= month_start,
        PublicationSchedule.publish_date < next_month,
        PublicationSchedule.is_suspended.is_(False),
    ).scalar()
    return last_publish_date == issue.publish_date


def _deferral_targets_issue(deferral: ShippingDeferral, issue: Issue) -> bool:
    has_target = (
        deferral.target_issue_number is not None
        or deferral.target_publish_date is not None
    )
    return bool(
        has_target
        and (
            deferral.target_issue_number is None
            or deferral.target_issue_number == issue.issue_number
        )
        and (
            deferral.target_publish_date is None
            or deferral.target_publish_date == issue.publish_date
        )
    )


def _matching_month_end_deferrals(
    db: Session,
    *,
    issue: Issue,
    row: ParsedWaybillRow,
    shipping_detail_id: int,
) -> tuple[list[ShippingDeferral], str | None]:
    if not (
        row.tracking_no
        and ("月底" in row.source_sheet or "整月" in row.source_sheet)
        and _is_last_issue_of_month(db, issue)
    ):
        return [], None
    candidates = db.query(ShippingDeferral).filter(
        ShippingDeferral.status == "pending",
        ShippingDeferral.deferral_type == MONTH_END_CONSOLIDATION,
        or_(
            ShippingDeferral.target_issue_number == issue.issue_number,
            ShippingDeferral.target_publish_date == issue.publish_date,
        ),
    ).order_by(ShippingDeferral.issue_number, ShippingDeferral.id).all()
    candidates = [item for item in candidates if _deferral_targets_issue(item, issue)]
    row_name, row_phone, row_address = _match_key(row.recipient_name, row.phone, row.address)
    matched = []
    for deferral in candidates:
        name, phone, address = _match_key(
            deferral.detail_name_snapshot or "",
            deferral.detail_phone_snapshot or "",
            deferral.detail_address_snapshot or "",
        )
        if _contact_key_matches(
            (name, phone, address),
            (row_name, row_phone, row_address),
        ):
            matched.append(deferral)
    if not matched:
        return [], None
    batches = {item.consolidation_batch for item in matched}
    if None in batches or len(batches) != 1:
        return [], "匹配到的月底待合寄记录批次不唯一，请先人工核对"
    if any(item.shipping_detail_id is None or item.quantity <= 0 for item in matched):
        return [], "匹配到的月底待合寄记录缺少明细归属或份数无效"
    current = [
        item for item in matched
        if item.issue_number == issue.issue_number
        or item.shipping_detail_id == shipping_detail_id
    ]
    if len(current) > 1:
        return [], "本期匹配到多条月底待合寄记录，请先人工核对"
    if current and (
        current[0].shipping_detail_id != shipping_detail_id
        or current[0].quantity != row.quantity
    ):
        return [], "本期月底待合寄记录与运单行份数或明细不一致"
    return matched, None


def _validate_sample_submission_document(
    *,
    db: Session,
    issue: Issue,
    document: ParsedImportDocument,
    row: ShippingWaybillImportRow,
) -> tuple[str, dict[str, Any], list[str]]:
    data = dict(document.extracted_data)
    expected_quantity = row.consolidation_quantity or row.quantity
    covered_issues = row.consolidation_issue_numbers or [issue.issue_number]
    month_start, next_month = _month_bounds(issue.publish_date)
    scheduled_issues = [
        issue_number for (issue_number,) in db.query(PublicationSchedule.issue_number).filter(
            PublicationSchedule.publish_date >= month_start,
            PublicationSchedule.publish_date < next_month,
            PublicationSchedule.is_suspended.is_(False),
            PublicationSchedule.issue_number.isnot(None),
        ).order_by(PublicationSchedule.publish_date).all()
    ]
    data["expected_quantity"] = expected_quantity
    data["covered_issue_numbers"] = covered_issues
    data["scheduled_issue_numbers"] = scheduled_issues
    errors: list[str] = []
    if _normalized(document.associated_name) != _normalized(row.recipient_name):
        errors.append("清单关联姓名与运单收件人不一致")
    if data.get("receiving_unit") != "中国版本图书馆":
        errors.append("接收单位不是中国版本图书馆")
    if _normalized(data.get("sender") or "") != _normalized("《中国经营报》社"):
        errors.append("缴送单位不是《中国经营报》社")
    if _normalized(data.get("publication_name") or "") != _normalized("中国经营报"):
        errors.append("报纸题名不是中国经营报")
    if re.sub(r"\D", "", data.get("cn_number") or "") != "110151":
        errors.append("CN号不是11-0151")
    if data.get("year") != issue.publish_date.year or data.get("submission_year") != issue.publish_date.year:
        errors.append("清单年份与当前刊期不一致")
    if data.get("month") != issue.publish_date.month:
        errors.append("清单月份与当前刊期不一致")
    if data.get("submission_sequence") != issue.publish_date.month:
        errors.append("当年寄送次序与当前月份不一致")
    if data.get("quantity") != expected_quantity:
        errors.append(f"清单散报总份数应为{expected_quantity}份")
    if scheduled_issues and sorted(covered_issues) != sorted(scheduled_issues):
        errors.append("月底合寄记录尚未覆盖当月全部出版期次")
    if data.get("unit_price") != 5:
        errors.append("每份定价不是5元")
    return ("verified" if not errors else "mismatch"), data, errors


def _store_import_documents(
    db: Session,
    *,
    batch: ShippingWaybillImportBatch,
    issue: Issue,
    parsed_documents: list[ParsedImportDocument],
) -> None:
    rows = db.query(ShippingWaybillImportRow).filter(
        ShippingWaybillImportRow.batch_id == batch.id
    ).order_by(ShippingWaybillImportRow.id).all()
    linked_row_ids: set[int] = set()
    for document in parsed_documents:
        candidates = [
            row for row in rows
            if _normalized(row.recipient_name) == _normalized(document.associated_name)
        ]
        linked_row = candidates[0] if len(candidates) == 1 else None
        if linked_row:
            status, extracted_data, errors = _validate_sample_submission_document(
                db=db,
                issue=issue,
                document=document,
                row=linked_row,
            )
            linked_row_ids.add(linked_row.id)
        else:
            status = "pending_review"
            extracted_data = document.extracted_data
            errors = [
                "清单关联姓名未能唯一对应运单行"
                if document.associated_name
                else "清单工作表名称缺少“关联姓名”"
            ]
        db.add(ShippingWaybillImportDocument(
            batch_id=batch.id,
            linked_import_row_id=linked_row.id if linked_row else None,
            document_type=document.document_type,
            source_sheet=document.source_sheet,
            status=status,
            extracted_data=extracted_data,
            validation_errors=errors,
            parser_version="1",
            checked_at=datetime.now(),
        ))
        if status != "verified":
            batch.warning_count += 1

    for row in rows:
        mentions_checklist = any(
            "样报缴送清单" in _text(value) for value in (row.raw_values or [])
        )
        if mentions_checklist and row.id not in linked_row_ids:
            db.add(ShippingWaybillImportDocument(
                batch_id=batch.id,
                linked_import_row_id=row.id,
                document_type="sample_submission_list",
                source_sheet="未检测到工作表",
                status="missing",
                extracted_data={"associated_name": row.recipient_name},
                validation_errors=["文件中未检测到对应的样报缴送清单工作表"],
                parser_version="1",
                checked_at=datetime.now(),
            ))
            batch.warning_count += 1


def preview_import(
    db: Session,
    issue_id: int,
    filename: str,
    content: bytes,
    user: User,
    *,
    reparse: bool = False,
) -> ShippingWaybillImportBatch:
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="刊期不存在")
    digest = sha256(content).hexdigest()
    existing = db.query(
        ShippingWaybillImportBatch.id,
        ShippingWaybillImportBatch.status,
    ).filter(
        ShippingWaybillImportBatch.issue_number == issue.issue_number,
        ShippingWaybillImportBatch.file_hash == digest,
    ).first()
    if existing:
        if existing.status == WaybillImportStatus.confirmed.value or not reparse:
            return _get_import_batch(db, existing.id)
    if reparse:
        draft_ids = [row.id for row in db.query(ShippingWaybillImportBatch.id).filter(
            ShippingWaybillImportBatch.issue_id == issue.id,
            ShippingWaybillImportBatch.status == WaybillImportStatus.previewed.value,
        ).all()]
        if draft_ids:
            db.query(ShippingWaybillImportRow).filter(
                ShippingWaybillImportRow.batch_id.in_(draft_ids)
            ).delete(synchronize_session="evaluate")
            db.query(ShippingWaybillImportBatch).filter(
                ShippingWaybillImportBatch.id.in_(draft_ids)
            ).delete(synchronize_session="evaluate")
            db.flush()
    else:
        active_draft_id = db.query(ShippingWaybillImportBatch.id).filter(
            ShippingWaybillImportBatch.issue_id == issue.id,
            ShippingWaybillImportBatch.status == WaybillImportStatus.previewed.value,
        ).order_by(ShippingWaybillImportBatch.id.desc()).scalar()
        if active_draft_id:
            raise HTTPException(status_code=409, detail="本期已有运单核对草稿，请先继续处理或选择重新解析")

    payload = _parse_waybill_workbook_payload(content)
    parsed = payload.rows
    details = _details_for_issue(db, issue.issue_number)
    by_id = {detail.id: detail for detail in details}
    full: dict[tuple[str, str, str], list[ShippingDetail]] = defaultdict(list)
    phone_address: dict[tuple[str, str], list[ShippingDetail]] = defaultdict(list)
    name_address: dict[tuple[str, str], list[ShippingDetail]] = defaultdict(list)
    for detail in details:
        detail_name, detail_phone, detail_address = _match_key(
            detail.name,
            detail.phone or "",
            detail.address or "",
        )
        full[(detail_name, detail_phone, detail_address)].append(detail)
        if detail_phone and detail_address:
            phone_address[(detail_phone, detail_address)].append(detail)
        if detail_name and detail_address:
            name_address[(detail_name, detail_address)].append(detail)

    existing_tracking = {
        (carrier, tracking_no)
        for carrier, tracking_no in db.query(ShippingPackage.carrier, ShippingPackage.tracking_no).all()
    }
    seen_tracking: set[tuple[str, str]] = set()
    row_results: list[
        tuple[ParsedWaybillRow, str, str | None, int | None, list[int], list[int], int]
    ] = []
    quantities_by_detail: dict[int, int] = defaultdict(int)

    for row in parsed:
        status = WaybillMatchStatus.unmatched.value
        reason: str | None = None
        match: ShippingDetail | None = None
        if row.parse_reason:
            status, reason = WaybillMatchStatus.invalid.value, row.parse_reason
        elif row.quantity <= 0 or not row.recipient_name:
            status, reason = WaybillMatchStatus.invalid.value, "姓名或份数无效"
        elif not row.no_tracking_required and not row.tracking_no:
            status, reason = WaybillMatchStatus.invalid.value, "缺少运单号"
        elif row.tracking_no and (row.carrier, row.tracking_no) in seen_tracking:
            status, reason = WaybillMatchStatus.duplicate.value, "文件内运单号重复"
        elif row.tracking_no and (row.carrier, row.tracking_no) in existing_tracking:
            status, reason = WaybillMatchStatus.duplicate.value, "运单号已存在"
        else:
            if row.tracking_no:
                seen_tracking.add((row.carrier, row.tracking_no))
            candidates: list[ShippingDetail] = []
            if row.source_detail_id:
                candidate = by_id.get(row.source_detail_id)
                candidates = [candidate] if candidate else []
                reason = None if candidate else "发货明细ID不属于本期"
            else:
                name, phone, address = _match_key(row.recipient_name, row.phone, row.address)
                candidates = full.get((name, phone, address), [])
                if not candidates and phone and address:
                    candidates = phone_address.get((phone, address), [])
                if not candidates and name and address:
                    candidates = name_address.get((name, address), [])
            if len(candidates) == 1:
                match = candidates[0]
                if match.is_mafei_warehouse_retention:
                    status = WaybillMatchStatus.invalid.value
                    reason = "马飞—库房留存请改为“转库留存/库存入库”核销"
                elif row.no_tracking_required and match.packages:
                    match = None
                    status, reason = WaybillMatchStatus.invalid.value, "该明细已有运单，不能改为无需运单"
                else:
                    status = WaybillMatchStatus.matched.value
                    quantities_by_detail[match.id] += row.quantity
            elif len(candidates) > 1:
                status, reason = WaybillMatchStatus.ambiguous.value, "匹配到多条发货明细"
            elif reason is None:
                reason = "未找到对应发货明细"
        deferral_ids: list[int] = []
        issue_numbers: list[int] = []
        consolidation_quantity = 0
        if status == WaybillMatchStatus.matched.value and match is not None:
            deferrals, consolidation_error = _matching_month_end_deferrals(
                db,
                issue=issue,
                row=row,
                shipping_detail_id=match.id,
            )
            if consolidation_error:
                status = WaybillMatchStatus.ambiguous.value
                reason = consolidation_error
                quantities_by_detail[match.id] -= row.quantity
                match = None
            elif deferrals:
                deferral_ids = [item.id for item in deferrals]
                issue_numbers = sorted({item.issue_number for item in deferrals} | {issue.issue_number})
                has_current_deferral = any(
                    item.shipping_detail_id == match.id for item in deferrals
                )
                consolidation_quantity = sum(item.quantity for item in deferrals)
                if not has_current_deferral:
                    consolidation_quantity += row.quantity
        row_results.append((
            row,
            status,
            reason,
            match.id if match else None,
            deferral_ids,
            issue_numbers,
            consolidation_quantity,
        ))

    expected = _expected_quantity(db, issue)
    matched_quantity = sum(
        row.quantity for row, status, *_ in row_results
        if status == WaybillMatchStatus.matched.value
    )
    current_handled = sum(detail.handled_quantity for detail in details)
    projected_handled = current_handled + matched_quantity
    warning_count = sum(
        1 for _, status, *_ in row_results
        if status != WaybillMatchStatus.matched.value
    )
    for detail_id, imported_quantity in quantities_by_detail.items():
        detail = by_id[detail_id]
        if detail.handled_quantity + imported_quantity != (detail.quantity or 0):
            warning_count += 1

    batch = ShippingWaybillImportBatch(
        issue_id=issue.id,
        issue_number=issue.issue_number,
        filename=filename or "运单导入.xlsx",
        file_hash=digest,
        status=WaybillImportStatus.previewed.value,
        expected_quantity=expected,
        parsed_quantity=sum(row.quantity for row in parsed),
        matched_quantity=matched_quantity,
        pending_quantity=max(expected - projected_handled, 0),
        extra_quantity=max(projected_handled - expected, 0),
        matched_rows=sum(
            1 for _, status, *_ in row_results
            if status == WaybillMatchStatus.matched.value
        ),
        unmatched_rows=sum(
            1 for _, status, *_ in row_results
            if status != WaybillMatchStatus.matched.value
        ),
        warning_count=warning_count,
        created_by=getattr(user, "id", None),
    )
    db.add(batch)
    db.flush()
    db.execute(insert(ShippingWaybillImportRow), [
        {
            "batch_id": batch.id,
            "source_sheet": row.source_sheet,
            "source_row": row.source_row,
            "carrier": row.carrier,
            "tracking_no": row.tracking_no,
            "recipient_name": row.recipient_name,
            "phone": row.phone or None,
            "address": row.address or None,
            "quantity": row.quantity,
            "no_tracking_required": row.no_tracking_required,
            "raw_values": row.raw_values,
            "manual_reviewed": False,
            "match_status": status,
            "match_reason": reason,
            "shipping_detail_id": detail_id,
            "consolidation_deferral_ids": deferral_ids or None,
            "consolidation_issue_numbers": issue_numbers or None,
            "consolidation_quantity": consolidation_quantity,
        }
        for (
            row,
            status,
            reason,
            detail_id,
            deferral_ids,
            issue_numbers,
            consolidation_quantity,
        ) in row_results
    ])
    _store_import_documents(
        db,
        batch=batch,
        issue=issue,
        parsed_documents=payload.documents,
    )
    batch_id = batch.id
    db.commit()
    return _get_import_batch(db, batch_id)


def _get_import_batch(db: Session, batch_id: int) -> ShippingWaybillImportBatch:
    batch = db.query(ShippingWaybillImportBatch).filter(
        ShippingWaybillImportBatch.id == batch_id
    ).first()
    if not batch:
        raise HTTPException(status_code=404, detail="运单导入批次不存在")
    return batch


def get_import_batch(db: Session, batch_id: int) -> ShippingWaybillImportBatch:
    return _get_import_batch(db, batch_id)


def get_draft_import(db: Session, issue_id: int) -> ShippingWaybillImportBatch | None:
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="刊期不存在")
    previewed = db.query(ShippingWaybillImportBatch).filter(
        ShippingWaybillImportBatch.issue_id == issue_id,
        ShippingWaybillImportBatch.status == WaybillImportStatus.previewed.value,
    ).order_by(
        ShippingWaybillImportBatch.created_at.desc(),
        ShippingWaybillImportBatch.id.desc(),
    ).first()
    if previewed:
        return previewed
    return db.query(ShippingWaybillImportBatch).filter(
        ShippingWaybillImportBatch.issue_id == issue_id,
        ShippingWaybillImportBatch.status == WaybillImportStatus.confirmed.value,
        or_(
            ShippingWaybillImportBatch.unmatched_rows > 0,
            ShippingWaybillImportBatch.pending_quantity > 0,
        ),
    ).order_by(
        ShippingWaybillImportBatch.confirmed_at.desc(),
        ShippingWaybillImportBatch.id.desc(),
    ).first()


def _candidate_details_for_contact(
    details: list[ShippingDetail],
    *,
    name: str,
    phone: str,
    address: str,
) -> list[ShippingDetail]:
    full: dict[tuple[str, str, str], list[ShippingDetail]] = defaultdict(list)
    phone_address: dict[tuple[str, str], list[ShippingDetail]] = defaultdict(list)
    name_address: dict[tuple[str, str], list[ShippingDetail]] = defaultdict(list)
    for detail in details:
        detail_name, detail_phone, detail_address = _match_key(
            detail.name,
            detail.phone or "",
            detail.address or "",
        )
        full[(detail_name, detail_phone, detail_address)].append(detail)
        if detail_phone and detail_address:
            phone_address[(detail_phone, detail_address)].append(detail)
        if detail_name and detail_address:
            name_address[(detail_name, detail_address)].append(detail)
    normalized_name, normalized_phone, normalized_address = _match_key(name, phone, address)
    candidates = full.get((normalized_name, normalized_phone, normalized_address), [])
    if not candidates and normalized_phone and normalized_address:
        candidates = phone_address.get((normalized_phone, normalized_address), [])
    if not candidates and normalized_name and normalized_address:
        candidates = name_address.get((normalized_name, normalized_address), [])
    return candidates


def _candidate_details(details: list[ShippingDetail], row: ShippingWaybillImportRow) -> list[ShippingDetail]:
    return _candidate_details_for_contact(
        details,
        name=row.recipient_name,
        phone=row.phone or "",
        address=row.address or "",
    )


def _refresh_draft_row_consolidation(
    db: Session,
    *,
    batch: ShippingWaybillImportBatch,
    row: ShippingWaybillImportRow,
) -> None:
    row.consolidation_deferral_ids = None
    row.consolidation_issue_numbers = None
    row.consolidation_quantity = 0
    if row.match_status != WaybillMatchStatus.matched.value or row.shipping_detail_id is None:
        return
    issue = db.query(Issue).filter(Issue.id == batch.issue_id).first()
    if issue is None:
        return
    parsed = ParsedWaybillRow(
        source_sheet=row.source_sheet,
        source_row=row.source_row,
        carrier=row.carrier,
        tracking_no=row.tracking_no,
        recipient_name=row.recipient_name,
        phone=row.phone or "",
        address=row.address or "",
        quantity=row.quantity,
        no_tracking_required=row.no_tracking_required,
        raw_values=row.raw_values,
    )
    deferrals, error = _matching_month_end_deferrals(
        db,
        issue=issue,
        row=parsed,
        shipping_detail_id=row.shipping_detail_id,
    )
    if error:
        row.match_status = WaybillMatchStatus.ambiguous.value
        row.match_reason = error
        row.shipping_detail_id = None
        return
    if deferrals:
        row.consolidation_deferral_ids = [item.id for item in deferrals]
        row.consolidation_issue_numbers = sorted(
            {item.issue_number for item in deferrals} | {batch.issue_number}
        )
        row.consolidation_quantity = sum(item.quantity for item in deferrals)
        if not any(item.shipping_detail_id == row.shipping_detail_id for item in deferrals):
            row.consolidation_quantity += row.quantity
    for document in row.documents:
        if document.document_type != "sample_submission_list":
            continue
        extracted = dict(document.extracted_data or {})
        extracted.pop("expected_quantity", None)
        extracted.pop("covered_issue_numbers", None)
        parsed_document = ParsedImportDocument(
            source_sheet=document.source_sheet,
            document_type=document.document_type,
            associated_name=extracted.get("associated_name") or "",
            extracted_data=extracted,
        )
        status, extracted, errors = _validate_sample_submission_document(
            db=db,
            issue=issue,
            document=parsed_document,
            row=row,
        )
        document.status = status
        document.extracted_data = extracted
        document.validation_errors = errors
        document.checked_at = datetime.now()


def _match_draft_row(
    db: Session,
    batch: ShippingWaybillImportBatch,
    row: ShippingWaybillImportRow,
    preferred_detail_id: int | None = None,
) -> None:
    row.consolidation_deferral_ids = None
    row.consolidation_issue_numbers = None
    row.consolidation_quantity = 0
    if row.match_status == WaybillMatchStatus.ignored.value:
        return
    details = _details_for_issue(db, batch.issue_number)
    by_id = {detail.id: detail for detail in details}
    preferred = by_id.get(preferred_detail_id) if preferred_detail_id is not None else None
    if preferred_detail_id is not None and preferred is None:
        raise HTTPException(status_code=400, detail="所选发货明细不属于本期")

    row.match_reason = None
    row.shipping_detail_id = preferred_detail_id
    if row.quantity <= 0 or not row.recipient_name.strip():
        row.match_status = WaybillMatchStatus.invalid.value
        row.match_reason = "姓名或份数无效"
        return
    if not row.no_tracking_required and not row.tracking_no:
        row.match_status = WaybillMatchStatus.invalid.value
        row.match_reason = "缺少运单号"
        return
    if row.tracking_no:
        package_duplicate = db.query(ShippingPackage.id).filter(
            ShippingPackage.carrier == row.carrier,
            ShippingPackage.tracking_no == row.tracking_no,
        ).first()
        draft_duplicate = db.query(ShippingWaybillImportRow.id).filter(
            ShippingWaybillImportRow.batch_id == batch.id,
            ShippingWaybillImportRow.id != row.id,
            ShippingWaybillImportRow.carrier == row.carrier,
            ShippingWaybillImportRow.tracking_no == row.tracking_no,
            ShippingWaybillImportRow.match_status != WaybillMatchStatus.ignored.value,
        ).first()
        if package_duplicate or draft_duplicate:
            row.match_status = WaybillMatchStatus.duplicate.value
            row.match_reason = "运单号已存在" if package_duplicate else "待导入数据中运单号重复"
            return

    candidates = [preferred] if preferred else _candidate_details(details, row)
    if len(candidates) == 1:
        detail = candidates[0]
        row.shipping_detail_id = detail.id
        if detail.is_mafei_warehouse_retention:
            row.match_status = WaybillMatchStatus.invalid.value
            row.match_reason = "马飞—库房留存只能按“转库留存/库存入库”核销"
            return
        if row.no_tracking_required and detail.packages:
            row.match_status = WaybillMatchStatus.invalid.value
            row.match_reason = "该明细已有运单，不能改为无需运单"
            return
        row.match_status = WaybillMatchStatus.matched.value
        _refresh_draft_row_consolidation(db, batch=batch, row=row)
        return
    row.shipping_detail_id = None
    if len(candidates) > 1:
        row.match_status = WaybillMatchStatus.ambiguous.value
        row.match_reason = "匹配到多条发货明细，请人工选择"
    else:
        row.match_status = WaybillMatchStatus.unmatched.value
        row.match_reason = "未找到对应发货明细"


def _recalculate_batch(db: Session, batch: ShippingWaybillImportBatch) -> None:
    rows = list(batch.rows)
    matched = [row for row in rows if row.match_status == WaybillMatchStatus.matched.value]
    unresolved = [
        row for row in rows
        if row.match_status not in {WaybillMatchStatus.matched.value, WaybillMatchStatus.ignored.value}
    ]
    batch.parsed_quantity = sum(max(row.quantity or 0, 0) for row in rows)
    batch.matched_quantity = sum(max(row.quantity or 0, 0) for row in matched)
    batch.matched_rows = len(matched)
    batch.unmatched_rows = len(unresolved)
    details = _details_for_issue(db, batch.issue_number)
    current_handled = sum(detail.handled_quantity for detail in details)
    projected = (
        current_handled
        if batch.status == WaybillImportStatus.confirmed.value
        else current_handled + batch.matched_quantity
    )
    batch.pending_quantity = max(batch.expected_quantity - projected, 0)
    batch.extra_quantity = max(projected - batch.expected_quantity, 0)

    quantities_by_detail: dict[int, int] = defaultdict(int)
    for row in matched:
        if row.shipping_detail_id is not None:
            quantities_by_detail[row.shipping_detail_id] += row.quantity
    by_id = {detail.id: detail for detail in details}
    if batch.status == WaybillImportStatus.confirmed.value:
        detail_warnings = sum(
            1
            for detail_id in quantities_by_detail
            if detail_id in by_id
            and by_id[detail_id].handled_quantity != (by_id[detail_id].quantity or 0)
        )
    else:
        detail_warnings = sum(
            1
            for detail_id, imported in quantities_by_detail.items()
            if detail_id in by_id
            and by_id[detail_id].handled_quantity + imported != (by_id[detail_id].quantity or 0)
        )
    document_warnings = sum(
        1 for document in batch.documents if document.status != "verified"
    )
    batch.warning_count = len(unresolved) + detail_warnings + document_warnings


def repair_postal_30_preview_rows(
    db: Session,
    *,
    issue_number: int,
    username: str = "system",
) -> WaybillPreviewRepairResult:
    """Repair compact F/G/H postal-30 rows from their stored raw-cell snapshot."""
    batch = db.query(ShippingWaybillImportBatch).filter(
        ShippingWaybillImportBatch.issue_number == issue_number,
        ShippingWaybillImportBatch.status == WaybillImportStatus.previewed.value,
    ).order_by(
        ShippingWaybillImportBatch.created_at.desc(),
        ShippingWaybillImportBatch.id.desc(),
    ).first()
    if not batch:
        raise ValueError(f"{issue_number} 期没有待确认的实际发货预览批次")

    result = WaybillPreviewRepairResult(
        issue_number=issue_number,
        batch_id=batch.id,
    )
    changes: list[dict[str, Any]] = []
    for row in batch.rows:
        if "邮政30" not in row.source_sheet or not row.raw_values:
            continue
        name, phone, address, quantity, layout = _parse_postal_30_fields(row.raw_values)
        if layout != "compact_fgh" or not name or quantity <= 0:
            continue

        before = {
            "name": row.recipient_name,
            "phone": row.phone,
            "address": row.address,
            "quantity": row.quantity,
            "match_status": row.match_status,
            "shipping_detail_id": row.shipping_detail_id,
        }
        row.carrier = "邮政"
        row.recipient_name = name
        row.phone = phone or None
        row.address = address or None
        row.quantity = quantity
        row.manual_reviewed = False
        _match_draft_row(db, batch, row)
        after = {
            "name": row.recipient_name,
            "phone": row.phone,
            "address": row.address,
            "quantity": row.quantity,
            "match_status": row.match_status,
            "shipping_detail_id": row.shipping_detail_id,
        }
        if before == after:
            continue
        result.repaired_rows += 1
        result.repaired_quantity += quantity
        result.row_ids.append(row.id)
        changes.append({"row_id": row.id, "source_row": row.source_row, "before": before, "after": after})

    if result.repaired_rows:
        db.flush()
        _recalculate_batch(db, batch)
        record_operation(
            db,
            table_name="shipping_waybill_import_batches",
            record_id=batch.id,
            record_name=f"{issue_number}期实际发货解析修复",
            action="repair_waybill_parse",
            username=username,
            issue_number=issue_number,
            changes={
                "layout": "compact_fgh",
                "repaired_rows": result.repaired_rows,
                "repaired_quantity": result.repaired_quantity,
                "rows": changes,
            },
        )
    return result


def update_import_row(
    db: Session,
    batch_id: int,
    row_id: int,
    body: WaybillImportRowUpdate,
    user: User | None = None,
) -> ShippingWaybillImportBatch:
    batch = _get_import_batch(db, batch_id)
    row = next((candidate for candidate in batch.rows if candidate.id == row_id), None)
    if not row:
        raise HTTPException(status_code=404, detail="运单导入行不存在")
    if (
        batch.status == WaybillImportStatus.confirmed.value
        and row.match_status == WaybillMatchStatus.matched.value
    ):
        raise HTTPException(status_code=409, detail="该行已经生成实际运单，不能在导入批次中修改")

    fields = body.model_fields_set
    if "carrier" in fields:
        row.carrier = (body.carrier or "").strip() or "中通"
    if "tracking_no" in fields:
        row.tracking_no = _tracking(body.tracking_no) or None
    if "recipient_name" in fields:
        row.recipient_name = (body.recipient_name or "").strip()
    if "phone" in fields:
        row.phone = (body.phone or "").strip() or None
    if "address" in fields:
        row.address = (body.address or "").strip() or None
    if "quantity" in fields:
        row.quantity = body.quantity or 0
    if "no_tracking_required" in fields:
        row.no_tracking_required = bool(body.no_tracking_required)
        if row.no_tracking_required:
            row.tracking_no = None
            if not row.carrier or row.carrier == "中通":
                row.carrier = "无需运单"
        elif row.carrier == "无需运单":
            row.carrier = "中通"

    ignored = body.ignored if "ignored" in fields else row.match_status == WaybillMatchStatus.ignored.value
    if ignored:
        reason = (body.ignore_reason or "").strip()
        if not reason:
            raise HTTPException(status_code=400, detail="忽略未匹配行时必须填写原因")
        row.match_status = WaybillMatchStatus.ignored.value
        row.match_reason = f"已人工忽略：{reason}"
        row.shipping_detail_id = None
    else:
        if row.match_status == WaybillMatchStatus.ignored.value:
            row.match_status = WaybillMatchStatus.unmatched.value
        preferred = body.shipping_detail_id if "shipping_detail_id" in fields else None
        _match_draft_row(db, batch, row, preferred)
    row.manual_reviewed = True
    db.flush()
    if (
        batch.status == WaybillImportStatus.confirmed.value
        and row.match_status == WaybillMatchStatus.matched.value
    ):
        _materialize_import_row(db, batch=batch, row=row)
        db.flush()
    _recalculate_batch(db, batch)
    record_operation(
        db,
        user=user,
        table_name="shipping_waybill_import_rows",
        record_id=row.id,
        record_name=f"{row.source_sheet} 第{row.source_row}行",
        action="review_waybill",
        issue_number=batch.issue_number,
        changes={
            "match_status": row.match_status,
            "shipping_detail_id": row.shipping_detail_id,
            "quantity": row.quantity,
        },
    )
    db.commit()
    db.refresh(batch)
    return batch


def add_import_row(
    db: Session,
    batch_id: int,
    body: WaybillImportRowCreate,
    user: User | None = None,
) -> ShippingWaybillImportBatch:
    batch = _get_import_batch(db, batch_id)
    next_row = max((row.source_row for row in batch.rows if row.source_sheet == "人工补充"), default=0) + 1
    row = ShippingWaybillImportRow(
        batch=batch,
        source_sheet="人工补充",
        source_row=next_row,
        carrier=(body.carrier or "中通").strip() or "中通",
        tracking_no=_tracking(body.tracking_no) or None,
        recipient_name=body.recipient_name.strip(),
        phone=(body.phone or "").strip() or None,
        address=(body.address or "").strip() or None,
        quantity=body.quantity,
        no_tracking_required=body.no_tracking_required,
        raw_values=[],
        manual_reviewed=True,
        match_status=WaybillMatchStatus.unmatched.value,
    )
    if row.no_tracking_required:
        row.tracking_no = None
        row.carrier = "无需运单"
    db.add(row)
    db.flush()
    _match_draft_row(db, batch, row, body.shipping_detail_id)
    db.flush()
    if (
        batch.status == WaybillImportStatus.confirmed.value
        and row.match_status == WaybillMatchStatus.matched.value
    ):
        _materialize_import_row(db, batch=batch, row=row)
        db.flush()
    _recalculate_batch(db, batch)
    record_operation(
        db,
        user=user,
        table_name="shipping_waybill_import_rows",
        record_id=row.id,
        record_name="人工补充运单",
        action="add_waybill_row",
        issue_number=batch.issue_number,
        changes={"match_status": row.match_status, "quantity": row.quantity},
    )
    db.commit()
    db.refresh(batch)
    return batch


def _refresh_legacy_shipping_fields(detail: ShippingDetail) -> None:
    if detail.shipping_requirement == "no_tracking_required":
        detail.shipped_at = None
        detail.shipped_quantity = None
        detail.tracking_no = None
        return
    direct_packages = [package for package in detail.packages if not package.allocations]
    allocated_packages = [allocation.package for allocation in detail.package_allocations]
    packages = list(dict.fromkeys([*direct_packages, *allocated_packages]))
    if not packages:
        return
    detail.shipped_quantity = sum(package.quantity or 0 for package in direct_packages) + sum(
        allocation.quantity or 0 for allocation in detail.package_allocations
    )
    detail.shipped_at = max(package.shipped_at for package in packages)
    detail.tracking_no = packages[0].tracking_no if len(packages) == 1 else None


def _materialize_matched_row(
    db: Session,
    row: ShippingWaybillImportRow,
    *,
    shipped_at: datetime | None = None,
) -> ShippingDetail:
    if row.match_status != WaybillMatchStatus.matched.value or row.shipping_detail_id is None:
        raise HTTPException(status_code=400, detail="运单行尚未关联发货明细")
    detail = row.shipping_detail or db.query(ShippingDetail).filter(
        ShippingDetail.id == row.shipping_detail_id
    ).first()
    if not detail:
        raise HTTPException(status_code=400, detail="关联的发货明细不存在")
    if detail.is_mafei_warehouse_retention:
        raise HTTPException(
            status_code=400,
            detail="马飞—库房留存不能生成运单，只能按“转库留存/库存入库”核销",
        )
    if row.no_tracking_required:
        if detail.packages:
            raise HTTPException(status_code=409, detail="该明细已有运单，不能改为无需运单")
        detail.shipping_requirement = "no_tracking_required"
        _refresh_legacy_shipping_fields(detail)
        return detail
    if not row.tracking_no:
        raise HTTPException(status_code=400, detail="缺少运单号")
    existing_for_row = db.query(ShippingPackage).filter(
        ShippingPackage.import_row_id == row.id
    ).first()
    if existing_for_row:
        return detail
    duplicate = db.query(ShippingPackage).filter(
        ShippingPackage.carrier == row.carrier,
        ShippingPackage.tracking_no == row.tracking_no,
    ).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="该运单号已经生成实际发货记录")
    detail.shipping_requirement = "tracking_required"
    package = ShippingPackage(
        shipping_detail_id=detail.id,
        import_row_id=row.id,
        carrier=row.carrier,
        tracking_no=row.tracking_no,
        quantity=row.quantity,
        shipped_at=shipped_at or datetime.now(),
    )
    db.add(package)
    detail.packages.append(package)
    db.flush()
    _refresh_legacy_shipping_fields(detail)
    return detail


def _validated_consolidation_deferrals(
    db: Session,
    *,
    batch: ShippingWaybillImportBatch,
    row: ShippingWaybillImportRow,
) -> list[ShippingDeferral]:
    ids = list(dict.fromkeys(row.consolidation_deferral_ids or []))
    if len(ids) != len(row.consolidation_deferral_ids or []):
        raise HTTPException(status_code=409, detail="月底合寄预览包含重复的待合寄记录，请重新解析")
    deferrals = db.query(ShippingDeferral).filter(
        ShippingDeferral.id.in_(ids)
    ).with_for_update().order_by(ShippingDeferral.issue_number, ShippingDeferral.id).all()
    issue = db.query(Issue).filter(Issue.id == batch.issue_id).first()
    if issue is None:
        raise HTTPException(status_code=409, detail="运单批次对应的刊期不存在")
    if len(deferrals) != len(ids) or any(item.status != "pending" for item in deferrals):
        raise HTTPException(status_code=409, detail="月底待合寄记录已经变化，请重新解析运单文件")
    if any(
            item.deferral_type != MONTH_END_CONSOLIDATION
            or item.shipping_detail_id is None
            or not _deferral_targets_issue(item, issue)
        for item in deferrals
    ):
        raise HTTPException(status_code=409, detail="月底待合寄记录的目标刊期或明细归属已经变化")
    if len({item.consolidation_batch for item in deferrals}) != 1:
        raise HTTPException(status_code=409, detail="月底待合寄记录不属于同一个合寄批次")
    row_key = _match_key(row.recipient_name, row.phone or "", row.address or "")
    if any(not _contact_key_matches(
        _match_key(
            item.detail_name_snapshot or "",
            item.detail_phone_snapshot or "",
            item.detail_address_snapshot or "",
        ),
        row_key,
    ) for item in deferrals):
        raise HTTPException(status_code=409, detail="月底待合寄记录的收件信息已经变化，请重新核对")
    current = [item for item in deferrals if item.shipping_detail_id == row.shipping_detail_id]
    if len(current) > 1 or (current and current[0].quantity != row.quantity):
        raise HTTPException(status_code=409, detail="本期月底待合寄记录与运单行份数不一致")
    expected_quantity = sum(item.quantity for item in deferrals)
    if not current:
        expected_quantity += row.quantity
    expected_issues = sorted({item.issue_number for item in deferrals} | {batch.issue_number})
    if (
        expected_quantity != row.consolidation_quantity
        or expected_issues != sorted(row.consolidation_issue_numbers or [])
    ):
        raise HTTPException(status_code=409, detail="月底合寄范围或总份数已经变化，请重新解析")
    return deferrals


def _materialize_consolidated_row(
    db: Session,
    *,
    row: ShippingWaybillImportRow,
    deferrals: list[ShippingDeferral],
    shipped_at: datetime,
) -> ShippingDetail:
    detail = row.shipping_detail
    if detail is None or row.shipping_detail_id is None or not row.tracking_no:
        raise HTTPException(status_code=400, detail="月底合寄运单缺少本期发货明细或运单号")
    duplicate = db.query(ShippingPackage.id).filter(
        ShippingPackage.carrier == row.carrier,
        ShippingPackage.tracking_no == row.tracking_no,
    ).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="该月底合寄运单号已经生成实际发货记录")
    package = ShippingPackage(
        shipping_detail=detail,
        import_row=row,
        carrier=row.carrier,
        tracking_no=row.tracking_no,
        quantity=row.consolidation_quantity,
        shipped_at=shipped_at,
    )
    db.add(package)
    current_has_deferral = False
    touched_details: set[ShippingDetail] = {detail}
    for deferral in deferrals:
        allocation = ShippingPackageAllocation(
            package=package,
            shipping_detail=deferral.shipping_detail,
            deferral=deferral,
            quantity=deferral.quantity,
        )
        db.add(allocation)
        current_has_deferral = current_has_deferral or deferral.shipping_detail_id == detail.id
        touched_details.add(deferral.shipping_detail)
        deferral.status = "fulfilled"
        deferral.fulfilled_package = package
        deferral.fulfilled_at = shipped_at
    if not current_has_deferral:
        db.add(ShippingPackageAllocation(
            package=package,
            shipping_detail=detail,
            quantity=row.quantity,
        ))
    for document in row.documents:
        document.shipping_package = package
    db.flush()
    for touched_detail in touched_details:
        _refresh_legacy_shipping_fields(touched_detail)
    return detail


def _materialize_import_row(
    db: Session,
    *,
    batch: ShippingWaybillImportBatch,
    row: ShippingWaybillImportRow,
    shipped_at: datetime | None = None,
) -> ShippingDetail:
    if row.consolidation_candidate:
        deferrals = _validated_consolidation_deferrals(db, batch=batch, row=row)
        return _materialize_consolidated_row(
            db,
            row=row,
            deferrals=deferrals,
            shipped_at=shipped_at or datetime.now(),
        )
    return _materialize_matched_row(db, row, shipped_at=shipped_at)


def restore_orphaned_confirmed_waybills(
    db: Session,
    *,
    issue: Issue,
) -> OrphanedWaybillRestoreResult:
    """Reconnect confirmed import rows after their plan details were recreated.

    Clearing legacy plan rows used to cascade-delete their packages while leaving
    the confirmed import rows behind. Re-uploading the plan can safely reconstruct
    those packages from the preserved carrier, tracking number and quantity. Exact
    recipient matches are restored first; split packages are restored only when a
    single same-name detail has exactly the same remaining quantity.
    """
    details = _details_for_issue(db, issue.issue_number)
    detail_ids = {detail.id for detail in details}
    batches = db.query(ShippingWaybillImportBatch).filter(
        ShippingWaybillImportBatch.issue_id == issue.id,
        ShippingWaybillImportBatch.status == WaybillImportStatus.confirmed.value,
    ).order_by(ShippingWaybillImportBatch.created_at, ShippingWaybillImportBatch.id).all()

    orphaned_by_batch: dict[int, list[ShippingWaybillImportRow]] = {}
    restored_rows: list[ShippingWaybillImportRow] = []
    for batch in batches:
        orphaned = [
            row for row in batch.rows
            if row.match_status == WaybillMatchStatus.matched.value
            and row.package is None
            and (row.shipping_detail_id is None or row.shipping_detail_id not in detail_ids)
        ]
        if orphaned:
            orphaned_by_batch[batch.id] = orphaned

    existing_tracking = {
        (carrier, tracking_no)
        for carrier, tracking_no in db.query(ShippingPackage.carrier, ShippingPackage.tracking_no).all()
    }
    seen_tracking: set[tuple[str, str]] = set()
    initial_handled = {detail.id: detail.handled_quantity for detail in details}
    initial_physical = {detail.id: detail.physical_shipped_quantity for detail in details}
    restored_by_detail: dict[int, int] = defaultdict(int)
    package_values: list[dict[str, Any]] = []
    package_tracking_by_detail: dict[int, list[str]] = defaultdict(list)
    package_shipped_at_by_detail: dict[int, list[datetime]] = defaultdict(list)

    def restore_row(
        row: ShippingWaybillImportRow,
        detail: ShippingDetail,
        batch: ShippingWaybillImportBatch,
    ) -> None:
        row.shipping_detail_id = detail.id
        row.match_status = WaybillMatchStatus.matched.value
        row.match_reason = None
        if row.no_tracking_required:
            detail.shipping_requirement = "no_tracking_required"
        else:
            shipped_at = batch.confirmed_at or datetime.now()
            package_values.append({
                "shipping_detail_id": detail.id,
                "import_row_id": row.id,
                "carrier": row.carrier,
                "tracking_no": row.tracking_no,
                "quantity": row.quantity,
                "shipped_at": shipped_at,
            })
            package_tracking_by_detail[detail.id].append(row.tracking_no)
            package_shipped_at_by_detail[detail.id].append(shipped_at)
        restored_by_detail[detail.id] += max(row.quantity or 0, 0)
        restored_rows.append(row)

    for batch in batches:
        for row in orphaned_by_batch.get(batch.id, []):
            row.shipping_detail_id = None
            row.match_reason = None
            if row.quantity <= 0 or not row.recipient_name.strip():
                row.match_status = WaybillMatchStatus.invalid.value
                row.match_reason = "姓名或份数无效"
                continue
            if not row.no_tracking_required and not row.tracking_no:
                row.match_status = WaybillMatchStatus.invalid.value
                row.match_reason = "缺少运单号"
                continue
            if row.tracking_no:
                tracking_key = (row.carrier, row.tracking_no)
                if tracking_key in existing_tracking or tracking_key in seen_tracking:
                    row.match_status = WaybillMatchStatus.duplicate.value
                    row.match_reason = "运单号已存在"
                    continue
                seen_tracking.add(tracking_key)
            candidates = _candidate_details(details, row)
            if len(candidates) > 1:
                row.match_status = WaybillMatchStatus.ambiguous.value
                row.match_reason = "匹配到多条发货明细，请人工选择"
                continue
            if not candidates:
                row.match_status = WaybillMatchStatus.unmatched.value
                row.match_reason = "未找到对应发货明细"
                continue
            restore_row(row, candidates[0], batch)

    # A single plan detail can be represented by several physical packages. Such
    # rows may have intentionally been bulk-matched because their addresses differ
    # from the plan. Restore that decision only when name and remaining total make
    # the target unambiguous.
    for batch in batches:
        unresolved = [
            row for row in orphaned_by_batch.get(batch.id, [])
            if row.match_status != WaybillMatchStatus.matched.value
        ]
        grouped: dict[str, list[ShippingWaybillImportRow]] = defaultdict(list)
        for row in unresolved:
            name_key = _normalized(row.recipient_name)
            if name_key and row.quantity > 0:
                grouped[name_key].append(row)
        for name_key, rows in grouped.items():
            row_quantity = sum(max(row.quantity or 0, 0) for row in rows)
            candidates = [
                detail for detail in details
                if _normalized(detail.name) == name_key
                and max(
                    (detail.quantity or 0)
                    - initial_handled[detail.id]
                    - restored_by_detail[detail.id],
                    0,
                ) == row_quantity
            ]
            if len(candidates) != 1:
                continue
            detail = candidates[0]
            for row in rows:
                restore_row(row, detail, batch)

    if package_values:
        db.execute(insert(ShippingPackage), package_values)
    for detail in details:
        if detail.id not in restored_by_detail:
            continue
        if detail.shipping_requirement == "no_tracking_required":
            detail.shipped_at = None
            detail.shipped_quantity = None
            detail.tracking_no = None
            continue
        detail.shipped_quantity = initial_physical[detail.id] + restored_by_detail[detail.id]
        shipped_times = package_shipped_at_by_detail[detail.id]
        if shipped_times:
            detail.shipped_at = max([time for time in [detail.shipped_at, *shipped_times] if time is not None])
        existing_tracking_numbers = [package.tracking_no for package in detail.packages]
        all_tracking_numbers = [*existing_tracking_numbers, *package_tracking_by_detail[detail.id]]
        detail.tracking_no = all_tracking_numbers[0] if len(all_tracking_numbers) == 1 else None
    db.flush()

    restored_adjustments = 0
    adjustments = db.query(ShippingFulfillmentAdjustment).filter(
        ShippingFulfillmentAdjustment.issue_id == issue.id,
        ShippingFulfillmentAdjustment.shipping_detail_id.is_(None),
        ShippingFulfillmentAdjustment.detail_name_snapshot.isnot(None),
    ).all()
    for adjustment in adjustments:
        candidates = _candidate_details_for_contact(
            details,
            name=adjustment.detail_name_snapshot or "",
            phone=adjustment.detail_phone_snapshot or "",
            address=adjustment.detail_address_snapshot or "",
        )
        if len(candidates) == 1:
            adjustment.shipping_detail_id = candidates[0].id
            restored_adjustments += 1

    restored_deferrals = 0
    deferrals = db.query(ShippingDeferral).filter(
        ShippingDeferral.issue_id == issue.id,
        ShippingDeferral.shipping_detail_id.is_(None),
        ShippingDeferral.detail_name_snapshot.isnot(None),
    ).all()
    for deferral in deferrals:
        candidates = _candidate_details_for_contact(
            details,
            name=deferral.detail_name_snapshot or "",
            phone=deferral.detail_phone_snapshot or "",
            address=deferral.detail_address_snapshot or "",
        )
        if len(candidates) == 1:
            deferral.shipping_detail_id = candidates[0].id
            restored_deferrals += 1

    db.flush()
    for batch in batches:
        _recalculate_batch(db, batch)

    restored_ids = {row.id for row in restored_rows}
    unresolved_rows = sum(
        1 for rows in orphaned_by_batch.values() for row in rows if row.id not in restored_ids
    )
    return OrphanedWaybillRestoreResult(
        restored_rows=len(restored_rows),
        restored_quantity=sum(max(row.quantity or 0, 0) for row in restored_rows),
        unresolved_rows=unresolved_rows,
        restored_adjustments=restored_adjustments,
        restored_deferrals=restored_deferrals,
    )


def bulk_match_import_rows(
    db: Session,
    batch_id: int,
    body: WaybillBulkMatchIn,
    user: User | None = None,
) -> ShippingWaybillImportBatch:
    batch = _get_import_batch(db, batch_id)
    requested_ids = list(dict.fromkeys(body.row_ids))
    rows_by_id = {row.id: row for row in batch.rows}
    rows = [rows_by_id[row_id] for row_id in requested_ids if row_id in rows_by_id]
    if len(rows) != len(requested_ids):
        raise HTTPException(status_code=404, detail="部分运单行不属于当前导入批次")
    if any(row.match_status == WaybillMatchStatus.matched.value for row in rows):
        raise HTTPException(status_code=409, detail="所选运单中包含已经关联的行")
    detail = next(
        (candidate for candidate in _details_for_issue(db, batch.issue_number) if candidate.id == body.shipping_detail_id),
        None,
    )
    if not detail:
        raise HTTPException(status_code=400, detail="所选发货明细不属于本期")
    selected_quantity = sum(max(row.quantity or 0, 0) for row in rows)
    other_draft_quantity = sum(
        max(row.quantity or 0, 0)
        for row in batch.rows
        if row.id not in requested_ids
        and row.match_status == WaybillMatchStatus.matched.value
        and row.shipping_detail_id == detail.id
        and row.package is None
    )
    if detail.handled_quantity + other_draft_quantity + selected_quantity > (detail.quantity or 0):
        raise HTTPException(status_code=400, detail="所选运单份数超过该发货明细的待核销份数")

    for row in rows:
        _match_draft_row(db, batch, row, detail.id)
        if row.match_status != WaybillMatchStatus.matched.value:
            raise HTTPException(status_code=400, detail=row.match_reason or "运单无法关联")
        row.manual_reviewed = True
    db.flush()
    if batch.status == WaybillImportStatus.confirmed.value:
        for row in rows:
            _materialize_import_row(db, batch=batch, row=row)
        db.flush()
    _recalculate_batch(db, batch)
    record_operation(
        db,
        user=user,
        table_name="shipping_waybill_import_batches",
        record_id=batch.id,
        record_name=batch.filename,
        action="bulk_match_waybills",
        issue_number=batch.issue_number,
        changes={
            "row_ids": requested_ids,
            "shipping_detail_id": detail.id,
            "quantity": selected_quantity,
        },
    )
    db.commit()
    db.refresh(batch)
    return batch


def confirm_import(db: Session, batch_id: int, user: User) -> ShippingWaybillImportBatch:
    batch = db.query(ShippingWaybillImportBatch).filter(
        ShippingWaybillImportBatch.id == batch_id
    ).with_for_update().first()
    if not batch:
        raise HTTPException(status_code=404, detail="运单导入批次不存在")
    if batch.status == WaybillImportStatus.confirmed.value:
        return batch

    consolidation_by_row: dict[int, list[ShippingDeferral]] = {}
    for row in batch.rows:
        if (
            row.match_status == WaybillMatchStatus.matched.value
            and row.shipping_detail_id is not None
            and row.consolidation_candidate
        ):
            consolidation_by_row[row.id] = _validated_consolidation_deferrals(
                db,
                batch=batch,
                row=row,
            )

    now = datetime.now()
    for row in batch.rows:
        if row.match_status != WaybillMatchStatus.matched.value or row.shipping_detail_id is None:
            continue
        if row.id in consolidation_by_row:
            _materialize_consolidated_row(
                db,
                row=row,
                deferrals=consolidation_by_row[row.id],
                shipped_at=now,
            )
        else:
            _materialize_matched_row(db, row, shipped_at=now)

    batch.status = WaybillImportStatus.confirmed.value
    batch.confirmed_at = now
    db.flush()
    _recalculate_batch(db, batch)
    record_operation(
        db,
        user=user,
        table_name="shipping_waybill_import_batches",
        record_id=batch.id,
        record_name=batch.filename,
        action="import_waybills",
        issue_number=batch.issue_number,
        changes={
            "matched_rows": batch.matched_rows,
            "matched_quantity": batch.matched_quantity,
            "pending_quantity": batch.pending_quantity,
            "warning_count": batch.warning_count,
        },
    )
    db.commit()
    db.refresh(batch)
    return batch


def fulfillment_summary(db: Session, issue_id: int) -> FulfillmentSummaryOut:
    latest_confirm_report = (
        select(IssueAuditSnapshot.report_total)
        .where(
            IssueAuditSnapshot.issue_id == Issue.id,
            IssueAuditSnapshot.snapshot_type == "confirm",
        )
        .order_by(IssueAuditSnapshot.created_at.desc(), IssueAuditSnapshot.id.desc())
        .limit(1)
        .correlate(Issue)
        .scalar_subquery()
    )
    issue_row = (
        db.query(Issue, latest_confirm_report.label("confirmed_report_total"))
        .filter(Issue.id == issue_id)
        .first()
    )
    if not issue_row:
        raise HTTPException(status_code=404, detail="刊期不存在")
    issue, confirmed_report_total = issue_row
    details = (
        db.query(ShippingDetail)
        .options(
            joinedload(ShippingDetail.packages),
            joinedload(ShippingDetail.package_allocations),
            noload(ShippingDetail.fulfillment_adjustments),
            noload(ShippingDetail.deferrals),
        )
        .filter(
            ShippingDetail.issue_number == issue.issue_number,
            ShippingDetail.source_type != ShippingDetailSourceType.complaint_makeup,
        )
        .order_by(ShippingDetail.id)
        .all()
    )
    planned = sum(detail.quantity or 0 for detail in details)
    expected = int(confirmed_report_total) if confirmed_report_total is not None else planned
    tracked = sum(
        detail.physical_shipped_quantity
        for detail in details
        if detail.shipping_requirement != "no_tracking_required"
    )
    no_tracking = sum(
        detail.quantity or 0 for detail in details if detail.shipping_requirement == "no_tracking_required"
    )
    adjustments = db.query(ShippingFulfillmentAdjustment).filter(
        ShippingFulfillmentAdjustment.issue_id == issue.id
    ).order_by(ShippingFulfillmentAdjustment.created_at, ShippingFulfillmentAdjustment.id).all()
    deferrals = db.query(ShippingDeferral).filter(
        ShippingDeferral.issue_id == issue.id
    ).order_by(ShippingDeferral.created_at, ShippingDeferral.id).all()
    adjustment_quantity = sum(max(adjustment.quantity or 0, 0) for adjustment in adjustments)
    warehouse_stock_in_quantity = sum(
        max(adjustment.quantity or 0, 0)
        for adjustment in adjustments
        if adjustment.adjustment_type == WAREHOUSE_STOCK_IN
    )
    no_shipment_quantity = adjustment_quantity - warehouse_stock_in_quantity
    deferred_quantity = sum(
        max(deferral.quantity or 0, 0) for deferral in deferrals if deferral.status == "pending"
    )
    twice_monthly_deferred_quantity = sum(
        max(deferral.quantity or 0, 0)
        for deferral in deferrals
        if deferral.status == "pending" and deferral.deferral_type == TWICE_MONTHLY_CONSOLIDATION
    )
    month_end_deferred_quantity = sum(
        max(deferral.quantity or 0, 0)
        for deferral in deferrals
        if deferral.status == "pending" and deferral.deferral_type == MONTH_END_CONSOLIDATION
    )
    attributed_adjustment_quantity = sum(
        max(adjustment.quantity or 0, 0) for adjustment in adjustments if adjustment.is_attributed
    )
    unattributed_adjustment_quantity = adjustment_quantity - attributed_adjustment_quantity
    actual_shipped = tracked + no_tracking
    handled = tracked + no_tracking + adjustment_quantity
    pending = max(expected - handled, 0)
    unexplained_pending = max(pending - deferred_quantity, 0)
    extra = max(handled - expected, 0)
    if extra:
        status = "exception"
    elif pending and handled:
        status = "partial"
    elif pending:
        status = "pending"
    else:
        status = "shipped"
    if actual_shipped > expected:
        shipment_status = "exception"
    elif actual_shipped == expected:
        shipment_status = "shipped"
    elif actual_shipped:
        shipment_status = "partial"
    else:
        shipment_status = "pending"
    latest = (
        db.query(ShippingWaybillImportBatch)
        .options(joinedload(ShippingWaybillImportBatch.rows))
        .filter(ShippingWaybillImportBatch.issue_number == issue.issue_number)
        .order_by(
            ShippingWaybillImportBatch.created_at.desc(),
            ShippingWaybillImportBatch.id.desc(),
        )
        .first()
    )
    source_by_detail: dict[int, int] = defaultdict(int)
    if latest:
        for row in latest.rows:
            if row.shipping_detail_id is not None and row.match_status == WaybillMatchStatus.matched.value:
                source_by_detail[row.shipping_detail_id] += max(row.quantity or 0, 0)
    pending_deferred_by_detail: dict[int, int] = defaultdict(int)
    pending_twice_monthly_by_detail: dict[int, int] = defaultdict(int)
    pending_month_end_by_detail: dict[int, int] = defaultdict(int)
    for deferral in deferrals:
        if deferral.status == "pending" and deferral.shipping_detail_id is not None:
            pending_deferred_by_detail[deferral.shipping_detail_id] += max(deferral.quantity or 0, 0)
            if deferral.deferral_type == TWICE_MONTHLY_CONSOLIDATION:
                pending_twice_monthly_by_detail[deferral.shipping_detail_id] += max(deferral.quantity or 0, 0)
            elif deferral.deferral_type == MONTH_END_CONSOLIDATION:
                pending_month_end_by_detail[deferral.shipping_detail_id] += max(deferral.quantity or 0, 0)
    adjustment_by_detail: dict[int, int] = defaultdict(int)
    for adjustment in adjustments:
        if adjustment.shipping_detail_id is not None:
            adjustment_by_detail[adjustment.shipping_detail_id] += max(adjustment.quantity or 0, 0)
    gap_details: list[ShippingGapDetailOut] = []
    if latest:
        for detail in details:
            source_quantity = source_by_detail.get(detail.id, 0)
            delivered_or_in_source = max(source_quantity, detail.physical_shipped_quantity)
            raw_gap = max(
                (detail.quantity or 0) - delivered_or_in_source - adjustment_by_detail.get(detail.id, 0),
                0,
            )
            deferred = min(pending_deferred_by_detail.get(detail.id, 0), raw_gap)
            twice_monthly_deferred = min(
                pending_twice_monthly_by_detail.get(detail.id, 0),
                deferred,
            )
            month_end_deferred = min(
                pending_month_end_by_detail.get(detail.id, 0),
                max(deferred - twice_monthly_deferred, 0),
            )
            remaining = max(raw_gap - deferred, 0)
            if raw_gap:
                marker = f"{detail.sheet_name or ''}{detail.frequency or ''}"
                gap_details.append(ShippingGapDetailOut(
                    shipping_detail_id=detail.id,
                    name=detail.name,
                    phone=detail.phone,
                    address=detail.address,
                    channel=detail.channel,
                    sheet_name=detail.sheet_name,
                    frequency=detail.frequency,
                    planned_quantity=detail.quantity or 0,
                    source_quantity=source_quantity,
                    deferred_quantity=deferred,
                    twice_monthly_deferred_quantity=twice_monthly_deferred,
                    month_end_deferred_quantity=month_end_deferred,
                    remaining_quantity=remaining,
                    suggested_month_end="月底" in marker or "整月" in marker,
                    required_adjustment_type=(
                        WAREHOUSE_STOCK_IN if detail.is_mafei_warehouse_retention else None
                    ),
                ))
        gap_details.sort(key=lambda item: (-item.remaining_quantity, item.shipping_detail_id))
    return FulfillmentSummaryOut(
        issue_id=issue.id,
        issue_number=issue.issue_number,
        expected_quantity=expected,
        planned_quantity=planned,
        handled_quantity=handled,
        tracked_quantity=tracked,
        no_tracking_quantity=no_tracking,
        actual_shipped_quantity=actual_shipped,
        adjustment_quantity=adjustment_quantity,
        no_shipment_quantity=no_shipment_quantity,
        warehouse_stock_in_quantity=warehouse_stock_in_quantity,
        deferred_quantity=deferred_quantity,
        twice_monthly_deferred_quantity=twice_monthly_deferred_quantity,
        month_end_deferred_quantity=month_end_deferred_quantity,
        unexplained_pending_quantity=unexplained_pending,
        attributed_adjustment_quantity=attributed_adjustment_quantity,
        unattributed_adjustment_quantity=unattributed_adjustment_quantity,
        pending_quantity=pending,
        extra_quantity=extra,
        package_count=len({
            *(package.id for detail in details for package in detail.packages),
            *(allocation.shipping_package_id for detail in details for allocation in detail.package_allocations),
        }),
        pending_detail_count=sum(
            1
            for detail in details
            if min(
                detail.physical_shipped_quantity + adjustment_by_detail.get(detail.id, 0),
                detail.quantity or 0,
            ) < (detail.quantity or 0)
        ),
        status=status,
        shipment_status=shipment_status,
        latest_import=latest,
        adjustments=adjustments,
        deferrals=deferrals,
        gap_details=gap_details,
    )


def _detail_for_adjustment(
    db: Session,
    *,
    issue: Issue,
    shipping_detail_id: int,
) -> ShippingDetail:
    detail = db.query(ShippingDetail).filter(
        ShippingDetail.id == shipping_detail_id,
        ShippingDetail.issue_number == issue.issue_number,
        ShippingDetail.source_type != ShippingDetailSourceType.complaint_makeup,
    ).first()
    if not detail:
        raise HTTPException(status_code=400, detail="所选发货明细不属于本期确认版计划")
    return detail


def _attribute_adjustment(
    adjustment: ShippingFulfillmentAdjustment,
    detail: ShippingDetail,
) -> None:
    adjustment.shipping_detail_id = detail.id
    adjustment.detail_name_snapshot = detail.name
    adjustment.detail_phone_snapshot = detail.phone
    adjustment.detail_address_snapshot = detail.address
    adjustment.detail_channel_snapshot = detail.channel
    adjustment.detail_company_snapshot = detail.company
    adjustment.detail_quantity_snapshot = detail.quantity


def create_fulfillment_adjustment(
    db: Session,
    issue_id: int,
    body: FulfillmentAdjustmentIn,
    user: User,
) -> FulfillmentSummaryOut:
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="刊期不存在")
    reason = body.reason.strip()
    if not reason:
        raise HTTPException(status_code=400, detail="核销原因不能为空")
    before = fulfillment_summary(db, issue_id)
    if body.quantity > before.pending_quantity:
        raise HTTPException(status_code=400, detail="无需发货份数不能超过当前待处理份数")
    detail = _detail_for_adjustment(
        db,
        issue=issue,
        shipping_detail_id=body.shipping_detail_id,
    )
    if detail.is_mafei_warehouse_retention and body.adjustment_type != WAREHOUSE_STOCK_IN:
        raise HTTPException(
            status_code=400,
            detail="马飞—库房留存必须按“转库留存/库存入库”核销",
        )
    if not detail.is_mafei_warehouse_retention and body.adjustment_type == WAREHOUSE_STOCK_IN:
        raise HTTPException(
            status_code=400,
            detail="“转库留存/库存入库”仅适用于马飞—库房留存",
        )
    automatic_stop_exists = bool(
        db.query(ShippingFulfillmentAdjustment.id)
        .filter(ShippingFulfillmentAdjustment.shipping_detail_id == detail.id)
        .filter(ShippingFulfillmentAdjustment.source == PLAN_STATUS_ADJUSTMENT_SOURCE)
        .first()
    )
    if automatic_stop_exists:
        raise HTTPException(
            status_code=400,
            detail="该明细已由计划停发自动核销，不能重复登记",
        )
    adjustment = ShippingFulfillmentAdjustment(
        issue_id=issue.id,
        issue_number=issue.issue_number,
        adjustment_type=body.adjustment_type,
        source=MANUAL_ADJUSTMENT_SOURCE,
        quantity=body.quantity,
        reason=reason,
        created_by=getattr(user, "id", None),
    )
    _attribute_adjustment(adjustment, detail)
    db.add(adjustment)
    db.flush()
    record_operation(
        db,
        user=user,
        table_name="shipping_fulfillment_adjustments",
        record_id=adjustment.id,
        record_name=adjustment.reason,
        action="create",
        issue_number=issue.issue_number,
        changes={
            "adjustment_type": adjustment.adjustment_type,
            "quantity": adjustment.quantity,
            "reason": adjustment.reason,
            "shipping_detail_id": adjustment.shipping_detail_id,
        },
    )
    db.commit()
    return fulfillment_summary(db, issue_id)


def convert_import_row_to_warehouse_stock_in(
    db: Session,
    batch_id: int,
    row_id: int,
    user: User | None = None,
    *,
    username: str | None = None,
) -> ShippingWaybillImportBatch:
    """Atomically convert a Mafei no-tracking import row to stock-in.

    The source row remains as an ignored audit record, while the full planned
    quantity is classified as warehouse stock-in.  Any legacy no-shipment
    adjustment already attributed to the same detail is reclassified instead
    of double-counted.
    """
    batch = db.query(ShippingWaybillImportBatch).filter(
        ShippingWaybillImportBatch.id == batch_id
    ).with_for_update().first()
    if not batch:
        raise HTTPException(status_code=404, detail="运单导入批次不存在")
    row = db.query(ShippingWaybillImportRow).filter(
        ShippingWaybillImportRow.id == row_id,
        ShippingWaybillImportRow.batch_id == batch.id,
    ).with_for_update().first()
    if not row:
        raise HTTPException(status_code=404, detail="运单导入行不存在")
    if row.shipping_detail_id is None:
        raise HTTPException(status_code=400, detail="请先关联马飞—库房留存发货明细")

    detail = db.query(ShippingDetail).filter(
        ShippingDetail.id == row.shipping_detail_id,
        ShippingDetail.issue_number == batch.issue_number,
    ).with_for_update().first()
    if not detail or not detail.is_mafei_warehouse_retention:
        raise HTTPException(status_code=400, detail="“转库留存/库存入库”仅适用于马飞—库房留存")

    linked_rows = db.query(ShippingWaybillImportRow).filter(
        ShippingWaybillImportRow.shipping_detail_id == detail.id,
        ShippingWaybillImportRow.match_status != WaybillMatchStatus.ignored.value,
    ).with_for_update().all()
    tracked_rows = [
        candidate for candidate in linked_rows
        if candidate.tracking_no or not candidate.no_tracking_required
    ]
    direct_packages = db.query(ShippingPackage.id).filter(
        ShippingPackage.shipping_detail_id == detail.id
    ).count()
    allocations = db.query(ShippingPackageAllocation.id).filter(
        ShippingPackageAllocation.shipping_detail_id == detail.id
    ).count()
    active_deferrals = db.query(ShippingDeferral.id).filter(
        ShippingDeferral.shipping_detail_id == detail.id,
        ShippingDeferral.status != "cancelled",
    ).count()
    has_legacy_physical_shipment = any(
        value is not None
        for value in (detail.shipped_at, detail.shipped_quantity, detail.tracking_no)
    )
    if tracked_rows or direct_packages or allocations or active_deferrals or has_legacy_physical_shipment:
        raise HTTPException(
            status_code=409,
            detail="马飞—库房留存已有真实运单、实发或合寄记录，不能自动改为库存入库",
        )

    planned_quantity = max(detail.quantity or 0, 0)
    if planned_quantity <= 0:
        raise HTTPException(status_code=400, detail="马飞—库房留存计划份数必须大于0")
    adjustments = db.query(ShippingFulfillmentAdjustment).filter(
        ShippingFulfillmentAdjustment.shipping_detail_id == detail.id
    ).with_for_update().all()
    adjustment_total = sum(max(adjustment.quantity or 0, 0) for adjustment in adjustments)
    if adjustment_total > planned_quantity:
        raise HTTPException(
            status_code=409,
            detail=f"马飞库房已核销 {adjustment_total} 份，超过当前计划 {planned_quantity} 份，请先人工核对",
        )

    audit_username = username or (None if user is not None else "系统修复")
    for adjustment in adjustments:
        previous_type = adjustment.adjustment_type
        previous_reason = adjustment.reason
        adjustment.adjustment_type = WAREHOUSE_STOCK_IN
        adjustment.reason = WAREHOUSE_STOCK_IN_REASON
        _attribute_adjustment(adjustment, detail)
        if previous_type != WAREHOUSE_STOCK_IN or previous_reason != WAREHOUSE_STOCK_IN_REASON:
            record_operation(
                db,
                user=user,
                username=audit_username,
                table_name="shipping_fulfillment_adjustments",
                record_id=adjustment.id,
                record_name=WAREHOUSE_STOCK_IN_REASON,
                action="update",
                issue_number=batch.issue_number,
                channel=detail.channel,
                changes={
                    "old_adjustment_type": previous_type,
                    "adjustment_type": WAREHOUSE_STOCK_IN,
                    "old_reason": previous_reason,
                    "reason": WAREHOUSE_STOCK_IN_REASON,
                    "shipping_detail_id": detail.id,
                },
            )

    missing_quantity = planned_quantity - adjustment_total
    if missing_quantity:
        adjustment = ShippingFulfillmentAdjustment(
            issue_id=batch.issue_id,
            issue_number=batch.issue_number,
            adjustment_type=WAREHOUSE_STOCK_IN,
            quantity=missing_quantity,
            reason=WAREHOUSE_STOCK_IN_REASON,
            created_by=getattr(user, "id", None),
        )
        _attribute_adjustment(adjustment, detail)
        db.add(adjustment)
        db.flush()
        record_operation(
            db,
            user=user,
            username=audit_username,
            table_name="shipping_fulfillment_adjustments",
            record_id=adjustment.id,
            record_name=WAREHOUSE_STOCK_IN_REASON,
            action="create",
            issue_number=batch.issue_number,
            channel=detail.channel,
            changes={
                "adjustment_type": WAREHOUSE_STOCK_IN,
                "quantity": missing_quantity,
                "reason": WAREHOUSE_STOCK_IN_REASON,
                "shipping_detail_id": detail.id,
                "source_import_row_id": row.id,
            },
        )

    affected_batches: dict[int, ShippingWaybillImportBatch] = {batch.id: batch}
    for linked_row in linked_rows:
        affected_batches[linked_row.batch.id] = linked_row.batch
        linked_row.match_status = WaybillMatchStatus.ignored.value
        linked_row.match_reason = WAREHOUSE_STOCK_IN_IMPORT_REASON
        linked_row.manual_reviewed = True
        record_operation(
            db,
            user=user,
            username=audit_username,
            table_name="shipping_waybill_import_rows",
            record_id=linked_row.id,
            record_name=f"{linked_row.source_sheet} 第{linked_row.source_row}行",
            action="review_waybill",
            issue_number=batch.issue_number,
            changes={
                "match_status": WaybillMatchStatus.ignored.value,
                "shipping_detail_id": detail.id,
                "quantity": linked_row.quantity,
                "conversion": WAREHOUSE_STOCK_IN,
            },
        )

    detail.shipping_requirement = "tracking_required"
    detail.shipped_at = None
    detail.shipped_quantity = None
    detail.tracking_no = None
    db.flush()
    db.expire(detail, ["fulfillment_adjustments"])
    for affected_batch in affected_batches.values():
        _recalculate_batch(db, affected_batch)
    db.commit()
    return _get_import_batch(db, batch.id)


def attribute_fulfillment_adjustment(
    db: Session,
    adjustment_id: int,
    body: FulfillmentAdjustmentAttributionIn,
    user: User,
) -> FulfillmentSummaryOut:
    adjustment = db.query(ShippingFulfillmentAdjustment).filter(
        ShippingFulfillmentAdjustment.id == adjustment_id
    ).first()
    if not adjustment:
        raise HTTPException(status_code=404, detail="无需发货核销记录不存在")
    if adjustment.source == PLAN_STATUS_ADJUSTMENT_SOURCE:
        raise HTTPException(status_code=409, detail="计划停发自动归因不能改绑，请在发货计划中调整停发状态")
    issue = db.query(Issue).filter(Issue.id == adjustment.issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="刊期不存在")
    detail = _detail_for_adjustment(
        db,
        issue=issue,
        shipping_detail_id=body.shipping_detail_id,
    )
    if detail.is_mafei_warehouse_retention:
        adjustment.adjustment_type = WAREHOUSE_STOCK_IN
        adjustment.reason = WAREHOUSE_STOCK_IN_REASON
    elif adjustment.adjustment_type == WAREHOUSE_STOCK_IN:
        raise HTTPException(
            status_code=400,
            detail="“转库留存/库存入库”仅适用于马飞—库房留存",
        )
    previous_detail_id = adjustment.shipping_detail_id
    _attribute_adjustment(adjustment, detail)
    db.flush()
    sync_plan_status_adjustment(db, detail=detail, user=user)
    record_operation(
        db,
        user=user,
        table_name="shipping_fulfillment_adjustments",
        record_id=adjustment.id,
        record_name=adjustment.reason,
        action="attribute",
        issue_number=issue.issue_number,
        changes={
            "old_shipping_detail_id": previous_detail_id,
            "shipping_detail_id": detail.id,
            "quantity": adjustment.quantity,
        },
    )
    db.commit()
    return fulfillment_summary(db, issue.id)


def delete_fulfillment_adjustment(
    db: Session,
    adjustment_id: int,
    user: User,
) -> FulfillmentSummaryOut:
    adjustment = db.query(ShippingFulfillmentAdjustment).filter(
        ShippingFulfillmentAdjustment.id == adjustment_id
    ).first()
    if not adjustment:
        raise HTTPException(status_code=404, detail="无需发货核销记录不存在")
    if adjustment.source == PLAN_STATUS_ADJUSTMENT_SOURCE:
        raise HTTPException(status_code=409, detail="计划停发自动归因不能单独撤销，请先恢复该发货计划")
    issue_id = adjustment.issue_id
    issue_number = adjustment.issue_number
    detail = adjustment.shipping_detail
    changes = {
        "adjustment_type": adjustment.adjustment_type,
        "source": adjustment.source,
        "quantity": adjustment.quantity,
        "reason": adjustment.reason,
    }
    db.delete(adjustment)
    db.flush()
    if detail is not None:
        sync_plan_status_adjustment(db, detail=detail, user=user)
    record_operation(
        db,
        user=user,
        table_name="shipping_fulfillment_adjustments",
        record_id=adjustment_id,
        record_name=adjustment.reason,
        action="delete",
        issue_number=issue_number,
        changes=changes,
    )
    db.commit()
    return fulfillment_summary(db, issue_id)


def refresh_detail_shipping_fields(detail: ShippingDetail) -> None:
    _refresh_legacy_shipping_fields(detail)


def _month_bounds(value: date) -> tuple[date, date]:
    start = value.replace(day=1)
    if value.month == 12:
        return start, date(value.year + 1, 1, 1)
    return start, date(value.year, value.month + 1, 1)


def _deferral_target(
    db: Session,
    issue: Issue,
    deferral_type: str,
) -> tuple[int | None, date, str]:
    """Resolve the immutable shipment batch for this issue within its natural month."""
    month_start, next_month = _month_bounds(issue.publish_date)
    schedules = db.query(PublicationSchedule).filter(
        PublicationSchedule.publish_date >= month_start,
        PublicationSchedule.publish_date < next_month,
        PublicationSchedule.is_suspended.is_(False),
    ).order_by(PublicationSchedule.publish_date).all()
    issue_numbers_by_date = {
        candidate.publish_date: candidate.issue_number
        for candidate in db.query(Issue).filter(
            Issue.publish_date >= month_start,
            Issue.publish_date < next_month,
        ).all()
    }
    monthly_issues = [
        (row.publish_date, row.issue_number or issue_numbers_by_date.get(row.publish_date))
        for row in schedules
    ]
    if issue.publish_date not in {publish_date for publish_date, _ in monthly_issues}:
        monthly_issues.append((issue.publish_date, issue.issue_number))
        monthly_issues.sort(key=lambda item: item[0])
    if not monthly_issues:
        monthly_issues = [(issue.publish_date, issue.issue_number)]

    current_index = next(
        index for index, item in enumerate(monthly_issues) if item[0] == issue.publish_date
    )
    if deferral_type == TWICE_MONTHLY_CONSOLIDATION:
        target_index = min(1, len(monthly_issues) - 1) if current_index <= 1 else len(monthly_issues) - 1
        phase = "first" if current_index <= 1 else "second"
    else:
        target_index = len(monthly_issues) - 1
        phase = "month_end"
    target_date, target_issue_number = monthly_issues[target_index]
    batch = f"{issue.publish_date:%Y-%m}-{phase}"
    return target_issue_number, target_date, batch


def create_shipping_deferrals(
    db: Session,
    issue_id: int,
    body: ShippingDeferralBulkIn,
    user: User,
) -> FulfillmentSummaryOut:
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="刊期不存在")
    target_issue_number, target_publish_date, consolidation_batch = _deferral_target(
        db,
        issue,
        body.deferral_type,
    )
    summary = fulfillment_summary(db, issue_id)
    gap_by_detail = {
        item.shipping_detail_id: item.remaining_quantity for item in summary.gap_details
    }
    requested_ids = [item.shipping_detail_id for item in body.items]
    if len(requested_ids) != len(set(requested_ids)):
        raise HTTPException(status_code=400, detail="同一发货明细不能重复选择")
    details = db.query(ShippingDetail).filter(
        ShippingDetail.id.in_(requested_ids),
        ShippingDetail.issue_number == issue.issue_number,
        ShippingDetail.source_type != ShippingDetailSourceType.complaint_makeup,
    ).all()
    by_id = {detail.id: detail for detail in details}
    if len(by_id) != len(requested_ids):
        raise HTTPException(status_code=400, detail="所选发货明细不属于本期确认版计划")
    total = 0
    created: list[ShippingDeferral] = []
    for item in body.items:
        available = gap_by_detail.get(item.shipping_detail_id, 0)
        if item.quantity > available:
            raise HTTPException(status_code=400, detail="待合寄份数不能超过该明细的计划缺口")
        detail = by_id[item.shipping_detail_id]
        if detail.is_mafei_warehouse_retention:
            raise HTTPException(
                status_code=400,
                detail="马飞—库房留存不能标记合寄，必须按“转库留存/库存入库”核销",
            )
        deferral = ShippingDeferral(
            issue_id=issue.id,
            issue_number=issue.issue_number,
            shipping_detail_id=detail.id,
            deferral_type=body.deferral_type,
            target_issue_number=target_issue_number,
            target_publish_date=target_publish_date,
            consolidation_batch=consolidation_batch,
            quantity=item.quantity,
            reason=body.reason.strip(),
            detail_name_snapshot=detail.name,
            detail_phone_snapshot=detail.phone,
            detail_address_snapshot=detail.address,
            detail_channel_snapshot=detail.channel,
            created_by=getattr(user, "id", None),
        )
        db.add(deferral)
        created.append(deferral)
        total += item.quantity
    if total > summary.unexplained_pending_quantity:
        raise HTTPException(status_code=400, detail="待合寄总数不能超过当前未解释待处理份数")
    db.flush()
    record_operation(
        db,
        user=user,
        table_name="shipping_deferrals",
        record_id=created[0].id,
        record_name=body.reason.strip(),
        action="bulk_create",
        issue_number=issue.issue_number,
        changes={
            "deferral_ids": [item.id for item in created],
            "detail_ids": requested_ids,
            "quantity": total,
            "type": body.deferral_type,
        },
    )
    db.commit()
    return fulfillment_summary(db, issue_id)


def list_pending_shipping_deferrals(db: Session) -> list[ShippingDeferral]:
    return db.query(ShippingDeferral).filter(
        ShippingDeferral.status == "pending"
    ).order_by(
        ShippingDeferral.target_publish_date,
        ShippingDeferral.deferral_type,
        ShippingDeferral.detail_name_snapshot,
        ShippingDeferral.issue_number,
        ShippingDeferral.id,
    ).all()


def delete_shipping_deferral(
    db: Session,
    deferral_id: int,
    user: User,
) -> FulfillmentSummaryOut:
    deferral = db.query(ShippingDeferral).filter(ShippingDeferral.id == deferral_id).first()
    if not deferral:
        raise HTTPException(status_code=404, detail="待合寄记录不存在")
    if deferral.status != "pending":
        raise HTTPException(status_code=409, detail="已完成的合寄记录不能删除")
    issue_id = deferral.issue_id
    db.delete(deferral)
    record_operation(
        db,
        user=user,
        table_name="shipping_deferrals",
        record_id=deferral.id,
        record_name=deferral.reason,
        action="delete",
        issue_number=deferral.issue_number,
        changes={"quantity": deferral.quantity, "shipping_detail_id": deferral.shipping_detail_id},
    )
    db.commit()
    return fulfillment_summary(db, issue_id)


def create_consolidated_package(
    db: Session,
    body: ConsolidatedPackageIn,
    user: User,
) -> ConsolidatedPackageOut:
    ids = [item.deferral_id for item in body.deferrals]
    if len(ids) != len(set(ids)):
        raise HTTPException(status_code=400, detail="待合寄记录不能重复选择")
    deferrals = db.query(ShippingDeferral).filter(
        ShippingDeferral.id.in_(ids)
    ).with_for_update().all()
    if len(deferrals) != len(ids) or any(item.status != "pending" for item in deferrals):
        raise HTTPException(status_code=409, detail="部分待合寄记录不存在或已经完成")
    if any(item.shipping_detail_id is None for item in deferrals):
        raise HTTPException(status_code=400, detail="待合寄记录缺少收件明细归属")
    deferral_types = {item.deferral_type for item in deferrals}
    if len(deferral_types) != 1:
        raise HTTPException(status_code=400, detail="一张合寄运单不能混合两种合寄方式")
    consolidation_batches = {
        item.consolidation_batch for item in deferrals if item.consolidation_batch
    }
    has_legacy_batch = any(not item.consolidation_batch for item in deferrals)
    if len(consolidation_batches) > 1 or (has_legacy_batch and consolidation_batches):
        raise HTTPException(status_code=400, detail="一张合寄运单只能核销同一个目标批次")
    recipient_keys = {
        _match_key(
            item.detail_name_snapshot or "",
            item.detail_phone_snapshot or "",
            item.detail_address_snapshot or "",
        )
        for item in deferrals
    }
    if len(recipient_keys) != 1:
        raise HTTPException(status_code=400, detail="一张合寄运单只能关联同一收件人")
    carrier = body.carrier.strip()
    tracking_no = _tracking(body.tracking_no)
    duplicate = db.query(ShippingPackage.id).filter(
        ShippingPackage.carrier == carrier,
        ShippingPackage.tracking_no == tracking_no,
    ).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="该运单号已存在")
    total = sum(max(item.quantity or 0, 0) for item in deferrals)
    primary = deferrals[0].shipping_detail
    package = ShippingPackage(
        shipping_detail=primary,
        carrier=carrier,
        tracking_no=tracking_no,
        quantity=total,
        shipped_at=body.shipped_at or datetime.now(),
    )
    db.add(package)
    db.flush()
    for deferral in deferrals:
        allocation = ShippingPackageAllocation(
            package=package,
            shipping_detail=deferral.shipping_detail,
            deferral=deferral,
            quantity=deferral.quantity,
        )
        db.add(allocation)
        deferral.status = "fulfilled"
        deferral.fulfilled_package_id = package.id
        deferral.fulfilled_at = package.shipped_at
    db.flush()
    for detail in {item.shipping_detail for item in deferrals}:
        refresh_detail_shipping_fields(detail)
    record_operation(
        db,
        user=user,
        table_name="shipping_packages",
        record_id=package.id,
        record_name=tracking_no,
        action="create_consolidated",
        issue_number=max(item.issue_number for item in deferrals),
        changes={"deferral_ids": ids, "quantity": total, "carrier": carrier},
    )
    db.commit()
    return ConsolidatedPackageOut(
        package_id=package.id,
        carrier=carrier,
        tracking_no=tracking_no,
        quantity=total,
        fulfilled_deferral_ids=ids,
    )


def transfer_shipping_plan_quantity(
    db: Session,
    issue_id: int,
    body: ShippingPlanTransferIn,
    user: User,
) -> ShippingPlanTransferOut:
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="刊期不存在")
    source = _detail_for_adjustment(db, issue=issue, shipping_detail_id=body.source_detail_id)
    if source.quantity - body.quantity < source.handled_quantity + source.deferred_quantity:
        raise HTTPException(status_code=409, detail="转出后份数不能少于已经发出、核销或待合寄的份数")
    if body.target_detail_id:
        target = _detail_for_adjustment(db, issue=issue, shipping_detail_id=body.target_detail_id)
        if target.id == source.id:
            raise HTTPException(status_code=400, detail="转入与转出明细不能相同")
        target.quantity = (target.quantity or 0) + body.quantity
    else:
        if not (body.target_name or "").strip():
            raise HTTPException(status_code=400, detail="新增收件明细时必须填写收件人")
        target = ShippingDetail(
            issue_number=issue.issue_number,
            sheet_name=body.target_sheet_name,
            channel=body.target_channel,
            sub_channel="",
            transport="中通物流",
            frequency=body.target_frequency,
            status="正常",
            name=body.target_name.strip(),
            phone=(body.target_phone or "").strip() or None,
            address=(body.target_address or "").strip() or None,
            quantity=body.quantity,
            notes="计划纠错新增",
        )
        db.add(target)
        db.flush()
    source.quantity -= body.quantity
    record_operation(
        db,
        user=user,
        table_name="shipping_details",
        record_id=source.id,
        record_name=source.name,
        action="transfer_quantity",
        issue_number=issue.issue_number,
        channel=source.channel,
        changes={
            "quantity": body.quantity,
            "target_detail_id": target.id,
            "reason": body.reason.strip(),
            "source_quantity": source.quantity,
            "target_quantity": target.quantity,
        },
    )
    db.commit()
    planned = sum(detail.quantity or 0 for detail in _details_for_issue(db, issue.issue_number))
    return ShippingPlanTransferOut(
        source_detail_id=source.id,
        source_quantity=source.quantity,
        target_detail_id=target.id,
        target_quantity=target.quantity,
        planned_quantity=planned,
    )
