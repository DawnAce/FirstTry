"""邮局年改地址 Excel 解析（纯提取 → ``ParsedAddressChange``）。

表头名可能带括注（如「原读者起月日 (邮局2024读者明细)」），匹配时先剥掉括注再比。
表头签名：含 修改日期 / 新地址 / 编号。
"""

import io
import re
from dataclasses import dataclass
from typing import List

import openpyxl

_SIGNATURE = {"修改日期", "新地址", "编号"}

_FIELDS = {
    "修改日期": "change_date_raw",
    "姓名": "old_name",
    "联系电话": "old_phone",
    "省": "old_province",
    "市": "old_city",
    "区": "old_district",
    "详细地址": "old_detail",
    "份数": "old_copies_raw",
    "新姓名": "new_name",
    "新电话": "new_phone",
    "新地址": "new_address",
    "处理情况": "handling",
    "原读者起月日": "original_start_month",
    "实际起月日": "effective_start_month",
    "份数2": "new_copies_raw",
    "编号": "external_no_raw",
    "备注": "notes",
}


@dataclass
class ParsedAddressChange:
    row_no: int
    change_date_raw: str = ""
    old_name: str = ""
    old_phone: str = ""
    old_province: str = ""
    old_city: str = ""
    old_district: str = ""
    old_detail: str = ""
    old_copies_raw: str = ""
    new_name: str = ""
    new_phone: str = ""
    new_address: str = ""
    new_copies_raw: str = ""
    original_start_month: str = ""
    effective_start_month: str = ""
    handling: str = ""
    external_no_raw: str = ""
    notes: str = ""


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm_header(h: str) -> str:
    """剥掉尾部括注：「原读者起月日 (邮局2024读者明细)」→「原读者起月日」。"""
    return re.sub(r"\s*[（(].*[)）]\s*$", "", h).strip()


def _find_sheet(wb):
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        hmap = {}
        for i, v in enumerate(first):
            n = _norm_header(_cell(v))
            if n and n not in hmap:
                hmap[n] = i
        if _SIGNATURE.issubset(hmap.keys()):
            return ws, hmap
    return None, None


def is_postal_address_change_export(file_bytes: bytes) -> bool:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception:
        return False
    ws, _ = _find_sheet(wb)
    wb.close()
    return ws is not None


def parse_postal_address_changes(file_bytes: bytes) -> List[ParsedAddressChange]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws, hmap = _find_sheet(wb)
    if ws is None:
        wb.close()
        raise ValueError("无法识别的邮局改地址表：未找到含「修改日期/新地址/编号」表头的工作表")
    col = {field: hmap[name] for name, field in _FIELDS.items() if name in hmap}
    out: List[ParsedAddressChange] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_cell(v) for v in row):
            continue

        def get(field: str) -> str:
            idx = col.get(field)
            if idx is None or idx >= len(row):
                return ""
            return _cell(row[idx])

        ac = ParsedAddressChange(row_no=i)
        for field in _FIELDS.values():
            setattr(ac, field, get(field))
        if not ac.external_no_raw and not ac.new_address:
            continue
        out.append(ac)
    wb.close()
    return out
