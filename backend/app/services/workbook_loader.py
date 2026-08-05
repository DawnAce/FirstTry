"""Load ordinary or password-protected OOXML workbooks from upload bytes."""

import io
from zipfile import BadZipFile

import msoffcrypto
from fastapi import HTTPException
from msoffcrypto.exceptions import DecryptionError, FileFormatError, InvalidKeyError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def _invalid_excel_error(file_label: str) -> HTTPException:
    return HTTPException(
        status_code=422,
        detail=f"无法解析{file_label}，请确保上传的是未损坏的 .xlsx 文件",
    )


def load_uploaded_workbook(
    workbook_bytes: bytes,
    *,
    password: str | None = None,
    file_label: str = "上传文件",
):
    """Return an openpyxl workbook, decrypting an Office-encrypted file in memory."""
    try:
        return load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as plain_exc:
        try:
            office_file = msoffcrypto.OfficeFile(io.BytesIO(workbook_bytes))
        except (FileFormatError, OSError) as encrypted_exc:
            raise _invalid_excel_error(file_label) from encrypted_exc

        if not office_file.is_encrypted():
            raise _invalid_excel_error(file_label) from plain_exc

        normalized_password = (password or "").strip()
        if not normalized_password:
            raise HTTPException(
                status_code=422,
                detail=f"{file_label}已加密，请输入文件密码后重试",
            ) from plain_exc

        decrypted = io.BytesIO()
        try:
            office_file.load_key(password=normalized_password, verify_password=True)
            office_file.decrypt(decrypted)
        except InvalidKeyError as exc:
            raise HTTPException(
                status_code=422,
                detail=f"{file_label}密码不正确，无法解密",
            ) from exc
        except DecryptionError as exc:
            raise HTTPException(
                status_code=422,
                detail=f"{file_label}无法解密，请确认密码和文件格式",
            ) from exc

        decrypted.seek(0)
        try:
            return load_workbook(decrypted, data_only=True)
        except (BadZipFile, InvalidFileException, OSError) as exc:
            raise _invalid_excel_error(file_label) from exc
