"""邮局订报数据生成模块 · 来源文件解析（来源A 订阅明细 / 来源B 读者统计）。

**表头已按 7月/8月 黄金样本锁定**（按表头名识别、不靠文件名，文档 §5）：

* 来源A《…订阅明细~陈海影.xls/.xlsx》—— 17 列，**全量纳入**明细。
* 来源B《读者统计表~M月.xlsx/.csv》—— 标题在第1行、表头在第2行；**仅取 物流平台=邮局
  且 起投日期月份=本批月份** 的行。CSV **必须 UTF-8/BOM**，其它编码拒绝（不猜）。

明细 = 全部A（顺序在前）+ 筛后B（追加），按 姓名+电话 去重（见 import_service）。
"""

import csv
import io
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

# --- 锁定表头（field -> 可接受列名） ---------------------------------------
SOURCE_A_HEADERS: Dict[str, List[str]] = {
    "name": ["姓名"],
    "phone": ["联系电话", "电话"],
    "address": ["详细地址", "地址"],
    "addr_alt": ["省"],   # 样本把完整地址写进「省」列、详细地址列常为空 → 兜底取「省」
    "postal_code": ["邮编", "邮政编码"],
    "copies": ["份数"],
    "channel": ["渠道"],
    "remittance_name": ["汇款名称"],
    "remittance_date": ["汇款日期"],
}
SOURCE_A_REQUIRED = ["name", "copies"]

SOURCE_B_HEADERS: Dict[str, List[str]] = {
    "name": ["姓名"],
    "phone": ["电话", "联系电话"],
    "address": ["地址", "详细地址"],
    "postal_code": ["邮政编码", "邮编"],
    "copies": ["份数"],
    "channel": ["订单平台", "来源平台"],
    "logistics": ["物流平台"],
    "start_date": ["起投日期"],
    "status": ["状态"],
}
SOURCE_B_REQUIRED = ["name", "address", "logistics", "start_date"]

# 省份 → 地区短名（与黄金样本文件名一致）。
REGION_SPECIAL = {
    "内蒙古自治区": "内蒙",
    "广西壮族自治区": "广西",
    "宁夏回族自治区": "宁夏",
    "新疆维吾尔自治区": "新疆",
    "西藏自治区": "西藏",
}


def province_to_region(province: Optional[str]) -> Optional[str]:
    if not province:
        return None
    p = province.strip()
    if p in REGION_SPECIAL:
        return REGION_SPECIAL[p]
    for suf in ("自治区", "省", "市"):
        if p.endswith(suf) and len(p) > len(suf):
            return p[: -len(suf)]
    return p


@dataclass
class ParsedRow:
    source_file_role: str
    source_row: int
    name: str = ""
    phone: str = ""
    address: str = ""
    postal_code: str = ""
    copies: Optional[int] = None
    channel: str = ""
    remittance_name: str = ""
    remittance_date: str = ""


@dataclass
class ParseIssue:
    level: str
    source: str
    sheet_or_file: str = ""
    row_no: Optional[int] = None
    field_name: str = ""
    code: str = ""
    message: str = ""


@dataclass
class ParseResult:
    rows: List[ParsedRow] = field(default_factory=list)
    issues: List[ParseIssue] = field(default_factory=list)


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _to_int(v) -> Optional[int]:
    s = _norm(v)
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _read_rows(content: bytes, filename: str) -> Tuple[str, List[list], Optional[str]]:
    """把 .xls/.xlsx/.csv 读成 (sheet_name, rows, error)。"""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        if content.startswith(b"\xef\xbb\xbf"):
            text = content.decode("utf-8-sig")
        else:
            try:
                text = content.decode("utf-8")
            except UnicodeDecodeError:
                return "csv", [], "CSV 编码无法识别（需 UTF-8/带 BOM），请转换后重传"
        return "csv", [list(r) for r in csv.reader(io.StringIO(text))], None
    if ext == "xls":
        import xlrd  # 仅旧格式需要
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)

        def _xls_cell(r, c):
            v = sh.cell_value(r, c)
            # xlrd 把所有数字读成 float；整数值还原成 int（否则电话/邮编带 .0）。
            if isinstance(v, float) and v.is_integer():
                return int(v)
            return v

        rows = [[_xls_cell(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        return sh.name, rows, None
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    return ws.title, [list(r) for r in ws.iter_rows(values_only=True)], None


def _header_map(header_row, config: Dict[str, List[str]]) -> Dict[str, int]:
    present: Dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        name = _norm(raw)
        if name and name not in present:
            present[name] = idx
    field_col: Dict[str, int] = {}
    for fld, names in config.items():
        for cand in names:
            if cand in present:
                field_col[fld] = present[cand]
                break
    return field_col


def _find_header_row(rows: List[list], config: Dict[str, List[str]], required: List[str], max_scan: int = 5):
    """在前几行里找表头（来源B 表头在第2行）。返回 (header_index, header_map) 或 (None, None)。"""
    for i in range(min(max_scan, len(rows))):
        hm = _header_map(rows[i], config)
        if all(f in hm for f in required):
            return i, hm
    return None, None


def _get(row, hm, fld) -> str:
    idx = hm.get(fld)
    return _norm(row[idx]) if idx is not None and idx < len(row) else ""


def parse_source_a(content: bytes, filename: str) -> ParseResult:
    result = ParseResult()
    sheet, rows, err = _read_rows(content, filename)
    if err:
        result.issues.append(ParseIssue("block", "A", sheet, None, "", "encoding", err))
        return result
    if not rows:
        result.issues.append(ParseIssue("block", "A", sheet, None, "", "empty", "来源A 无数据行"))
        return result
    hidx, hm = _find_header_row(rows, SOURCE_A_HEADERS, SOURCE_A_REQUIRED)
    if hm is None:
        result.issues.append(ParseIssue(
            "block", "A", sheet, 1, "", "missing_header",
            f"来源A 未找到含 {SOURCE_A_REQUIRED} 的表头（请核对样本表头）"))
        return result
    for i, row in enumerate(rows[hidx + 1:], start=hidx + 2):
        if not row or all(_norm(c) == "" for c in row):
            continue
        name = _get(row, hm, "name")
        if not name:
            continue  # 无姓名 = 非订户记录（空行/占位），跳过
        result.rows.append(ParsedRow(
            source_file_role="A", source_row=i,
            name=name, phone=_get(row, hm, "phone"),
            address=_get(row, hm, "address") or _get(row, hm, "addr_alt"),
            postal_code=_get(row, hm, "postal_code"),
            copies=_to_int(row[hm["copies"]]) if hm["copies"] < len(row) else None,
            channel=_get(row, hm, "channel"),
            remittance_name=_get(row, hm, "remittance_name"),
            remittance_date=_get(row, hm, "remittance_date"),
        ))
    return result


def _month_of(start_raw: str) -> Optional[int]:
    """来源B 起投日期 'YYYY/M/D' → month。"""
    s = start_raw.strip().replace("-", "/")
    parts = [p for p in s.split("/") if p]
    if len(parts) >= 2:
        try:
            return int(parts[1])
        except ValueError:
            return None
    return None


def _year_of(start_raw: str) -> Optional[int]:
    s = start_raw.strip().replace("-", "/")
    parts = [p for p in s.split("/") if p]
    if parts:
        try:
            return int(parts[0])
        except ValueError:
            return None
    return None


def parse_source_b(content: bytes, filename: str, year: int, month: int) -> ParseResult:
    """来源B → 仅 物流平台=邮局 且 起投日期(年,月)=本批 的行。"""
    result = ParseResult()
    sheet, rows, err = _read_rows(content, filename)
    if err:
        result.issues.append(ParseIssue("block", "B", sheet, None, "", "encoding", err))
        return result
    if not rows:
        result.issues.append(ParseIssue("warn", "B", sheet, None, "", "empty", "来源B 无数据行"))
        return result
    hidx, hm = _find_header_row(rows, SOURCE_B_HEADERS, SOURCE_B_REQUIRED)
    if hm is None:
        result.issues.append(ParseIssue(
            "warn", "B", sheet, None, "", "missing_header",
            f"来源B 未找到含 {SOURCE_B_REQUIRED} 的表头，跳过来源B"))
        return result
    for i, row in enumerate(rows[hidx + 1:], start=hidx + 2):
        if not row or all(_norm(c) == "" for c in row):
            continue
        logistics = _get(row, hm, "logistics")
        if logistics != "邮局":
            continue
        start_raw = _get(row, hm, "start_date")
        if _month_of(start_raw) != month or _year_of(start_raw) != year:
            continue
        name = _get(row, hm, "name")
        if not name:
            continue  # 无姓名 = 空行/占位，跳过
        result.rows.append(ParsedRow(
            source_file_role="B", source_row=i,
            name=name, phone=_get(row, hm, "phone"),
            address=_get(row, hm, "address"), postal_code=_get(row, hm, "postal_code"),
            copies=_to_int(row[hm["copies"]]) if hm["copies"] < len(row) else None,
            channel=_get(row, hm, "channel"),
            remittance_name="未到", remittance_date="",
        ))
    return result
