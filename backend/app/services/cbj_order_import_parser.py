"""Parse a CBJ 小程序 order-export workbook into clean structured rows.

The export is a single sheet: a header row then one row per platform order, with
a multi-line 产品名称 cell. This module ONLY parses the messy Excel into typed
rows — resolving products, mapping status, computing coverage, dedup and order
creation are the import service's job (Phase 3b-3).

Known columns (matched by header name, tolerant of column reordering):
订单号 / 产品名称 / 数量 / 原价 / 付款金额 / 支付方式 / 发票 / 地址 / 备注 /
下单时间 / 支付时间 / 订单状态
"""

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import List, Optional

import openpyxl


_HEADER_MAP = {
    "订单号": "external_order_no",
    "产品名称": "product",
    "数量": "quantity",
    "原价": "original_amount",
    "付款金额": "paid_amount",
    "支付方式": "payment_method",
    "发票": "invoice",
    "地址": "address",
    "备注": "notes",
    "下单时间": "order_time",
    "支付时间": "payment_time",
    "订单状态": "status",
}
_REQUIRED = {"external_order_no", "product", "paid_amount", "address", "status"}

# name X<qty>,单价:<price>  (X / x / ×; half or full-width comma/colon)
_PRODUCT_RE = re.compile(
    r"^(?P<name>.*?)\s*[xX×]\s*(?P<qty>\d+)\s*[,，]?\s*单价\s*[:：]\s*(?P<price>[\d.]+)"
)


@dataclass
class ProductLine:
    raw: str
    name: str
    quantity: int
    unit_price: Decimal
    is_shipping: bool      # 运费补拍 line — not a publication item
    mentions_zto: bool     # contains 中通 → signal to flip delivery to 中通


@dataclass
class ParsedOrder:
    external_order_no: str
    status_raw: str
    paid_amount: Decimal
    original_amount: Decimal
    order_date: Optional[date]
    payment_time: Optional[datetime]
    payment_method_raw: str
    invoice_raw: str
    recipient_name: str
    recipient_phone: str
    recipient_address: str
    recipient_postal_code: Optional[str]
    notes: str
    product_lines: List[ProductLine] = field(default_factory=list)


def _dec(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _to_datetime(value) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    # "2026-06-01 13:37:10.0" → take the second-precision prefix.
    m = re.match(r"(\d{4}-\d{2}-\d{2})[ T](\d{2}:\d{2}:\d{2})", text)
    if m:
        return datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y-%m-%d %H:%M:%S")
    m = re.match(r"(\d{4}-\d{2}-\d{2})", text)
    if m:
        return datetime.strptime(m.group(1), "%Y-%m-%d")
    return None


def parse_product_field(value) -> List[ProductLine]:
    """Split the multi-line 产品名称 cell into product lines.

    Drops X0 promo placeholders (qty 0, not a shipping line). Shipping-surcharge
    lines (运费补拍…) are kept but flagged. Unparseable segments are kept verbatim
    (qty 1, price 0) so the import can surface them rather than silently lose them.
    """
    lines: List[ProductLine] = []
    for segment in str(value or "").split("\n"):
        seg = segment.strip()
        if not seg:
            continue
        match = _PRODUCT_RE.match(seg)
        if match:
            name = match.group("name").strip()
            qty = int(match.group("qty"))
            price = _dec(match.group("price"))
        else:
            name, qty, price = seg, 1, Decimal("0")
        is_shipping = "运费" in name
        mentions_zto = "中通" in name
        if qty == 0 and not is_shipping:
            continue  # X0 promo placeholder — not purchased
        lines.append(
            ProductLine(
                raw=seg,
                name=name,
                quantity=qty,
                unit_price=price,
                is_shipping=is_shipping,
                mentions_zto=mentions_zto,
            )
        )
    return lines


def parse_address(value):
    """Split '姓名,电话,详细地址[,邮编]' (half- or full-width commas)."""
    parts = re.split(r"[,，]", str(value or ""), maxsplit=2)
    name = parts[0].strip() if len(parts) > 0 else ""
    phone = parts[1].strip() if len(parts) > 1 else ""
    rest = parts[2].strip() if len(parts) > 2 else ""
    postal = None
    tail = re.search(r"[,，]\s*(\d{3,6})\s*$", rest)
    if tail:
        postal = tail.group(1)
        rest = rest[: tail.start()].strip()
    return name, phone, rest, postal


def _find_header(ws):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        cells = ["" if v is None else str(v).strip() for v in row]
        if "订单号" in cells and "产品名称" in cells:
            index = {}
            for col_idx, cell in enumerate(cells):
                if cell in _HEADER_MAP:
                    index[_HEADER_MAP[cell]] = col_idx
            return row_idx, index
    return None, None


def parse_cbj_orders(file_bytes: bytes) -> List[ParsedOrder]:
    """Parse the CBJ export bytes into ParsedOrder rows.

    Raises ValueError if the file is not a recognizable CBJ order export.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface a clean message to the API
        raise ValueError("无法读取 Excel 文件") from exc

    ws = wb[wb.sheetnames[0]]
    header_row, index = _find_header(ws)
    if header_row is None:
        raise ValueError("未找到 CBJ 订单表头（需要「订单号」「产品名称」等列）")
    missing = _REQUIRED - set(index)
    if missing:
        labels = {v: k for k, v in _HEADER_MAP.items()}
        raise ValueError("缺少必需列：" + "、".join(labels[m] for m in missing))

    def cell(row, key):
        col = index.get(key)
        return row[col] if col is not None and col < len(row) else None

    orders: List[ParsedOrder] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        external = cell(row, "external_order_no")
        if external is None or str(external).strip() == "":
            continue  # blank / trailing row
        name, phone, addr, postal = parse_address(cell(row, "address"))
        orders.append(
            ParsedOrder(
                external_order_no=str(external).strip(),
                status_raw=str(cell(row, "status") or "").strip(),
                paid_amount=_dec(cell(row, "paid_amount")),
                original_amount=_dec(cell(row, "original_amount")),
                order_date=(_to_datetime(cell(row, "order_time")) or datetime.min).date()
                if cell(row, "order_time")
                else None,
                payment_time=_to_datetime(cell(row, "payment_time")),
                payment_method_raw=str(cell(row, "payment_method") or "").strip(),
                invoice_raw=str(cell(row, "invoice") or "").strip(),
                recipient_name=name,
                recipient_phone=phone,
                recipient_address=addr,
                recipient_postal_code=postal,
                notes=str(cell(row, "notes") or "").strip(),
                product_lines=parse_product_field(cell(row, "product")),
            )
        )
    return orders
