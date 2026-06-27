import io
import os
from datetime import date, timedelta
from typing import Optional

from openpyxl import load_workbook, Workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session
from app.models import Issue, Order, OrderStatus, ReportEntry, ShippingDetail
from app.models.shipping_detail import ShippingDetailSyncStatus
from app.services.issue_service import format_chinese_issue_number, get_year_issue_index

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "templates")

SHIPPING_DETAIL_EXPORT_COLUMNS: list[tuple[str, str | None]] = [
    ("序号", None),
    ("期号", "issue_number"),
    ("原工作表", "sheet_name"),
    ("渠道", "channel"),
    ("子渠道", "sub_channel"),
    ("签约公司", "company"),
    ("姓名", "name"),
    ("电话", "phone"),
    ("地址", "address"),
    ("份数", "quantity"),
    ("频率", "frequency"),
    ("运输方式", "transport"),
    ("发货时间", "shipped_at"),
    ("截止日期", "deadline"),
    ("状态", "status"),
    ("备注", "notes"),
    ("附加信息", "extra_info"),
    ("站点", "station_name"),
    ("站厅", "station_hall"),
    ("联系人", "contact_person"),
    ("高铁序号", "seq_number"),
    ("期数", "period_count"),
    ("信息确认", "confirmation"),
]

# Mapping: (category, sub_category) → list of (sheet_name, cell) to write current value
CELL_MAPPING: dict[tuple[str, str], list[tuple[str, str]]] = {
    # 邮发 → 人民日报印厂`
    ("postal", "本市"): [("人民日报印厂`", "C8")],
    ("postal", "外埠"): [("人民日报印厂`", "C9")],
    # 零售 → 零售渠道`
    ("retail", "东部"): [("零售渠道`", "B4")],
    ("retail", "西部"): [("零售渠道`", "B5")],
    ("guangzhou", "零售"): [("零售渠道`", "B6")],
    # 订阅 → 订阅渠道`
    ("guangzhou", "订阅"): [("订阅渠道`", "B6")],
    ("guotumao", "国图贸"): [("订阅渠道`", "B5")],
    ("chengdu", "成都杂志铺"): [("订阅渠道`", "B7")],
    # 收发室部门 → 收发室自留分发（需打印）B column
    ("social_use", "营报传媒_收发室"): [("收发室自留分发（需打印）", "B5")],
    ("social_use", "中经传媒智库"): [("收发室自留分发（需打印）", "B6")],
    ("social_use", "新闻中心"): [("收发室自留分发（需打印）", "B7")],
    ("social_use", "行政"): [("收发室自留分发（需打印）", "B8")],
    ("social_use", "财经中心"): [("收发室自留分发（需打印）", "B9")],
    ("social_use", "产经中心"): [("收发室自留分发（需打印）", "B10")],
    ("social_use", "出版中心"): [("收发室自留分发（需打印）", "B11")],
    ("social_use", "品牌中心"): [("收发室自留分发（需打印）", "B12")],
    ("social_use", "经营网"): [("收发室自留分发（需打印）", "B13")],
    ("social_use", "法务"): [("收发室自留分发（需打印）", "B14")],
    ("social_use", "社科院、工经所"): [("收发室自留分发（需打印）", "B15")],
    ("social_use", "财务"): [("收发室自留分发（需打印）", "B16")],
    ("social_use", "库房"): [("收发室自留分发（需打印）", "B17")],
    # 社用报` direct values (I column and B column non-formula cells)
    ("social_use", "营报传媒_读者"): [("社用报`", "I5")],
    ("social_use", "营报传媒_备用报"): [("社用报`", "I6")],
    ("social_use", "营报传媒_上犹"): [("社用报`", "I18")],
    ("social_use", "高铁展示"): [("社用报`", "I19")],
    ("social_use", "上海站用"): [("社用报`", "B19")],
    ("social_use", "广东站用"): [("社用报`", "B20")],
    ("social_use", "成都站用"): [("社用报`", "B21")],
    ("social_use", "西安站用"): [("社用报`", "B22")],
    # 临时加印: split into 自留 (收发室) and 快递 (社用报) — handled specially in export
    # ("social_use", "临时加印") — 快递部分 = 总数 - 自留, written to 社用报` B27
    # ("social_use", "临时加印_自留") — written to 收发室自留分发（需打印） B18
    ("social_use", "临时加印_自留"): [("收发室自留分发（需打印）", "B18")],
    # 合订本 → 3 locations (社用报` B17, 北京印厂 C11, 人民日报印厂` I14)
    ("binding", "合订本（印厂留存）"): [
        ("社用报`", "B17"),
        ("北京印厂", "C11"),
        ("人民日报印厂`", "I14"),
    ],
}

# Mapping for "上期" (previous issue) columns.
# (category, sub_category) → list of (sheet_name, cell) for previous issue value
PREV_CELL_MAPPING: dict[tuple[str, str], list[tuple[str, str]]] = {
    # 人民日报印厂` D column (direct items)
    ("postal", "本市"): [("人民日报印厂`", "D8")],
    ("postal", "外埠"): [("人民日报印厂`", "D9")],
    # 零售渠道` C column
    ("retail", "东部"): [("零售渠道`", "C4")],
    ("retail", "西部"): [("零售渠道`", "C5")],
    ("guangzhou", "零售"): [("零售渠道`", "C6")],
    # 订阅渠道` C column
    ("guotumao", "国图贸"): [("订阅渠道`", "C5")],
    ("guangzhou", "订阅"): [("订阅渠道`", "C6")],
    ("chengdu", "成都杂志铺"): [("订阅渠道`", "C7")],
    # 社用报` C column
    ("social_use", "营报传媒_读者"): [("社用报`", "C4")],
    # C4 is 营报传媒 aggregate — handled specially below
    ("social_use", "中经传媒智库"): [("社用报`", "C5")],
    ("social_use", "新闻中心"): [("社用报`", "C6")],
    ("social_use", "行政"): [("社用报`", "C7")],
    ("social_use", "财经中心"): [("社用报`", "C8")],
    ("social_use", "产经中心"): [("社用报`", "C9")],
    ("social_use", "出版中心"): [("社用报`", "C10")],
    ("social_use", "品牌中心"): [("社用报`", "C11")],
    ("social_use", "经营网"): [("社用报`", "C12")],
    ("social_use", "法务"): [("社用报`", "C13")],
    ("social_use", "社科院、工经所"): [("社用报`", "C14")],
    ("social_use", "财务"): [("社用报`", "C15")],
    ("social_use", "库房"): [("社用报`", "C16")],
    ("binding", "合订本（印厂留存）"): [("社用报`", "C17")],
    ("social_use", "营报传媒_上犹"): [("社用报`", "C18")],
    # C18 is 报社订阅/展示 aggregate — handled specially below
    ("social_use", "上海站用"): [("社用报`", "C19")],
    ("social_use", "广东站用"): [("社用报`", "C20")],
    ("social_use", "成都站用"): [("社用报`", "C21")],
    ("social_use", "西安站用"): [("社用报`", "C22")],

    # 临时加印 快递部分 handled specially
    ("social_use", "临时加印_自留"): [("收发室自留分发（需打印）", "C18")],
}


def _get_template_path(filename: str) -> str:
    path = os.path.join(TEMPLATE_DIR, filename)
    if os.path.exists(path):
        return path
    return None


def _get_prev_issue(issue: Issue, db: Session) -> Optional[Issue]:
    """Find the previous issue by issue_number."""
    return (
        db.query(Issue)
        .filter(Issue.issue_number < issue.issue_number)
        .order_by(Issue.issue_number.desc())
        .first()
    )


def _build_entry_map(entries: list[ReportEntry]) -> dict[tuple[str, str], int]:
    """Build a lookup dict from (category, sub_category) → value."""
    return {(e.category, e.sub_category): e.value for e in entries}


def _fill_prev_issue_aggregates(
    wb, prev_entries: dict[tuple[str, str], int]
) -> None:
    """Fill computed aggregate cells in 人民日报印厂` D column for previous issue."""
    ws = wb["人民日报印厂`"]
    get = prev_entries.get

    # D10: 上期北京零售 = 东部 + 西部
    ws["D10"] = get(("retail", "东部"), 0) + get(("retail", "西部"), 0)
    # D11: 上期外埠零售 = 广州零售
    ws["D11"] = get(("guangzhou", "零售"), 0)
    # D12: 上期驻外 = 上海站 + 广东站 + 西安站 + 成都站
    ws["D12"] = (
        get(("social_use", "上海站用"), 0)
        + get(("social_use", "广东站用"), 0)
        + get(("social_use", "西安站用"), 0)
        + get(("social_use", "成都站用"), 0)
    )
    # D13: 上期发行商订阅 = 国图贸 + 广州订阅 + 杂志铺
    ws["D13"] = (
        get(("guotumao", "国图贸"), 0)
        + get(("guangzhou", "订阅"), 0)
        + get(("chengdu", "成都杂志铺"), 0)
    )
    # D14: 上期报社订阅/展示 = 上犹 + 高铁展示
    ws["D14"] = (
        get(("social_use", "营报传媒_上犹"), 0)
        + get(("social_use", "高铁展示"), 0)
    )
    # D15: 上期社内用 = 收发室合计 + 读者 + 备用 + 印厂留存
    dept_keys = [
        ("social_use", "营报传媒_收发室"),
        ("social_use", "中经传媒智库"),
        ("social_use", "新闻中心"),
        ("social_use", "行政"),
        ("social_use", "财经中心"),
        ("social_use", "产经中心"),
        ("social_use", "出版中心"),
        ("social_use", "品牌中心"),
        ("social_use", "经营网"),
        ("social_use", "法务"),
        ("social_use", "社科院、工经所"),
        ("social_use", "财务"),
        ("social_use", "库房"),
    ]
    shoufashi_total = sum(get(k, 0) for k in dept_keys)
    ws["D15"] = (
        shoufashi_total
        + get(("social_use", "营报传媒_读者"), 0)
        + get(("social_use", "营报传媒_备用报"), 0)
        + get(("binding", "合订本（印厂留存）"), 0)
    )
    # D16: 上期临时加印（快递部分）
    prev_temp_total_d16 = get(("social_use", "临时加印"), 0)
    prev_temp_self_d16 = get(("social_use", "临时加印_自留"), 0)
    ws["D16"] = prev_temp_total_d16 - prev_temp_self_d16
    # D4: 上期北京印厂印数 (total of all above)
    ws["D4"] = sum(
        (ws[f"D{r}"].value or 0) for r in range(8, 17)
    )

    # 社用报` C4: 上期营报传媒 = 收发室 + 读者 + 备用
    ws_sy = wb["社用报`"]
    ws_sy["C4"] = (
        get(("social_use", "营报传媒_收发室"), 0)
        + get(("social_use", "营报传媒_读者"), 0)
        + get(("social_use", "营报传媒_备用报"), 0)
    )
    # C18: 上期报社订阅/展示 = 上犹 + 高铁展示
    ws_sy["C18"] = (
        get(("social_use", "营报传媒_上犹"), 0)
        + get(("social_use", "高铁展示"), 0)
    )


def export_report_excel(issue_id: int, db: Session) -> io.BytesIO:
    """Generate the 报数 Excel file for a given issue.

    Uses the decrypted original Excel as a template, preserving all
    formatting, merged cells, formulas, and cross-sheet references.
    Only writes to source data cells; formulas compute the rest.
    """
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise ValueError("Issue not found")

    template_path = _get_template_path("report_template.xlsx")
    if not template_path:
        raise FileNotFoundError("report_template.xlsx not found in templates/")

    wb = load_workbook(template_path)

    # --- Fill current issue data ---
    entries = db.query(ReportEntry).filter(ReportEntry.issue_id == issue_id).all()
    entry_map = _build_entry_map(entries)

    for key, cells in CELL_MAPPING.items():
        if key in entry_map:
            value = entry_map[key]
            for sheet_name, cell_ref in cells:
                if sheet_name in wb.sheetnames:
                    wb[sheet_name][cell_ref] = value

    # 临时加印 快递部分 = 总数 - 自留
    temp_total = entry_map.get(("social_use", "临时加印"), 0)
    temp_self = entry_map.get(("social_use", "临时加印_自留"), 0)
    temp_express = temp_total - temp_self
    if "社用报`" in wb.sheetnames:
        wb["社用报`"]["B27"] = temp_express

    # --- Fill previous issue data (上期 columns) ---
    prev_issue = _get_prev_issue(issue, db)
    if prev_issue:
        prev_entries = db.query(ReportEntry).filter(
            ReportEntry.issue_id == prev_issue.id
        ).all()
        prev_map = _build_entry_map(prev_entries)

        for key, cells in PREV_CELL_MAPPING.items():
            if key in prev_map:
                value = prev_map[key]
                for sheet_name, cell_ref in cells:
                    if sheet_name in wb.sheetnames:
                        wb[sheet_name][cell_ref] = value

        # 上期临时加印 快递部分
        prev_temp_total = prev_map.get(("social_use", "临时加印"), 0)
        prev_temp_self = prev_map.get(("social_use", "临时加印_自留"), 0)
        prev_temp_express = prev_temp_total - prev_temp_self
        if "社用报`" in wb.sheetnames:
            wb["社用报`"]["C27"] = prev_temp_express

        _fill_prev_issue_aggregates(wb, prev_map)

    # --- Update header in 北京印厂 ---
    d = issue.publish_date
    page_count = issue.page_count if issue.page_count else 24
    year_issue_label = format_chinese_issue_number(get_year_issue_index(db, issue))
    year_issue_part = f" 第{year_issue_label}期" if year_issue_label else ""
    header = f"期数：{issue.issue_number}{year_issue_part}   版数：{page_count}    出版日期：{d.year}年{d.month}月{d.day}日"
    wb["北京印厂"]["A3"] = header
    # 制表时间 = 出版日期的上一周周五 (weekday: Mon=0 ... Fri=4)
    days_since_friday = (d.weekday() - 4) % 7 or 7
    report_date = d - timedelta(days=days_since_friday)
    wb["北京印厂"]["D15"] = f"制表时间：{report_date.year}年{report_date.month}月{report_date.day}日"

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_shipping_excel(issue_id: int, db: Session) -> io.BytesIO:
    """Generate the ZTO-MF Excel file for a given issue."""
    issue = db.query(Issue).filter(Issue.id == issue_id).first()
    if not issue:
        raise ValueError("Issue not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "ZTO-MF"

    # 排除已作废订单产生的发货行：order_generated 行在订单作废时被置 orphaned，
    # 且兜底排除任何 link 到 void 订单的行（覆盖本次修复前历史遗留的孤儿行）。
    details = (
        db.query(ShippingDetail)
        .outerjoin(Order, ShippingDetail.order_id == Order.id)
        .filter(ShippingDetail.issue_number == issue.issue_number)
        .filter(ShippingDetail.sync_status != ShippingDetailSyncStatus.orphaned)
        .filter(
            or_(ShippingDetail.order_id.is_(None), Order.status != OrderStatus.void)
        )
        .order_by(ShippingDetail.id)
        .all()
    )

    for col, (header, _) in enumerate(SHIPPING_DETAIL_EXPORT_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_index, detail in enumerate(details, start=2):
        for col, (_, field_name) in enumerate(SHIPPING_DETAIL_EXPORT_COLUMNS, start=1):
            if field_name is None:
                value = row_index - 1
            else:
                value = getattr(detail, field_name)
                if field_name == "shipped_at" and value:
                    value = value.strftime("%Y-%m-%d")
                if value is None:
                    value = ""
            ws.cell(row=row_index, column=col, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_report_filename(issue: Issue) -> str:
    year = issue.publish_date.year
    return f"{year}年《中国经营报》（总第{issue.issue_number}期）报数.xlsx"


def get_shipping_filename(issue: Issue) -> str:
    d = issue.publish_date
    return f"{d.year}年{d.month}月{d.day}日《中国经营报》中通快递发货明细（{issue.issue_number}）.xlsx"
