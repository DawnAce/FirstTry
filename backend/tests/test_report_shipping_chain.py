import asyncio
import io
import unittest
from datetime import date

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.api.exports import export_all, export_report, export_shipping
from app.api.reports import confirm_report, get_report
from app.database import Base
from app.models import Issue, IssueAuditSnapshot, IssueStatus, PublicationSchedule, ReportEntry, ShippingDetail, User, UserRole


def _admin_user() -> User:
    return User(id=1, username="admin", role=UserRole.admin, password_hash="x")


class ReportShippingChainTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(bind=self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

    def _read_streaming_response_bytes(self, response) -> bytes:
        async def _collect_body() -> bytes:
            chunks = []
            async for chunk in response.body_iterator:
                chunks.append(chunk)
            return b"".join(chunks)

        return asyncio.run(_collect_body())

    def test_issue_audit_snapshot_model_is_wired_to_issue(self):
        from app.models import IssueAuditSnapshot

        self.assertEqual(IssueAuditSnapshot.__tablename__, "issue_audit_snapshots")
        self.assertTrue(hasattr(Issue, "audit_snapshots"))

    def test_confirm_uses_shipping_details_total_and_returns_snapshot_and_drift_totals(self):
        db = self.SessionLocal()
        issue = Issue(issue_number=3001, publish_date=date(2026, 5, 25), status=IssueStatus.draft)
        db.add(issue)
        db.flush()
        db.add_all(
            [
                ReportEntry(issue_id=issue.id, category="postal", sub_category="本市", value=10),
                ReportEntry(issue_id=issue.id, category="social_use", sub_category="营报传媒_读者", value=40),
                ShippingDetail(issue_number=3001, sheet_name="测试", channel="渠道订阅", name="甲", quantity=15),
                ShippingDetail(issue_number=3001, sheet_name="测试", channel="渠道订阅", name="乙", quantity=25),
            ]
        )
        db.commit()

        # Step 1: Confirm the report with shipping_total = 40
        result = confirm_report(
            issue.id,
            db=db,
            user=User(id=1, username="admin", role=UserRole.admin, password_hash="x"),
        )

        self.assertEqual(result["zt_report_total"], 40)
        self.assertEqual(result["zt_shipping_total"], 40)

        # Step 2: Mutate shipping_details after confirmation
        shipping_details = db.query(ShippingDetail).filter_by(issue_number=3001).all()
        shipping_details[0].quantity = 100  # Change 甲 from 15 to 100
        db.commit()

        # Step 3: Get report and verify confirmation_summary separates confirmed snapshot totals
        # from current drift totals after the live rows have changed.
        report = get_report(issue.id, db=db)
        self.assertEqual(report.confirmation_summary.confirmed_report_total, 40)
        self.assertEqual(report.confirmation_summary.confirmed_shipping_total, 40)
        self.assertEqual(report.confirmation_summary.confirmed_delta, 0)
        self.assertEqual(report.confirmation_summary.confirmed_is_match, True)
        self.assertEqual(report.confirmation_summary.current_shipping_total, 125)
        self.assertEqual(report.confirmation_summary.current_delta, -85)
        self.assertEqual(report.confirmation_summary.current_is_match, False)
        self.assertEqual(report.confirmation_summary.has_shipping_drift, True)

    def test_confirm_report_persists_issue_audit_snapshot_row(self):
        db = self.SessionLocal()
        issue = Issue(issue_number=3010, publish_date=date(2026, 7, 20), status=IssueStatus.draft)
        db.add(issue)
        db.flush()
        db.add_all(
            [
                ReportEntry(issue_id=issue.id, category="social_use", sub_category="营报传媒_读者", value=22),
                ShippingDetail(issue_number=3010, sheet_name="测试", channel="渠道订阅", name="甲", quantity=10),
            ]
        )
        db.commit()

        result = confirm_report(
            issue.id,
            db=db,
            user=User(id=1, username="admin", role=UserRole.admin, password_hash="x"),
        )

        self.assertIn("warning", result)
        snapshot = db.query(IssueAuditSnapshot).filter_by(issue_id=issue.id, snapshot_type="confirm").one()
        self.assertEqual(snapshot.report_total, 22)
        self.assertEqual(snapshot.shipping_total, 10)
        self.assertEqual(snapshot.delta, 12)
        self.assertEqual(snapshot.is_match, False)

    def test_shipping_export_writes_single_full_detail_sheet(self):
        db = self.SessionLocal()
        issue = Issue(issue_number=3002, publish_date=date(2026, 6, 1), status=IssueStatus.confirmed)
        db.add(issue)
        db.flush()
        db.add_all(
            [
                ShippingDetail(
                    issue_number=3002,
                    sheet_name="原始表",
                    channel="渠道订阅",
                    sub_channel="监管",
                    transport="中通物流",
                    frequency="周",
                    status="正常",
                    name="甲",
                    address="北京市朝阳区测试路1号",
                    phone="13800000000",
                    quantity=7,
                    deadline="长期",
                    notes="备注",
                    extra_info="附加信息",
                    station_name="北京站",
                    station_hall="A厅",
                    contact_person="联系人甲",
                    seq_number=12,
                    period_count=3,
                    confirmation="已确认",
                    company="测试公司",
                ),
                ShippingDetail(issue_number=3002, sheet_name="测试", channel="对公订阅", name="乙", quantity=9),
            ]
        )
        db.commit()

        response = export_shipping(issue.id, db=db, user=_admin_user())

        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.media_type,
        )
        workbook = load_workbook(io.BytesIO(self._read_streaming_response_bytes(response)))
        self.assertEqual(workbook.sheetnames, ["ZTO-MF"])

        sheet = workbook["ZTO-MF"]
        headers = [sheet.cell(row=1, column=col).value for col in range(1, 24)]
        self.assertEqual(
            headers,
            [
                "序号",
                "期号",
                "原工作表",
                "渠道",
                "子渠道",
                "签约公司",
                "姓名",
                "电话",
                "地址",
                "份数",
                "频率",
                "运输方式",
                "发货时间",
                "截止日期",
                "状态",
                "备注",
                "附加信息",
                "站点",
                "站厅",
                "联系人",
                "高铁序号",
                "期数",
                "信息确认",
            ],
        )
        first_row = [sheet.cell(row=2, column=col).value for col in range(1, 24)]
        self.assertEqual(
            first_row,
            [
                1,
                3002,
                "原始表",
                "渠道订阅",
                "监管",
                "测试公司",
                "甲",
                "13800000000",
                "北京市朝阳区测试路1号",
                7,
                "周",
                "中通物流",
                None,
                "长期",
                "正常",
                "备注",
                "附加信息",
                "北京站",
                "A厅",
                "联系人甲",
                12,
                3,
                "已确认",
            ],
        )

    def test_shipping_export_persists_export_audit_snapshot(self):
        db = self.SessionLocal()
        issue = Issue(issue_number=3004, publish_date=date(2026, 6, 15), status=IssueStatus.confirmed)
        db.add(issue)
        db.flush()
        db.add_all(
            [
                ReportEntry(issue_id=issue.id, category="social_use", sub_category="营报传媒_读者", value=30),
                ShippingDetail(issue_number=3004, sheet_name="测试", channel="渠道订阅", name="甲", quantity=12),
                ShippingDetail(issue_number=3004, sheet_name="测试", channel="对公订阅", name="乙", quantity=8),
            ]
        )
        db.commit()

        response = export_shipping(issue.id, db=db, user=_admin_user())

        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.media_type,
        )
        snapshot = (
            db.query(IssueAuditSnapshot)
            .filter_by(issue_id=issue.id, snapshot_type="shipping_export")
            .one()
        )
        self.assertEqual(snapshot.report_total, 30)
        self.assertEqual(snapshot.shipping_total, 20)
        self.assertEqual(snapshot.delta, 10)
        self.assertEqual(snapshot.is_match, False)

    def test_report_export_persists_export_audit_snapshot(self):
        db = self.SessionLocal()
        issue = Issue(issue_number=3005, publish_date=date(2026, 6, 22), status=IssueStatus.confirmed)
        db.add(issue)
        db.flush()
        db.add_all(
            [
                ReportEntry(issue_id=issue.id, category="social_use", sub_category="营报传媒_读者", value=18),
                ShippingDetail(issue_number=3005, sheet_name="测试", channel="渠道订阅", name="甲", quantity=18),
            ]
        )
        db.commit()

        response = export_report(issue.id, db=db, user=_admin_user())

        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.media_type,
        )
        snapshot = (
            db.query(IssueAuditSnapshot)
            .filter_by(issue_id=issue.id, snapshot_type="report_export")
            .one()
        )
        self.assertEqual(snapshot.report_total, 18)
        self.assertEqual(snapshot.shipping_total, 18)
        self.assertEqual(snapshot.delta, 0)
        self.assertEqual(snapshot.is_match, True)

    def test_report_export_header_includes_annual_sequence_label(self):
        db = self.SessionLocal()
        schedule_rows = []
        for index in range(14):
            schedule_rows.append(
                PublicationSchedule(
                    year=2026,
                    issue_number=2635 + index,
                    publish_date=date.fromordinal(date(2026, 1, 5).toordinal() + index * 7),
                    is_suspended=False,
                )
            )
        issue = Issue(issue_number=2648, publish_date=schedule_rows[-1].publish_date, status=IssueStatus.confirmed)
        db.add_all(schedule_rows)
        db.add(issue)
        db.commit()

        response = export_report(issue.id, db=db, user=_admin_user())
        workbook = load_workbook(io.BytesIO(self._read_streaming_response_bytes(response)))

        self.assertIn("第十四期", workbook["北京印厂"]["A3"].value)
        db.close()

    def test_export_all_persists_both_export_audit_snapshots(self):
        db = self.SessionLocal()
        issue = Issue(issue_number=3006, publish_date=date(2026, 6, 29), status=IssueStatus.confirmed)
        db.add(issue)
        db.flush()
        db.add_all(
            [
                ReportEntry(issue_id=issue.id, category="social_use", sub_category="营报传媒_读者", value=26),
                ShippingDetail(issue_number=3006, sheet_name="测试", channel="渠道订阅", name="甲", quantity=20),
            ]
        )
        db.commit()

        response = export_all(issue.id, db=db, user=_admin_user())

        self.assertEqual(response.media_type, "application/zip")
        snapshot_types = {
            snapshot.snapshot_type
            for snapshot in db.query(IssueAuditSnapshot).filter_by(issue_id=issue.id).all()
        }
        self.assertIn("report_export", snapshot_types)
        self.assertIn("shipping_export", snapshot_types)

    def test_confirm_mismatch_returns_snapshot_and_current_drift_values(self):
        db = self.SessionLocal()
        issue = Issue(issue_number=3003, publish_date=date(2026, 6, 8), status=IssueStatus.draft)
        db.add(issue)
        db.flush()
        db.add_all(
            [
                ReportEntry(issue_id=issue.id, category="social_use", sub_category="营报传媒_读者", value=30),
                ShippingDetail(issue_number=3003, sheet_name="测试", channel="渠道订阅", name="甲", quantity=18),
            ]
        )
        db.commit()

        # Step 1: Confirm the report with mismatch (report=30, shipping=18, delta=12)
        result = confirm_report(
            issue.id,
            db=db,
            user=User(id=1, username="admin", role=UserRole.admin, password_hash="x"),
        )

        self.assertIn("warning", result)

        # Step 2: Mutate shipping_details after confirmation
        shipping_detail = db.query(ShippingDetail).filter_by(issue_number=3003).first()
        shipping_detail.quantity = 50  # Change from 18 to 50
        db.commit()

        # Step 3: Get report and verify confirmation_summary keeps the confirmed snapshot
        # while also surfacing the current live totals after drift.
        report = get_report(issue.id, db=db)
        self.assertEqual(report.confirmation_summary.confirmed_report_total, 30)
        self.assertEqual(report.confirmation_summary.confirmed_shipping_total, 18)
        self.assertEqual(report.confirmation_summary.confirmed_delta, 12)
        self.assertEqual(report.confirmation_summary.confirmed_is_match, False)
        self.assertEqual(report.confirmation_summary.current_shipping_total, 50)
        self.assertEqual(report.confirmation_summary.current_delta, -20)
        self.assertEqual(report.confirmation_summary.current_is_match, False)
        self.assertEqual(report.confirmation_summary.has_shipping_drift, True)


if __name__ == "__main__":
    unittest.main()
