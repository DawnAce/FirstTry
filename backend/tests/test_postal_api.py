"""邮局投递 API · HTTP 连通测试（Task 5）。

In-memory SQLite + TestClient，覆盖 get_db / get_current_user / require_admin，
与 test_products_api 同风格。
"""

import io
from types import SimpleNamespace

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.auth import get_current_user, require_admin
from app.database import Base, get_db
from app.main import app
from app.models import Partner, PartnerType

_HEADERS = [
    "编号", "地区", "姓名", "联系电话", "省", "市", "区", "详细地址", "邮编", "年度",
    "产品名称", "起月日", "止月日", "份数", "金额", "渠道", "汇款名称", "汇款日期",
    "投递单位", "赠阅/关联", "备注",
]

_ROWS = [
    {"编号": "4784", "姓名": "高占军", "联系电话": "13764491959", "省": "上海市",
     "市": "上海市", "区": "宝山区", "详细地址": "华灵路1900弄355号501室", "邮编": "201900",
     "年度": "2026年", "产品名称": "中国经营报", "起月日": "0101", "止月日": "0131",
     "份数": "1", "金额": "20", "渠道": "CBJ+小程序", "投递单位": "北京集订分送"},
    {"编号": "4837", "姓名": "孙琪", "联系电话": "18353360185", "省": "山东省",
     "市": "淄博市", "区": "临淄区", "详细地址": "齐兴路88号临淄农村商业银行",
     "年度": "2026年", "产品名称": "中国经营报", "起月日": "0101", "止月日": "0531",
     "份数": "50", "金额": "5000", "渠道": "对公转账", "投递单位": "山东集订分送"},
    {"编号": "4801", "姓名": "郑天敏", "联系电话": "18323045917", "省": "重庆市",
     "市": "重庆市", "区": "九龙坡区", "详细地址": "石坪桥街道骏逸新视界19栋25-6", "邮编": "400050",
     "年度": "2026年", "产品名称": "中国经营报", "起月日": "0101", "止月日": "0131",
     "份数": "1", "金额": "20", "渠道": "CBJ+小程序", "投递单位": ""},
]


def _workbook_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "邮局读者明细"
    ws.append(_HEADERS)
    for r in _ROWS:
        ws.append([r.get(h, "") for h in _HEADERS])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

    seed = TestingSessionLocal()
    for n in ("北京集订分送", "山东集订分送"):
        seed.add(Partner(name=n, partner_type=PartnerType.distribution))
    seed.commit()
    seed.close()

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    fake = SimpleNamespace(id=1, role="admin")
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[get_current_user] = lambda: fake
    app.dependency_overrides[require_admin] = lambda: fake

    c = TestClient(app)
    try:
        yield c
    finally:
        app.dependency_overrides.clear()


def test_full_flow(client):
    data = _workbook_bytes()

    # 1) 预览
    resp = client.post(
        "/api/postal/import/preview",
        files={"file": ("postal.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["counts"]["import"] == 3
    assert body["can_commit"] is True
    sid = body["session_id"]

    # 2) 提交
    resp = client.post("/api/postal/import/commit", json={"session_id": sid})
    assert resp.status_code == 200, resp.text
    assert resp.json()["created"] == 3

    # 3) 投递名册可查到导入的 3 条
    resp = client.get("/api/postal/deliveries")
    assert resp.status_code == 200
    assert resp.json()["total"] == 3


def test_reject_non_postal_upload(client):
    wb = openpyxl.Workbook()
    wb.active.append(["不相关", "表头"])
    bio = io.BytesIO()
    wb.save(bio)
    resp = client.post(
        "/api/postal/import/preview",
        files={"file": ("x.xlsx", bio.getvalue(), "application/octet-stream")},
    )
    assert resp.status_code == 400


def test_cannot_delete_in_use_distribution_unit(client):
    """回归：投递单位被邮局订单目标引用时，删除该 Partner 应 409（不触发 FK 500）。"""
    data = _workbook_bytes()
    r = client.post(
        "/api/postal/import/preview",
        files={"file": ("postal.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    client.post("/api/postal/import/commit", json={"session_id": r.json()["session_id"]})

    partners = client.get("/api/partners").json()
    pid = next(p["id"] for p in partners if p["name"] == "北京集订分送")
    resp = client.delete(f"/api/partners/{pid}")
    assert resp.status_code == 409


_COMPLAINT_HEADERS = ["接诉日期", "姓名", "联系电话", "省", "市", "区", "详细地址", "邮编",
                      "年度", "投诉情况", "处理情况", "回访", "处理次数", "编号", "投递渠道单位"]
_COMPLAINT_ROWS = [
    {"接诉日期": "2024-01-03", "姓名": "马宁", "省": "江苏省", "市": "徐州市", "区": "泉山区",
     "年度": "2024年", "投诉情况": "2024年1月1日第一期没有收到", "处理情况": "转徐州11185",
     "回访": "已收到", "处理次数": "1", "编号": "000680", "投递渠道单位": "北京集订分送"},
    {"接诉日期": "2024-02-01", "姓名": "李四", "省": "北京市", "市": "北京市", "区": "朝阳区",
     "年度": "2024年", "投诉情况": "2月没收到", "处理情况": "北京局", "处理次数": "2", "编号": "888888"},
]


def _complaint_wb() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "邮局年投诉"
    ws.append(_COMPLAINT_HEADERS)
    for r in _COMPLAINT_ROWS:
        ws.append([r.get(h, "") for h in _COMPLAINT_HEADERS])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_complaint_import_and_list(client):
    data = _complaint_wb()
    r = client.post(
        "/api/postal/complaints/import/preview",
        files={"file": ("c.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["counts"]["import"] == 2
    assert client.post("/api/postal/complaints/import/commit", json={"session_id": body["session_id"]}).json()["created"] == 2

    # 列表 + 年度筛选
    lst = client.get("/api/postal/complaints?year=2024").json()
    assert lst["total"] == 2
    # 状态筛选（马宁有回访 → resolved）
    resolved = client.get("/api/postal/complaints?status=resolved").json()
    assert resolved["total"] == 1
    assert resolved["rows"][0]["snap_name"] == "马宁"
    assert resolved["rows"][0]["routed_label"] == "徐州11185"
    assert resolved["rows"][0]["routed_unit_name"] == "北京集订分送"
    # 处理次数≥2 筛选
    assert client.get("/api/postal/complaints?min_handling_count=2").json()["total"] == 1


def test_complaint_invalid_status_is_422(client):
    """非法 status 值 → 422（不 500）。"""
    assert client.get("/api/postal/complaints?status=spam").status_code == 422


_ADDR_HEADERS = ["修改日期", "姓名", "新姓名", "新电话", "新地址", "处理情况",
                 "原读者起月日 (邮局2024读者明细)", "编号", "备注"]
_ADDR_ROWS = [
    {"修改日期": "2024-01-03", "姓名": "韩博武", "新地址": "陕西省西安市碑林区X", "处理情况": "转北京局微信", "编号": "000402"},
    {"修改日期": "2024-01-05", "姓名": "赵旭", "新姓名": "肖老师", "新电话": "18616817895",
     "新地址": "上海市浦东新区Y", "处理情况": "转广东局微信", "编号": "000637"},
]


def _addr_wb() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "邮局年改地址"; ws.append(_ADDR_HEADERS)
    for r in _ADDR_ROWS:
        ws.append([r.get(h, "") for h in _ADDR_HEADERS])
    b = io.BytesIO(); wb.save(b); return b.getvalue()


def test_address_change_import_list_apply(client):
    r = client.post("/api/postal/address-changes/import/preview",
                    files={"file": ("a.xlsx", _addr_wb(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["counts"]["import"] == 2
    assert client.post("/api/postal/address-changes/import/commit", json={"session_id": body["session_id"]}).json()["created"] == 2

    lst = client.get("/api/postal/address-changes").json()
    assert lst["total"] == 2
    # 未挂订单的改地址回流 → 400
    cid = lst["rows"][0]["id"]
    assert client.post(f"/api/postal/address-changes/{cid}/apply").status_code == 400


_FU_HEADERS = ["编号", "姓名", "起月日", "止月日", "投递单位", "年度", "20240227回访", "2025回访"]
_FU_ROWS = [
    {"编号": "719", "姓名": "张三", "起月日": "0101", "止月日": "1231", "投递单位": "北京集订分送",
     "年度": "2024年", "20240227回访": "——", "2025回访": "拒接"},
]


def _fu_wb() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "邮局读者明细"; ws.append(_FU_HEADERS)
    for r in _FU_ROWS:
        ws.append([r.get(h, "") for h in _FU_HEADERS])
    b = io.BytesIO(); wb.save(b); return b.getvalue()


def test_follow_up_import_and_list(client):
    r = client.post("/api/postal/follow-ups/import/preview",
                    files={"file": ("r.xlsx", _fu_wb(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["counts"]["import"] == 2  # 张三的两条回访
    assert client.post("/api/postal/follow-ups/import/commit", json={"session_id": body["session_id"]}).json()["created"] == 2
    assert client.get("/api/postal/follow-ups").json()["total"] == 2


_FIN_HEADERS = ["姓名", "商品名称", "份数", "金额", "手续费", "到款金额", "到款日期",
                "开票金额", "发票信息", "发票类型", "订单平台"]
_FIN_ROWS = [
    {"姓名": "张翠", "商品名称": "《中国经营报》", "份数": "1", "金额": "240", "手续费": "1.3",
     "到款金额": "238.7", "到款日期": "2024-01-30", "开票金额": "240",
     "发票信息": "发票抬头：某公司\n购方税号：91ABC", "发票类型": "普票", "订单平台": "CBJ+小程序"},
    {"姓名": "吴婷", "商品名称": "《商学院》", "份数": "1", "金额": "20", "手续费": "0.11",
     "到款金额": "", "到款日期": "2024-01-16", "开票金额": "20",
     "发票信息": "不开票", "发票类型": "专票", "订单平台": "商学院APP"},
]


def _fin_wb() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "提现发票合集"; ws.append(_FIN_HEADERS)
    for r in _FIN_ROWS:
        ws.append([r.get(h, "") for h in _FIN_HEADERS])
    b = io.BytesIO(); wb.save(b); return b.getvalue()


def test_finance_import_and_list(client):
    r = client.post("/api/finance/postal-receipts/import/preview",
                    files={"file": ("f.xlsx", _fin_wb(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["counts"]["import"] == 2
    assert client.post("/api/finance/postal-receipts/import/commit", json={"session_id": body["session_id"]}).json()["created"] == 2
    assert client.get("/api/finance/postal-receipts").json()["total"] == 2
    # 平台 + 普专票筛选
    assert client.get("/api/finance/postal-receipts?platform=商学院APP").json()["total"] == 1
    assert client.get("/api/finance/postal-receipts?tax_category=专票").json()["rows"][0]["payer_name"] == "吴婷"
