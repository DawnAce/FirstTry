import io
from datetime import date

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models import (
    PublicationSchedule,
    ShippingDetail,
    ShippingDetailSourceType,
)
from app.schemas.history_import import ShippingImportRow
from app.services.recurring_shipping_detail_service import (
    SHANGYOU_GOVERNMENT_NAMES,
    SHANGYOU_GOVERNMENT_RECIPIENTS,
    ensure_recurring_shipping_details,
    exclude_recurring_shipping_import_rows,
)
from app.services.recurring_shipping_cleanup_service import (
    build_recurring_duplicate_cleanup_plan,
    delete_recurring_duplicate_candidates,
)
from app.services.shipping_plan_import_service import _parse_shipping_file


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(autocommit=False, autoflush=False, bind=engine)()
    yield session
    session.close()


def test_uploaded_recurring_rows_are_ignored_with_warning():
    fixed = ShippingImportRow(
        sheet_name="上犹",
        channel="赠阅",
        sub_channel="政府",
        transport="邮政物流",
        frequency="周",
        status="正常",
        name="上犹县政府办",
        quantity=10,
    )
    ordinary = fixed.model_copy(update={"sheet_name": "每周（读者）", "name": "普通读者"})

    rows, warnings = exclude_recurring_shipping_import_rows([fixed, ordinary], year=2026)

    assert [row.name for row in rows] == ["普通读者"]
    assert len(warnings) == 1
    assert "本次导入会忽略" in warnings[0]

    rows_2027, warnings_2027 = exclude_recurring_shipping_import_rows([fixed, ordinary], year=2027)
    assert [row.name for row in rows_2027] == ["上犹县政府办", "普通读者"]
    assert warnings_2027 == []


def test_full_recurring_set_uses_explicit_30_copy_preview_notice():
    rows = [
        ShippingImportRow(
            sheet_name="上犹",
            channel="赠阅",
            sub_channel="政府",
            transport="邮政物流",
            frequency="周",
            status="正常",
            name=name,
            quantity=quantity,
        )
        for name, quantity in (
            ("上犹县政府办", 10),
            ("上犹县人大办", 11),
            ("上犹县政协办", 9),
        )
    ]

    kept, warnings = exclude_recurring_shipping_import_rows(rows, year=2026)

    assert kept == []
    assert warnings == [
        "现有系统里已有2026年「上犹」的3个政府单位，导入数据会忽略该30份明细。"
    ]


def test_original_upload_ignores_fixed_rows_before_legacy_adjustments():
    workbook = Workbook()
    summary = workbook.active
    summary.title = "每周合计"
    summary.append(["2026年1月5日《中国经营报》中通发货表", "总第2635期"])
    corporate = workbook.create_sheet("每周（对公）")
    corporate.append(["2026年1月5日《中国经营报》中通发货表", "", "总第2635期"])
    corporate.append([
        "姓名", "地址", "电话", "份数", "刊物", "渠道", "子渠道",
        "签约公司", "频率", "运输方式", "城市", "备注",
    ])
    corporate.append([
        "普通收件人", "测试地址", "13800000000", 1, "中国经营报", "个人订阅", "",
        "", "周", "中通物流", "北京", "",
    ])
    shangyou = workbook.create_sheet("上犹")
    shangyou.append(["2026年1月5日《中国经营报》中通发货表", "", "总第2635期"])
    shangyou.append([
        "姓名", "地址", "电话", "份数", "刊物", "渠道", "子渠道",
        "签约公司", "频率", "运输方式", "城市", "备注",
    ])
    shangyou.append([
        "上犹县政府办", "测试地址", "0797-8542306", 10, "中国经营报", "赠阅", "政府",
        "上犹县政府", "周", "邮政物流", "赣州", "政府赠报，邮政",
    ])
    buffer = io.BytesIO()
    workbook.save(buffer)

    issue_number, rows, warnings, adjustments = _parse_shipping_file(
        buffer.getvalue(), recurring_year=2026
    )

    assert issue_number == 2635
    assert [row.name for row in rows] == ["普通收件人"]
    assert adjustments == []
    assert any("本次导入会忽略" in warning for warning in warnings)


def test_backfill_corrects_existing_and_fills_only_active_issues(db):
    db.add_all([
        PublicationSchedule(year=2026, issue_number=2635, publish_date=date(2026, 1, 5), is_suspended=False),
        PublicationSchedule(year=2026, issue_number=None, publish_date=date(2026, 1, 12), is_suspended=True),
        PublicationSchedule(year=2026, issue_number=2637, publish_date=date(2026, 1, 19), is_suspended=False),
    ])
    wrong = ShippingDetail(
        issue_number=2635,
        sheet_name="上犹",
        channel="",
        sub_channel=None,
        transport="周",
        frequency="上犹县政府",
        status="正常",
        name="上犹县政府办",
        address="旧错位地址",
        phone="0797-8542306",
        quantity=10,
        company="政府",
        notes="历史说明：赠阅",
        shipped_quantity=10,
        tracking_no="KEEP-ME",
        source_type=ShippingDetailSourceType.historical_import,
    )
    correct_but_manual = ShippingDetail(
        issue_number=2637,
        sheet_name="上犹",
        channel="赠阅",
        sub_channel="政府",
        transport="邮政物流",
        frequency="周",
        status="正常",
        name="上犹县政府办",
        address="江西省赣州市上犹县东山镇犹江大道16号县政府大楼325室政府办",
        phone="0797-8542306",
        quantity=10,
        deadline="",
        notes="政府赠报，邮政",
        extra_info="",
        station_name="",
        station_hall="",
        contact_person="",
        company="上犹县政府",
        source_type=ShippingDetailSourceType.manual,
    )
    db.add_all([wrong, correct_but_manual])
    db.commit()

    result = ensure_recurring_shipping_details(db, year=2026)
    db.commit()

    assert result.active_issue_count == 2
    assert result.created_count == 4
    assert result.updated_count == 2
    assert result.changed_issue_numbers == [2635, 2637]
    all_rows = db.query(ShippingDetail).order_by(ShippingDetail.issue_number, ShippingDetail.name).all()
    assert len(all_rows) == 6
    assert {row.issue_number for row in all_rows} == {2635, 2637}
    assert {row.name for row in all_rows} == SHANGYOU_GOVERNMENT_NAMES
    assert all(row.channel == "赠阅" and row.sub_channel == "政府" for row in all_rows)
    assert all(row.source_type == ShippingDetailSourceType.recurring_generated for row in all_rows)
    db.refresh(wrong)
    assert wrong.tracking_no == "KEEP-ME"
    assert wrong.shipped_quantity == 10

    second = ensure_recurring_shipping_details(db, year=2026)
    db.rollback()
    assert second.created_count == 0
    assert second.updated_count == 0
    assert second.unchanged_count == 6


def test_backfill_rejects_duplicate_existing_recipient(db):
    db.add(PublicationSchedule(
        year=2026,
        issue_number=2635,
        publish_date=date(2026, 1, 5),
        is_suspended=False,
    ))
    for _ in range(2):
        db.add(ShippingDetail(
            issue_number=2635,
            sheet_name="上犹",
            channel="赠阅",
            sub_channel="政府",
            transport="邮政物流",
            frequency="周",
            status="正常",
            name="上犹县政府办",
            quantity=10,
        ))
    db.commit()

    with pytest.raises(ValueError, match="存在重复固定收件人"):
        ensure_recurring_shipping_details(db, year=2026)


def test_duplicate_cleanup_only_applies_safe_identity_matches(db):
    db.add(PublicationSchedule(
        year=2026,
        issue_number=2635,
        publish_date=date(2026, 1, 5),
        is_suspended=False,
    ))
    db.commit()
    ensure_recurring_shipping_details(db, year=2026)
    db.commit()
    recipient_by_name = {
        recipient["name"]: recipient
        for recipient in SHANGYOU_GOVERNMENT_RECIPIENTS
    }

    safe_data = recipient_by_name["上犹县政府办"]
    safe = ShippingDetail(
        issue_number=2635,
        sheet_name="上犹",
        channel="赠阅",
        name=safe_data["name"],
        address=safe_data["address"],
        phone=safe_data["phone"],
        quantity=safe_data["quantity"],
        source_type=ShippingDetailSourceType.manual,
    )
    protected_data = recipient_by_name["上犹县人大办"]
    protected = ShippingDetail(
        issue_number=2635,
        sheet_name="上犹",
        channel="赠阅",
        name=protected_data["name"],
        address=protected_data["address"],
        phone=protected_data["phone"],
        quantity=protected_data["quantity"],
        shipped_quantity=protected_data["quantity"],
        source_type=ShippingDetailSourceType.historical_import,
    )
    mismatch_data = recipient_by_name["上犹县政协办"]
    mismatch = ShippingDetail(
        issue_number=2635,
        sheet_name="上犹",
        channel="赠阅",
        name=mismatch_data["name"],
        address="不同地址",
        phone=mismatch_data["phone"],
        quantity=mismatch_data["quantity"],
        source_type=ShippingDetailSourceType.manual,
    )
    db.add_all([safe, protected, mismatch])
    db.commit()

    plan = build_recurring_duplicate_cleanup_plan(db, year=2026)

    assert plan.candidate_ids == [safe.id]
    assert [item.detail_id for item in plan.protected] == [protected.id]
    assert [item.detail_id for item in plan.skipped] == [mismatch.id]
    assert db.get(ShippingDetail, safe.id) is not None

    assert delete_recurring_duplicate_candidates(db, plan) == 1
    db.commit()
    assert db.get(ShippingDetail, safe.id) is None
    assert db.get(ShippingDetail, protected.id) is not None
    assert db.get(ShippingDetail, mismatch.id) is not None
