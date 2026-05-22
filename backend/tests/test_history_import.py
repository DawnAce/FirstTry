import io
import unittest
from datetime import date, datetime as _dt

from fastapi import HTTPException
from openpyxl import load_workbook, Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.history_import_cache import get_history_import_session
from app.models import ReportItemTemplate, Issue, IssueStatus, ReportEntry, TempPrintDetail, ShippingDetail
from app.services.history_import_template_service import (
    build_report_import_template,
    build_shipping_import_template,
)
from app.services.history_import_service import preview_history_import, commit_history_import
from app.services.original_zto_shipping_import_service import read_original_zto_shipping_rows
from app.services.raw_report_import_service import parse_raw_report_workbook


_SHIPPING_HEADERS = [
    "工作表名称", "渠道", "子渠道", "运输方式", "频次", "状态",
    "姓名", "地址", "电话", "数量", "截止日期", "备注", "附加信息",
    "网点名称", "网点大厅", "联系人", "序号", "期数", "公司",
]


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_report_upload_with_datetime_date(issue_number: int = 2648) -> bytes:
    """Same structure as build_report_upload but with a datetime cell for publish_date."""
    wb = Workbook()
    basic = wb.active
    basic.title = "基本信息"
    basic.append(["字段", "值"])
    basic.append(["期号", issue_number])
    basic.append(["出版日期", _dt(2026, 4, 20)])   # Excel date cell (datetime object)
    basic.append(["版数", 24])
    basic.append(["备注", ""])

    report = wb.create_sheet("报数项")
    report.append(["分类编码", "分类名称", "项目名称", "去向", "是否变动", "数值"])
    report.append(["postal", "北京邮发", "本市", "邮局", "否", 100])

    temp = wb.create_sheet("临时加印明细")
    temp.append(["部门", "自定义名称", "数量", "自留分发数量"])

    return _wb_to_bytes(wb)


def build_report_upload_with_unknown_row(issue_number: int = 2648) -> bytes:
    """Upload containing one valid row and one row not in any ReportItemTemplate."""
    wb = Workbook()
    basic = wb.active
    basic.title = "基本信息"
    basic.append(["字段", "值"])
    basic.append(["期号", issue_number])
    basic.append(["出版日期", "2026-04-20"])
    basic.append(["版数", 24])
    basic.append(["备注", ""])

    report = wb.create_sheet("报数项")
    report.append(["分类编码", "分类名称", "项目名称", "去向", "是否变动", "数值"])
    report.append(["postal", "北京邮发", "本市", "邮局", "否", 100])         # valid
    report.append(["unknown", "未知分类", "未知项目", "", "否", 5])           # not in templates

    temp = wb.create_sheet("临时加印明细")
    temp.append(["部门", "自定义名称", "数量", "自留分发数量"])

    return _wb_to_bytes(wb)


def build_report_upload(issue_number: int = 2648) -> bytes:
    wb = Workbook()
    basic = wb.active
    basic.title = "基本信息"
    basic.append(["字段", "值"])
    basic.append(["期号", issue_number])
    basic.append(["出版日期", "2026-04-20"])
    basic.append(["版数", 24])
    basic.append(["备注", "测试备注"])

    report = wb.create_sheet("报数项")
    report.append(["分类编码", "分类名称", "项目名称", "去向", "是否变动", "数值"])
    report.append(["postal", "北京邮发", "本市", "邮局", "否", 100])
    report.append(["retail", "北京报零", "西部", "零售点", "是", 50])

    temp = wb.create_sheet("临时加印明细")
    temp.append(["部门", "自定义名称", "数量", "自留分发数量"])
    temp.append(["编辑部", "赠送用", 20, 5])

    return _wb_to_bytes(wb)


def build_report_upload_with_blank_publish_date(issue_number: int = 2648) -> bytes:
    wb = Workbook()
    basic = wb.active
    basic.title = "基本信息"
    basic.append(["字段", "值"])
    basic.append(["期号", issue_number])
    basic.append(["出版日期", ""])
    basic.append(["版数", 24])
    basic.append(["备注", "测试备注"])

    report = wb.create_sheet("报数项")
    report.append(["分类编码", "分类名称", "项目名称", "去向", "是否变动", "数值"])
    report.append(["postal", "北京邮发", "本市", "邮局", "否", 100])

    temp = wb.create_sheet("临时加印明细")
    temp.append(["部门", "自定义名称", "数量", "自留分发数量"])

    return _wb_to_bytes(wb)


def build_report_upload_with_invalid_issue_number(issue_number: str = "第2648期") -> bytes:
    wb = Workbook()
    basic = wb.active
    basic.title = "基本信息"
    basic.append(["字段", "值"])
    basic.append(["期号", issue_number])
    basic.append(["出版日期", "2026-04-20"])
    basic.append(["版数", 24])
    basic.append(["备注", "测试备注"])

    report = wb.create_sheet("报数项")
    report.append(["分类编码", "分类名称", "项目名称", "去向", "是否变动", "数值"])
    report.append(["postal", "北京邮发", "本市", "邮局", "否", 100])

    temp = wb.create_sheet("临时加印明细")
    temp.append(["部门", "自定义名称", "数量", "自留分发数量"])

    return _wb_to_bytes(wb)


def build_shipping_upload(issue_number: int = 2648, quantity: int = 10) -> bytes:
    wb = Workbook()
    basic = wb.active
    basic.title = "基本信息"
    basic.append(["字段", "值"])
    basic.append(["期号", issue_number])
    basic.append(["出版日期", "2026-04-20"])

    detail = wb.create_sheet("发货明细")
    detail.append(_SHIPPING_HEADERS)
    detail.append([
        "发货明细", "邮发", "本市", "中通物流", "每周", "正常",
        "张三", "北京市朝阳区xx路1号", "13800138000", quantity,
        "2026-04-19", "", "", "", "", "", 1, issue_number, "",
    ])

    return _wb_to_bytes(wb)


def build_original_zto_shipping_upload(issue_number: int = 2648, high_speed_sheet_name: str = "高铁展示") -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "每周合计"
    summary.append([f"2026年4月20日《中国经营报》中通发货表", f"总第{issue_number}期"])
    summary.append(["各渠道统计合计"])

    weekly_corporate = wb.create_sheet("每周（对公）")
    weekly_corporate.append([f"2026年4月20日《中国经营报》中通发货表", "", f"总第{issue_number}期"])
    weekly_corporate.append(["姓名", "地址", "电话", "份数", "刊物", "渠道", "子渠道", "签约公司", "频率", "运输方式", "城市", "备注"])
    weekly_corporate.append(["叶剑", "广州市白云区增槎路1113号广州日报印务中心北门", "13556046615", 531, "中国经营报", "渠道订阅", "", "广州日报", "周", "中通物流", "广州", "广州日报，要求务必打电话通知！"])
    weekly_corporate.append(["", "中通库房，详细地址待定", "", 0, "中国经营报", "", "", "", "", "", "", "加印，"])
    weekly_corporate.append(["", "", "合计", 531])

    weekly_reader = wb.create_sheet("每周（读者）")
    weekly_reader.append([f"2026年4月20日《中国经营报》中通发货表", "", f"总第{issue_number}期"])
    weekly_reader.append(["姓名", "地址", "电话", "份数", "刊物", "截止日期", "渠道", "子渠道", "签约公司", "频率", "运输方式", "城市", "备注", "附加信息"])
    weekly_reader.append(["黄雪", "北京市海淀区杏石口路5号", "15110271926", 1, "中国经营报", "长期", "赠阅", "监管", "", "周", "中通物流", "北京", "", ""])
    weekly_reader.append(["", "", "合计", 1])

    rail = wb.create_sheet(high_speed_sheet_name)
    rail.append([f"2026年4月20日《中国经营报》中通发货表", "", "", "", f"总第{issue_number}期"])
    rail.append(["《中国经营报》《商学院》高铁贵宾厅投放站点信息统计表"])
    rail.append(["城市", "序号", "站名", "站厅名称", "联系人", "联系电话", "收货地址", "投放数量", "信息确认", "附加信息", "渠道", "子渠道", "签约公司", "频率", "运输方式"])
    rail.append(["北京", 1, "北京站", "商务座候车区（北京站）", "赵叶", "15810698235", "北京市东城区北京站广场西侧", 5, "☑", "", "对公订阅", "", "北京悦途出行", "周", "中通物流"])
    rail.append(["", 2, "北京南站", "商务座候车区（北京南站）", "李四", "15810698236", "北京市丰台区北京南站", 5, "☑", "", "对公订阅", "", "北京悦途出行", "周", "中通物流"])

    shangyou = wb.create_sheet("上犹")
    shangyou.append([f"2026年4月20日《中国经营报》中通发货表", "", f"总第{issue_number}期"])
    shangyou.append(["姓名", "地址", "电话", "份数", "刊物", "渠道", "子渠道", "签约公司", "频率", "运输方式", "城市", "备注"])
    shangyou.append(["上犹县政府办", "江西省赣州市上犹县东山镇犹江大道16号", "0797-8542306", 10, "中国经营报", "赠阅", "政府", "上犹县政府", "周", "邮政物流", "赣州", "政府赠报，邮政"])
    shangyou.append(["", "", "合计", 10])

    suspended = wb.create_sheet("停发-双周（读者）")
    suspended.append([f"2026年4月上半月《中国经营报》中通发货表", "", f"总第{issue_number}期"])
    suspended.append(["姓名", "地址", "电话", "期数", "份数", "刊物", "截止日期", "备注"])
    suspended.append(["丁联诚", "上海市普陀区长寿路1086号", "13319400970", 2, 1, "中国经营报", _dt(2026, 5, 1), "5-13开始停发"])
    suspended.append(["", "", "合计", "", 1])

    monthly = wb.create_sheet("月底-整月")
    monthly.append([f"2026年4月《中国经营报》中通发货表", "", f"总 第{issue_number - 3}期、第{issue_number - 2}期 第{issue_number - 1}期、第{issue_number}期"])
    monthly.append(["姓名", "地址", "电话", "期数", "份数", "刊物", "截止日期", "渠道", "子渠道", "签约公司", "频率", "运输方式", "城市", "备注", "附加信息"])
    monthly.append(["宣传部5号格", "北京市通州区运河东大街56号院", "", 5, 3, "中国经营报", "", "赠阅", "监管", "", "月", "中通物流", "北京", "", ""])
    monthly.append(["", "", "合计", "", 3])

    return _wb_to_bytes(wb)


def build_raw_report_upload() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "北京印厂"
    ws.append(["中国经营报印数表"])
    ws.append(["（此表给人民日报印厂邮一科）"])
    ws.append(["期数：2647   版数：24    出版日期：2026年4月13日"])
    ws.append(["渠道名称", None, "数量", "插报要求"])
    ws.append(["北京报刊局\n(邮发）", "本市", 1213, None])
    ws.append([None, "外埠", 5573, None])
    ws.append(["报社总部（社用）", None, 200, "早4：00前插报完成"])
    ws.append(["报社总部（其他）", None, 1212, "早4：00前插报完成"])
    ws.append(["北京零售公司\n(邮局-快速邮路）", "东部", 460, "早2：30前插报完成"])
    ws.append([None, "西部", 592, "早2：30前插报完成"])
    ws.append(["合订本专用", None, 15, "印厂留存"])
    ws.append(["合计", None, 9265, None])

    retail = wb.create_sheet("零售渠道`")
    retail.append(["传统零售渠道报数统计"])
    retail.append(["期数：2647   版数：24    出版日期：2026年4月13日"])
    retail.append(["渠道名称", "本期印数", "上期印数", "同上期比", "印刷地点"])
    retail.append(["北京报零", 460, 460, 0, "东部"])
    retail.append([None, 592, 592, 0, "西部"])
    retail.append(["广州日报\n（零售）", 500, 500, 0, "北京快递"])
    retail.append(["报数合计", 1552, 1552, 0, None])

    subscription = wb.create_sheet("订阅渠道`")
    subscription.append(["订阅渠道报数统计"])
    subscription.append(["期数：2647   版数：24    出版日期：2026年4月13日"])
    subscription.append(["渠道名称", "本期印数", "上期印数", "同上期比", "印刷地点"])
    subscription.append(["北京报刊局\n（邮发）", 6786, 6782, 4, None])
    subscription.append(["国图贸", 1, 1, 0, "北京快递"])
    subscription.append(["广州日报\n（订户）", 31, 31, 0, "北京快递"])
    subscription.append(["杂志铺", 366, 366, 0, "北京快递"])
    subscription.append(["报数合计", 7184, 7180, 4, None])

    social = wb.create_sheet("社用报`")
    social.append(["《中国经营报》社用报统计表"])
    social.append(["期数：2647   版数：24    出版日期：2026年4月13日"])
    social.append(["地点", "本期印数", "上期印数", "同上期比", "分发地", "印刷地点", None, None, None, None])
    social.append(["营报传媒", 151, 151, 0, "报社", "北京", None, None, 29, "报社"])
    social.append(["中经传媒智库", 3, 3, 0, "报社", None, None, None, 50, "读者"])
    social.append(["新闻中心", 45, 45, 0, "报社", None, None, None, 72, "备用"])
    social.append(["行政", 4, 4, 0, "报社", None, None, None, None, None])
    social.append(["财经中心", 15, 24, -9, "报社", None, None, None, None, None])
    social.append(["产经中心", 5, 5, 0, "报社", None, None, None, None, None])
    social.append(["出版中心", 10, 10, 0, "报社", None, None, None, None, None])
    social.append(["品牌中心", 5, 5, 0, "报社", None, None, None, None, None])
    social.append(["经营网", 7, 7, 0, "报社", None, None, None, None, None])
    social.append(["法务", 2, 2, 0, "报社", None, None, None, None, None])
    social.append(["社科院、工经所", 64, 64, 0, "报社", None, None, None, None, None])
    social.append(["财务", 1, 1, 0, "报社", None, None, None, None, None])
    social.append(["库房", 10, 10, 0, "报社", None, None, None, None, None])
    social.append(["印厂留存", 15, 15, 0, "报社", None, None, None, None, None])
    social.append(["报社订阅自投/展示", 140, 140, 0, "报社", "北京快递", None, None, 30, "上犹"])
    social.append(["上海站用", 10, 10, 0, "上海", "北京快递", None, None, 110, "高铁展示"])
    social.append(["广东站用", 30, 30, 0, "广东", "北京快递", None, None, None, None])
    social.append(["成都站用", 2, 2, 0, "成都", "北京快递", None, None, None, None])
    social.append(["西安站用", 10, 10, 0, "西安", "北京快递", None, None, None, None])
    social.append(["营报传媒加印", 0, 0, 0, "报社", "北京", None, None, None, None])
    social.append(["临时加印", 0, 0, 0, "展示", "北京快递", None, None, None, None])
    social.append(["合计", 529, 538, -9, None, None, None, None, None, None])

    distribution = wb.create_sheet("收发室自留分发（需打印）")
    distribution.append([])
    distribution.append(["《中国经营报》社用报分发表"])
    distribution.append(["期数：2647   版数：24    出版日期：2026年4月13日"])
    distribution.append(["部门", "本期印数", "领用人", "备注", "领用人签字"])
    distribution.append(["营报传媒", 29, "卢娅丽", None, None])
    distribution.append(["中经传媒智库", 3, "赵震平", None, None])
    distribution.append(["新闻中心", 45, "翟军", None, None])
    distribution.append(["行政", 4, "黄鹤", None, None])
    distribution.append(["财经中心", 15, "齐士娟", None, None])
    distribution.append(["产经中心", 5, "刘家懿", None, None])
    distribution.append(["出版中心", 10, "翟军", None, None])
    distribution.append(["品牌中心", 5, "张威", None, None])
    distribution.append(["经营网", 7, "孙明胜", None, None])
    distribution.append(["法务", 2, "张伊萍", None, None])
    distribution.append(["社科院、工经所", 64, None, "社科院43、工经所21", None])
    distribution.append(["财务", 1, "王钧", None, None])
    distribution.append(["库房", 10, None, None, None])
    distribution.append(["临时加印（报社内存）", 0, None, None, None])
    distribution.append(["合计", 200, None, None, None])

    return _wb_to_bytes(wb)


def build_raw_report_upload_with_temp_print() -> bytes:
    wb = load_workbook(io.BytesIO(build_raw_report_upload()))
    ws = wb["社用报`"]
    ws["B24"] = 12
    ws["B25"] = 8
    ws["B26"] = 4
    ws["B27"] = 24
    ws["B28"] = 553
    return _wb_to_bytes(wb)


def build_raw_report_upload_with_total_mismatch() -> bytes:
    wb = load_workbook(io.BytesIO(build_raw_report_upload()))
    wb["北京印厂"]["C12"] = 9999
    return _wb_to_bytes(wb)


def build_raw_report_upload_with_unmapped_item() -> bytes:
    wb = load_workbook(io.BytesIO(build_raw_report_upload()))
    ws = wb["订阅渠道`"]
    ws.insert_rows(8)
    ws["A8"] = "神秘渠道"
    ws["B8"] = 12
    return _wb_to_bytes(wb)


def build_raw_report_upload_with_unmapped_distribution_item() -> bytes:
    wb = load_workbook(io.BytesIO(build_raw_report_upload()))
    ws = wb["收发室自留分发（需打印）"]
    ws.insert_rows(18)
    ws["A18"] = "临时新增部门"
    ws["B18"] = 6
    return _wb_to_bytes(wb)


def build_raw_report_upload_with_resolved_temp_print() -> bytes:
    wb = load_workbook(io.BytesIO(build_raw_report_upload_with_temp_print()))
    wb["收发室自留分发（需打印）"]["B18"] = 12
    wb["北京印厂"]["C12"] = 9289
    return _wb_to_bytes(wb)


class RawReportImportParserTests(unittest.TestCase):
    def test_parses_original_report_metadata_and_mapped_rows(self):
        workbook = load_workbook(io.BytesIO(build_raw_report_upload()), data_only=True)

        result = parse_raw_report_workbook(workbook)

        self.assertEqual(result.issue_number, 2647)
        self.assertEqual(result.publish_date, "2026-04-13")
        self.assertEqual(result.page_count, 24)
        row_map = {(row.category, row.sub_category): row.value for row in result.report_rows}
        self.assertEqual(row_map[("postal", "本市")], 1213)
        self.assertEqual(row_map[("postal", "外埠")], 5573)
        self.assertEqual(row_map[("retail", "东部")], 460)
        self.assertEqual(row_map[("retail", "西部")], 592)
        self.assertEqual(row_map[("guangzhou", "零售")], 500)
        self.assertEqual(row_map[("guangzhou", "订阅")], 31)
        self.assertEqual(row_map[("chengdu", "成都杂志铺")], 366)
        self.assertEqual(row_map[("guotumao", "国图贸")], 1)
        self.assertEqual(row_map[("social_use", "营报传媒_收发室")], 29)
        self.assertEqual(row_map[("social_use", "营报传媒_读者")], 50)
        self.assertEqual(row_map[("social_use", "营报传媒_备用报")], 72)
        self.assertEqual(row_map[("social_use", "营报传媒_上犹")], 30)
        self.assertEqual(row_map[("social_use", "高铁展示")], 110)
        self.assertEqual(row_map[("social_use", "临时加印")], 0)
        self.assertEqual(row_map[("social_use", "临时加印_自留")], 0)
        self.assertEqual(row_map[("binding", "合订本（印厂留存）")], 15)
        self.assertEqual(result.source_total, 9265)
        self.assertEqual(result.mapped_total, 9265)
        self.assertEqual(result.unmapped_items, [])
        workbook.close()

    def test_detects_unmapped_original_report_rows(self):
        workbook = load_workbook(io.BytesIO(build_raw_report_upload_with_unmapped_item()), data_only=True)

        result = parse_raw_report_workbook(workbook)

        self.assertEqual(result.unmapped_items, ["订阅渠道`：神秘渠道（12份）"])
        workbook.close()

    def test_detects_unmapped_distribution_rows(self):
        workbook = load_workbook(io.BytesIO(build_raw_report_upload_with_unmapped_distribution_item()), data_only=True)

        result = parse_raw_report_workbook(workbook)

        self.assertEqual(result.unmapped_items, ["收发室自留分发（需打印）：临时新增部门（6份）"])
        workbook.close()


class HistoryImportTemplateTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(bind=self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

    def test_report_template_has_expected_sheets(self):
        db = self.SessionLocal()

        template_bytes = build_report_import_template(db)
        workbook = load_workbook(io.BytesIO(template_bytes))

        self.assertEqual(workbook.sheetnames, ["基本信息", "报数项", "临时加印明细"])
        db.close()

    def test_shipping_template_has_expected_sheets(self):
        template_bytes = build_shipping_import_template()
        workbook = load_workbook(io.BytesIO(template_bytes))

        self.assertEqual(workbook.sheetnames, ["基本信息", "发货明细"])

    def test_report_template_uses_required_headers(self):
        db = self.SessionLocal()

        template_bytes = build_report_import_template(db)
        workbook = load_workbook(io.BytesIO(template_bytes))

        self.assertEqual(
            [cell.value for cell in workbook["基本信息"][1]],
            ["字段", "值"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["报数项"][1]],
            ["分类编码", "分类名称", "项目名称", "去向", "是否变动", "数值"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["临时加印明细"][1]],
            ["部门", "自定义名称", "数量", "自留分发数量"],
        )
        db.close()

    def test_report_template_includes_basic_info_notes(self):
        db = self.SessionLocal()

        template_bytes = build_report_import_template(db)
        workbook = load_workbook(io.BytesIO(template_bytes))

        basic_rows = [
            [cell.value for cell in row]
            for row in workbook["基本信息"].iter_rows(values_only=False)
        ]

        self.assertIn(["填写说明", "1. 只填“值”列；2. 报数项只改“数值”列；3. 临时加印总数填在报数项 sheet。"], basic_rows)
        self.assertIn(["临时加印说明", "先在报数项 sheet 填“临时加印”总数；再在“临时加印明细” sheet 按行填写部门、数量、自留分发数量；明细数量合计应等于总数。"], basic_rows)
        self.assertIn(["临时加印示例", "例如总数 20：营报传媒 12、自留 2；财经中心 8、自留 0。没有临时加印时，总数填 0，明细可留空。"], basic_rows)
        db.close()

    def test_shipping_template_uses_required_headers(self):
        template_bytes = build_shipping_import_template()
        workbook = load_workbook(io.BytesIO(template_bytes))

        self.assertEqual(
            [cell.value for cell in workbook["基本信息"][1]],
            ["字段", "值"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["发货明细"][1]],
            [
                "工作表名称",
                "渠道",
                "子渠道",
                "运输方式",
                "频次",
                "状态",
                "姓名",
                "地址",
                "电话",
                "数量",
                "截止日期",
                "备注",
                "附加信息",
                "网点名称",
                "网点大厅",
                "联系人",
                "序号",
                "期数",
                "公司",
            ],
        )

    def test_report_template_rows_keep_category_code_and_label_in_order(self):
        db = self.SessionLocal()
        db.add_all(
            [
                ReportItemTemplate(
                    category="retail",
                    sub_category="西部",
                    display_name="北京报零-西部",
                    default_value=8,
                    is_variable=True,
                    destination="零售点",
                    sort_order=20,
                ),
                ReportItemTemplate(
                    category="postal",
                    sub_category="本市",
                    display_name="北京邮发-本市",
                    default_value=12,
                    is_variable=False,
                    destination="邮局",
                    sort_order=10,
                ),
            ]
        )
        db.commit()

        template_bytes = build_report_import_template(db)
        workbook = load_workbook(io.BytesIO(template_bytes))
        rows = [
            row[:6]
            for row in workbook["报数项"].iter_rows(min_row=2, max_row=3, values_only=True)
        ]

        self.assertEqual(
            rows,
            [
                ("postal", "北京邮发", "本市", "邮局", "否", 12),
                ("retail", "北京报零", "西部", "零售点", "是", 8),
            ],
        )
        db.close()


class HistoryImportPreviewTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(bind=self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

    def _seed_upload_templates(self, db) -> None:
        """Seed templates that match the rows produced by build_report_upload."""
        db.add_all([
            ReportItemTemplate(
                category="postal", sub_category="本市", display_name="北京邮发-本市",
                default_value=0, is_variable=False, destination="邮局", sort_order=1,
            ),
            ReportItemTemplate(
                category="retail", sub_category="西部", display_name="北京报零-西部",
                default_value=0, is_variable=True, destination="零售点", sort_order=2,
            ),
        ])
        db.commit()

    def _seed_raw_report_templates(self, db) -> None:
        rows = [
            ("postal", "本市", "北京邮发-本市"),
            ("postal", "外埠", "北京邮发-外埠"),
            ("retail", "东部", "北京报零-东部"),
            ("retail", "西部", "北京报零-西部"),
            ("guangzhou", "零售", "广州日报-零售"),
            ("guangzhou", "订阅", "广州日报-订阅"),
            ("chengdu", "成都杂志铺", "成都杂志铺"),
            ("guotumao", "国图贸", "国图贸"),
            ("social_use", "营报传媒_收发室", "营报传媒-收发室"),
            ("social_use", "营报传媒_读者", "营报传媒-读者"),
            ("social_use", "营报传媒_备用报", "营报传媒-备用报"),
            ("social_use", "营报传媒_上犹", "营报传媒-上犹"),
            ("social_use", "中经传媒智库", "中经传媒智库"),
            ("social_use", "新闻中心", "新闻中心"),
            ("social_use", "行政", "行政"),
            ("social_use", "财经中心", "财经中心"),
            ("social_use", "产经中心", "产经中心"),
            ("social_use", "出版中心", "出版中心"),
            ("social_use", "品牌中心", "品牌中心"),
            ("social_use", "经营网", "经营网"),
            ("social_use", "法务", "法务"),
            ("social_use", "社科院、工经所", "社科院、工经所"),
            ("social_use", "财务", "财务"),
            ("social_use", "库房", "库房"),
            ("social_use", "高铁展示", "高铁展示"),
            ("social_use", "上海站用", "上海站用"),
            ("social_use", "广东站用", "广东站用"),
            ("social_use", "成都站用", "成都站用"),
            ("social_use", "西安站用", "西安站用"),
            ("social_use", "临时加印", "临时加印"),
            ("social_use", "临时加印_自留", "临时加印（自留）"),
            ("binding", "合订本（印厂留存）", "合订本（印厂留存）"),
        ]
        db.add_all([
            ReportItemTemplate(
                category=category,
                sub_category=sub_category,
                display_name=display_name,
                default_value=0,
                is_variable=True,
                destination="",
                sort_order=index,
            )
            for index, (category, sub_category, display_name) in enumerate(rows, start=1)
        ])
        db.commit()

    def test_preview_returns_counts_and_session_id(self):
        db = self.SessionLocal()
        self._seed_upload_templates(db)

        result = preview_history_import(db, build_report_upload(), build_shipping_upload())

        self.assertEqual(result.issue_number, 2648)
        self.assertEqual(result.publish_date, "2026-04-20")
        self.assertEqual(result.report_entry_count, 2)
        self.assertEqual(result.temp_detail_count, 1)
        self.assertEqual(result.shipping_detail_count, 1)
        self.assertTrue(result.can_commit)
        self.assertNotEqual(result.import_session_id, "")
        # readiness object must be present and correctly structured
        self.assertIsNotNone(result.readiness)
        self.assertTrue(result.readiness.same_issue)
        self.assertFalse(result.readiness.issue_exists)
        self.assertTrue(result.readiness.can_commit)
        self.assertEqual(result.readiness.errors, [])
        db.close()

    def test_preview_accepts_original_report_workbook_with_template_shipping_file(self):
        db = self.SessionLocal()
        self._seed_raw_report_templates(db)

        result = preview_history_import(db, build_raw_report_upload(), build_shipping_upload(2647, quantity=1412))

        self.assertEqual(result.issue_number, 2647)
        self.assertEqual(result.publish_date, "2026-04-13")
        self.assertEqual(result.report_entry_count, 32)
        self.assertEqual(result.temp_detail_count, 0)
        self.assertEqual(result.shipping_detail_count, 1)
        self.assertTrue(result.can_commit)

        payload = get_history_import_session(result.import_session_id)
        self.assertIsNotNone(payload)
        self.assertEqual(payload["page_count"], 24)
        row_map = {
            (row["category"], row["sub_category"]): row["value"]
            for row in payload["report_rows"]
        }
        self.assertEqual(row_map[("chengdu", "成都杂志铺")], 366)
        self.assertEqual(row_map[("binding", "合订本（印厂留存）")], 15)
        db.close()

    def test_preview_blocks_zto_report_and_shipping_total_mismatch(self):
        db = self.SessionLocal()
        self._seed_raw_report_templates(db)

        result = preview_history_import(db, build_raw_report_upload(), build_shipping_upload(2647, quantity=10))

        self.assertFalse(result.can_commit)
        self.assertTrue(any("中通物流份数不一致" in error for error in result.errors))
        self.assertTrue(any("报数合计 1412 份" in error and "发货明细合计 10 份" in error for error in result.errors))
        db.close()

    def test_preview_blocks_original_report_with_unresolved_temp_print(self):
        db = self.SessionLocal()
        self._seed_raw_report_templates(db)

        result = preview_history_import(
            db,
            build_raw_report_upload_with_temp_print(),
            build_shipping_upload(2647),
        )

        self.assertFalse(result.can_commit)
        self.assertTrue(any("临时加印未处理" in error for error in result.errors))
        db.close()

    def test_preview_allows_original_report_when_temp_print_is_fully_resolved(self):
        db = self.SessionLocal()
        self._seed_raw_report_templates(db)

        result = preview_history_import(
            db,
            build_raw_report_upload_with_resolved_temp_print(),
            build_shipping_upload(2647, quantity=1436),
        )

        self.assertTrue(result.can_commit)
        self.assertEqual(result.errors, [])
        db.close()

    def test_preview_blocks_original_report_when_totals_do_not_match(self):
        db = self.SessionLocal()
        self._seed_raw_report_templates(db)

        result = preview_history_import(
            db,
            build_raw_report_upload_with_total_mismatch(),
            build_shipping_upload(2647),
        )

        self.assertFalse(result.can_commit)
        self.assertTrue(any("原表总印数 9999" in error and "映射后总数 9265" in error for error in result.errors))
        db.close()

    def test_preview_raw_report_errors_include_actionable_counts(self):
        db = self.SessionLocal()
        self._seed_raw_report_templates(db)

        result = preview_history_import(
            db,
            build_raw_report_upload_with_temp_print(),
            build_shipping_upload(2647),
        )

        self.assertFalse(result.can_commit)
        self.assertIn(
            "临时加印未处理：原表临时加印 12 份，其中自留分发 0 份，待手工确认 12 份",
            result.errors,
        )

        mismatch_result = preview_history_import(
            db,
            build_raw_report_upload_with_total_mismatch(),
            build_shipping_upload(2647),
        )

        self.assertIn(
            "原表总印数 9999 与映射后总数 9265 不一致，相差 734 份",
            mismatch_result.errors,
        )
        db.close()

    def test_preview_blocks_original_report_with_unmapped_items(self):
        db = self.SessionLocal()
        self._seed_raw_report_templates(db)

        result = preview_history_import(
            db,
            build_raw_report_upload_with_unmapped_item(),
            build_shipping_upload(2647),
        )

        self.assertFalse(result.can_commit)
        self.assertTrue(any("未命中映射项" in error and "神秘渠道" in error for error in result.errors))
        db.close()

    def test_preview_accepts_original_zto_shipping_workbook(self):
        db = self.SessionLocal()
        self._seed_upload_templates(db)

        result = preview_history_import(
            db,
            build_report_upload(),
            build_original_zto_shipping_upload(),
        )

        self.assertEqual(result.issue_number, 2648)
        self.assertEqual(result.publish_date, "2026-04-20")
        self.assertEqual(result.shipping_detail_count, 7)
        self.assertTrue(result.can_commit)

        payload = get_history_import_session(result.import_session_id)
        self.assertIsNotNone(payload)
        rows = payload["shipping_rows"]
        row_map = {
            (row["sheet_name"], row["name"]): row
            for row in rows
        }
        self.assertEqual(row_map[("每周（对公）", "叶剑")]["channel"], "渠道订阅")
        self.assertEqual(row_map[("每周（对公）", "叶剑")]["company"], "广州日报")
        self.assertEqual(row_map[("每周（对公）", "叶剑")]["quantity"], 531)
        self.assertEqual(row_map[("每周（读者）", "黄雪")]["channel"], "赠阅")
        self.assertEqual(row_map[("每周（读者）", "黄雪")]["sub_channel"], "监管")
        self.assertEqual(row_map[("高铁展示", "赵叶")]["company"], "北京悦途出行")
        self.assertEqual(row_map[("高铁展示", "赵叶")]["quantity"], 5)
        self.assertEqual(row_map[("上犹", "上犹县政府办")]["channel"], "赠阅")
        self.assertEqual(row_map[("上犹", "上犹县政府办")]["company"], "上犹县政府")
        self.assertEqual(row_map[("停发-双周（读者）", "丁联诚")]["status"], "停发")
        self.assertEqual(row_map[("停发-双周（读者）", "丁联诚")]["quantity"], 1)
        self.assertEqual(row_map[("月底-整月", "宣传部5号格")]["channel"], "赠阅")
        self.assertEqual(row_map[("月底-整月", "宣传部5号格")]["sub_channel"], "监管")
        self.assertEqual(row_map[("月底-整月", "宣传部5号格")]["quantity"], 3)
        db.close()

    def test_preview_skips_weekly_corporate_zero_quantity_reprint_placeholder(self):
        db = self.SessionLocal()
        self._seed_upload_templates(db)

        result = preview_history_import(
            db,
            build_report_upload(),
            build_original_zto_shipping_upload(),
        )

        self.assertTrue(result.can_commit)
        payload = get_history_import_session(result.import_session_id)
        self.assertIsNotNone(payload)
        rows = payload["shipping_rows"]
        self.assertFalse(any(
            row["sheet_name"] == "每周（对公）"
            and row["name"] == "(未填写)"
            and row["quantity"] == 0
            and "加印" in row["notes"]
            for row in rows
        ))
        db.close()

    def test_original_zto_parser_omits_city_from_rows(self):
        wb = load_workbook(io.BytesIO(build_original_zto_shipping_upload()))
        wb["每周（对公）"].append([
            "马飞", "中通库房", "", 1, "中国经营报", "库房留存", "", "", "周", "库房留存", "", "",
        ])

        rows = read_original_zto_shipping_rows(wb)

        row = next(row for row in rows if row.name == "马飞")
        self.assertFalse(hasattr(row, "city"))

    def test_preview_accepts_original_zto_high_speed_alias_sheet(self):
        db = self.SessionLocal()
        self._seed_upload_templates(db)

        result = preview_history_import(
            db,
            build_report_upload(),
            build_original_zto_shipping_upload(high_speed_sheet_name="北京悦途出行（高铁）"),
        )

        self.assertTrue(result.can_commit)
        self.assertEqual(result.shipping_detail_count, 7)
        payload = get_history_import_session(result.import_session_id)
        self.assertIsNotNone(payload)
        row_map = {
            (row["sheet_name"], row["name"]): row
            for row in payload["shipping_rows"]
        }
        self.assertEqual(row_map[("高铁展示", "赵叶")]["company"], "北京悦途出行")
        self.assertEqual(row_map[("高铁展示", "赵叶")]["quantity"], 5)
        db.close()

    def test_preview_blocks_duplicate_issue_and_cross_issue_upload(self):
        db = self.SessionLocal()

        # Seed an existing issue to trigger duplicate check
        existing = Issue(
            issue_number=2648,
            publish_date=date(2026, 4, 20),
            status=IssueStatus.confirmed,
        )
        db.add(existing)
        db.commit()

        dup_result = preview_history_import(db, build_report_upload(2648), build_shipping_upload(2648))
        self.assertFalse(dup_result.can_commit)
        self.assertTrue(any("该期已存在" in e for e in dup_result.errors))
        # readiness must reflect the correct state
        self.assertTrue(dup_result.readiness.same_issue)
        self.assertTrue(dup_result.readiness.issue_exists)
        self.assertFalse(dup_result.readiness.can_commit)

        # Cross-issue: report=2648, shipping=2649
        cross_result = preview_history_import(db, build_report_upload(2648), build_shipping_upload(2649))
        self.assertFalse(cross_result.can_commit)
        self.assertTrue(any("两份文件不是同一期" in e for e in cross_result.errors))
        # readiness must reflect cross-issue mismatch
        self.assertFalse(cross_result.readiness.same_issue)
        self.assertFalse(cross_result.readiness.can_commit)

        db.close()

    def test_preview_rejects_rows_with_unknown_template_structure(self):
        db = self.SessionLocal()
        # Seed only the known row; the unknown row has no matching template
        db.add(ReportItemTemplate(
            category="postal", sub_category="本市", display_name="北京邮发-本市",
            default_value=0, is_variable=False, destination="邮局", sort_order=1,
        ))
        db.commit()

        result = preview_history_import(
            db, build_report_upload_with_unknown_row(), build_shipping_upload()
        )

        self.assertFalse(result.can_commit)
        self.assertFalse(result.readiness.can_commit)
        self.assertTrue(result.readiness.same_issue)
        self.assertFalse(result.readiness.issue_exists)
        # Error message should identify the unknown category code
        self.assertTrue(any("unknown" in e for e in result.errors))
        db.close()

    def test_preview_normalizes_excel_datetime_publish_date(self):
        db = self.SessionLocal()
        db.add(ReportItemTemplate(
            category="postal", sub_category="本市", display_name="北京邮发-本市",
            default_value=0, is_variable=False, destination="邮局", sort_order=1,
        ))
        db.commit()

        result = preview_history_import(
            db, build_report_upload_with_datetime_date(), build_shipping_upload()
        )

        self.assertEqual(result.publish_date, "2026-04-20")
        db.close()


    def test_preview_payload_uses_display_name_from_template(self):
        """Cached payload must use display_name from ReportItemTemplate, not the raw Excel label."""
        db = self.SessionLocal()
        self._seed_upload_templates(db)

        result = preview_history_import(db, build_report_upload(), build_shipping_upload())
        self.assertTrue(result.can_commit)

        payload = get_history_import_session(result.import_session_id)
        self.assertIsNotNone(payload)

        postal_row = next(r for r in payload["report_rows"] if r["category"] == "postal")
        self.assertEqual(postal_row["display_name"], "北京邮发-本市")

        retail_row = next(r for r in payload["report_rows"] if r["category"] == "retail")
        self.assertEqual(retail_row["display_name"], "北京报零-西部")

        # category_name must NOT be present in the cached payload (replaced by display_name)
        self.assertNotIn("category_name", postal_row)
        db.close()

    def test_preview_rejects_blank_publish_date(self):
        db = self.SessionLocal()
        db.add(ReportItemTemplate(
            category="postal", sub_category="本市", display_name="北京邮发-本市",
            default_value=0, is_variable=False, destination="邮局", sort_order=1,
        ))
        db.commit()

        result = preview_history_import(
            db, build_report_upload_with_blank_publish_date(), build_shipping_upload()
        )

        self.assertFalse(result.can_commit)
        self.assertFalse(result.readiness.can_commit)
        self.assertTrue(any("出版日期" in e for e in result.errors))
        db.close()

    def test_preview_rejects_non_numeric_issue_number(self):
        db = self.SessionLocal()
        db.add(ReportItemTemplate(
            category="postal", sub_category="本市", display_name="北京邮发-本市",
            default_value=0, is_variable=False, destination="邮局", sort_order=1,
        ))
        db.commit()

        result = preview_history_import(
            db, build_report_upload_with_invalid_issue_number(), build_shipping_upload()
        )

        self.assertFalse(result.can_commit)
        self.assertFalse(result.readiness.can_commit)
        self.assertTrue(any("期号" in e for e in result.errors))
        db.close()

    def test_preview_raises_422_for_invalid_workbook_bytes(self):
        db = self.SessionLocal()

        with self.assertRaises(HTTPException) as ctx:
            preview_history_import(db, b"not-an-xlsx", build_shipping_upload())

        self.assertEqual(ctx.exception.status_code, 422)
        db.close()


class HistoryImportCommitTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(bind=self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

    def _seed_templates(self, db) -> None:
        db.add_all([
            ReportItemTemplate(
                category="postal", sub_category="本市", display_name="北京邮发-本市",
                default_value=0, is_variable=False, destination="邮局", sort_order=1,
            ),
            ReportItemTemplate(
                category="retail", sub_category="西部", display_name="北京报零-西部",
                default_value=0, is_variable=True, destination="零售点", sort_order=2,
            ),
        ])
        db.commit()

    def test_commit_creates_issue_and_all_records(self):
        """Successful commit creates Issue, ReportEntry, TempPrintDetail, ShippingDetail from session."""
        db = self.SessionLocal()
        self._seed_templates(db)

        preview = preview_history_import(db, build_report_upload(), build_shipping_upload())
        self.assertTrue(preview.can_commit)

        result = commit_history_import(db, preview.import_session_id)

        issue = db.query(Issue).filter(Issue.issue_number == 2648).first()
        self.assertIsNotNone(issue)
        self.assertEqual(issue.issue_number, 2648)
        self.assertEqual(str(issue.publish_date), "2026-04-20")
        self.assertEqual(issue.status, IssueStatus.draft)
        self.assertEqual(issue.page_count, 24)
        self.assertEqual(issue.notes, "测试备注")

        entries = db.query(ReportEntry).filter(ReportEntry.issue_id == issue.id).all()
        self.assertEqual(len(entries), 2)
        cats = {e.category for e in entries}
        self.assertIn("postal", cats)
        self.assertIn("retail", cats)

        temps = db.query(TempPrintDetail).filter(TempPrintDetail.issue_id == issue.id).all()
        self.assertEqual(len(temps), 1)
        self.assertEqual(temps[0].department, "编辑部")
        self.assertEqual(temps[0].quantity, 20)

        shipping = db.query(ShippingDetail).filter(ShippingDetail.issue_number == 2648).all()
        self.assertEqual(len(shipping), 1)
        self.assertEqual(shipping[0].channel, "邮发")

        self.assertEqual(result.report_entry_count, 2)
        self.assertEqual(result.temp_detail_count, 1)
        self.assertEqual(result.shipping_detail_count, 1)
        self.assertEqual(result.issue_number, 2648)
        self.assertIsNotNone(result.issue_id)
        db.close()

    def test_commit_persists_original_zto_shipping_fields(self):
        db = self.SessionLocal()
        self._seed_templates(db)

        preview = preview_history_import(
            db,
            build_report_upload(),
            build_original_zto_shipping_upload(),
        )
        self.assertTrue(preview.can_commit)

        result = commit_history_import(db, preview.import_session_id)

        shipping = db.query(ShippingDetail).filter(ShippingDetail.issue_number == 2648).all()
        self.assertEqual(len(shipping), 7)
        by_name = {row.name: row for row in shipping}
        self.assertEqual(by_name["叶剑"].channel, "渠道订阅")
        self.assertEqual(by_name["叶剑"].company, "广州日报")
        self.assertEqual(by_name["赵叶"].station_name, "北京站")
        self.assertEqual(by_name["赵叶"].confirmation, "☑")
        self.assertEqual(by_name["丁联诚"].period_count, 2)
        self.assertEqual(by_name["丁联诚"].deadline, "2026-05-01")
        self.assertEqual(by_name["宣传部5号格"].frequency, "月")
        self.assertEqual(result.shipping_detail_count, 7)
        db.close()

    def test_commit_raises_400_for_missing_session(self):
        """Missing or expired session raises HTTPException with status_code 400."""
        db = self.SessionLocal()
        with self.assertRaises(HTTPException) as ctx:
            commit_history_import(db, "nonexistent-session-id")
        self.assertEqual(ctx.exception.status_code, 400)
        db.close()

    def test_commit_raises_409_if_issue_already_exists(self):
        """If issue already exists at commit time, raises HTTPException with status_code 409."""
        db = self.SessionLocal()
        self._seed_templates(db)

        preview = preview_history_import(db, build_report_upload(), build_shipping_upload())
        self.assertTrue(preview.can_commit)

        # Simulate issue created between preview and commit (race / double-commit)
        db.add(Issue(
            issue_number=2648,
            publish_date=date(2026, 4, 20),
            status=IssueStatus.draft,
        ))
        db.commit()

        with self.assertRaises(HTTPException) as ctx:
            commit_history_import(db, preview.import_session_id)
        self.assertEqual(ctx.exception.status_code, 409)
        db.close()


if __name__ == "__main__":
    unittest.main()
