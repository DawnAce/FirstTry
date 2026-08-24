from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.api.reports import get_report
from app.models import (
    Issue,
    IssueAuditSnapshot,
    IssueStatus,
    PublicationSchedule,
    ShippingDetail,
    ShippingDeferral,
    ShippingFulfillmentAdjustment,
    ShippingPackageAllocation,
    ShippingWaybillImportBatch,
    ShippingWaybillImportRow,
    WaybillImportStatus,
    WaybillMatchStatus,
)
from app.models.user import User, UserRole
from app.schemas.shipping_detail import ShippingDetailOut
from app.schemas.shipping_waybill import (
    FulfillmentAdjustmentAttributionIn,
    FulfillmentAdjustmentIn,
    ConsolidatedAllocationIn,
    ConsolidatedPackageIn,
    ShippingDeferralBulkIn,
    ShippingDeferralItemIn,
    ShippingPlanTransferIn,
    WaybillBulkMatchIn,
    WaybillImportBatchOut,
    WaybillImportRowCreate,
    WaybillImportRowUpdate,
)
from app.services.shipping_waybill_service import (
    _deferral_target,
    add_import_row,
    attribute_fulfillment_adjustment,
    bulk_match_import_rows,
    confirm_import,
    convert_import_row_to_warehouse_stock_in,
    create_fulfillment_adjustment,
    create_shipping_deferrals,
    create_consolidated_package,
    fulfillment_summary,
    get_draft_import,
    parse_waybill_workbook,
    preview_import,
    repair_postal_30_preview_rows,
    transfer_shipping_plan_quantity,
    update_import_row,
)


def _workbook_bytes() -> bytes:
    wb = Workbook()
    main = wb.active
    main.title = "中通+顺丰到付"
    main.append([None, None, None, None, "电话", "地址", "姓名", "份数"])
    main.append(["经营报", None, "73592817527861", None, "13800000000", "北京市测试路1号", "张三", 1])
    main.append(["经营报", None, "73592817528444", None, "13800000000", "北京市测试路1号", "张三", 2])

    internal = wb.create_sheet("299（备用74+社用225）")
    internal.append(["姓名", "地址", "电话", "份数", "刊物", "备注"])
    internal.append(["库房", "中通库房", "13900000000", 2, "中国经营报", "备用报"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _mafei_warehouse_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "306（备用71+社用235）"
    ws.append(["姓名", "地址", "电话", "份数", "刊物", "备注"])
    ws.append(["马飞", "中通库房", "13800000000", 71, "中国经营报", "库房留存"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _supplement_workbook_bytes(detail_id: int, tracking_no: str = "73592817529999") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "运单补录"
    ws.append(["发货明细ID", "快递公司", "运单号", "实发份数", "姓名", "电话", "地址"])
    ws.append([detail_id, "中通", tracking_no, 1, "待补客户", "13700000000", "北京市测试路2号"])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _unrecognized_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "临时格式"
    ws.append(["说明", "内容"])
    ws.append(["王五", "此行未按已知列布局排列", 3])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _high_speed_rail_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "巴出高铁110"
    ws.append(["刊物", "打印名称", "单号", "打印名称", "电话", "地址", "展示名称", "姓名", "份数"])
    ws.append([
        "中国经营报5-18日",
        "赵叶5",
        "73708644153509",
        "赵叶5",
        "15810698235",
        "北京市东城区北京站广场西侧商务专用通道 赵叶 15810698235",
        "商务座候车区（北京站）",
        "赵叶",
        5,
        "☑",
        "北京站",
    ])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _compact_high_speed_rail_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "已出高铁110"
    ws.append([None, None, None, None, "联系电话", "收货地址", "联系人", "投放数量", "站名", "站厅名称"])
    ws.append([
        "中国经营报8-24日",
        "测试客户5",
        "73708644153510",
        "测试客户5",
        "13800000000",
        "北京市测试地址",
        "测试客户",
        5,
        "北京站",
        "商务候车厅",
    ])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _postal_30_workbook_bytes(*, compact: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "已出邮政30"
    ws.append(["刊物", "打印名称", "单号", "打印名称", "电话", "空列/地址", "地址/姓名", "姓名/份数", "份数"])
    if compact:
        ws.append([
            "经营报1-5日", "上犹县政协办9", "9442663534703", "上犹县政协办9",
            "0797-8541235", "江西省赣州市上犹县县政府大楼232室政协办",
            "上犹县政协办", 9, "中国经营报",
        ])
    else:
        ws.append([
            "经营报5-18日", "上犹县政协办9", "9407632598208", "上犹县政协办9",
            "0797-8541235", None, "江西省赣州市上犹县县政府大楼232室政协办",
            "上犹县政协办", 9,
        ])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _split_chengdu_packages_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "中通+顺丰到付970"
    ws.append([None, None, None, None, "电话", "地址", "姓名", "份数"])
    for tracking_no, quantity in [
        ("73592817556132", 65),
        ("73592817560608", 100),
        ("73592817561275", 100),
        ("73592817561872", 100),
    ]:
        ws.append([
            "经营报1-26日", None, tracking_no, None, "15719468023",
            "成都市双流文星镇通关路86号A1－A4杂志铺/\n028－85312807",
            "肖波", quantity,
        ])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _bulk_waybill_workbook_bytes(row_count: int = 67) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "中通批量运单"
    ws.append([None, None, None, None, "电话", "地址", "姓名", "份数"])
    for index in range(row_count):
        ws.append([
            "经营报",
            None,
            f"73592817{index:06d}",
            None,
            f"138000{index:05d}",
            f"北京市测试路{index}号",
            f"批量客户{index}",
            1,
        ])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _single_waybill_bytes(name: str, phone: str, address: str, quantity: int = 1) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "中通运单"
    ws.append([None, None, None, None, "电话", "地址", "姓名", "份数"])
    ws.append(["经营报", None, f"735928176{phone[-5:]}", None, phone, address, name, quantity])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _original_zto_shipping_detail_bytes() -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "每周合计"
    summary.append(["2026年1月5日《中国经营报》中通发货表", "总第2635期"])
    summary.append(["各渠道统计合计"])
    summary.append(["传统零售渠道", 500])

    corporate = wb.create_sheet("每周（对公）")
    corporate.append(["2026年1月5日《中国经营报》中通发货表", None, "总第2635期"])
    corporate.append(["姓名", "地址", "电话", "份数", "刊物", "渠道", "子渠道", "签约公司", "频率", "运输方式", "备注"])
    corporate.append(["张三", "北京市测试路1号", "13800000000", 2, "中国经营报", "渠道订阅", "", "", "周", "中通物流", ""])
    corporate.append([None, None, "合计", 2])

    reader = wb.create_sheet("每周（读者）")
    reader.append(["2026年1月5日《中国经营报》中通发货表", None, "总第2635期"])
    reader.append(["姓名", "地址", "电话", "份数", "刊物", "截止日期", "渠道", "子渠道", "签约公司", "频率", "运输方式", "备注"])
    reader.append(["李四", "上海市测试路2号", "13900000000", 1, "中国经营报", "长期", "个人订阅", "", "", "周", "中通物流", ""])

    rail = wb.create_sheet("高铁展示")
    rail.append(["2026年1月5日《中国经营报》中通发货表", None, "总第2635期"])
    rail.append([])
    rail.append(["城市", "序号", "车站", "候车厅", "联系人", "电话", "地址", "份数", "确认", "备注", "渠道", "子渠道", "公司", "频率", "运输方式"])
    rail.append(["北京", 1, "北京站", "商务厅", "王五", "13700000000", "北京站测试地址", 3, "", "", "赠阅", "", "", "周", "中通物流"])

    monthly = wb.create_sheet("月底-整月")
    monthly.append(["2026年1月《中国经营报》中通发货表", None, "总第2635期、第2636期"])
    monthly.append(["姓名", "地址", "电话", "期数", "份数", "刊物", "截止日期", "渠道", "子渠道", "签约公司", "频率", "运输方式", "备注"])
    monthly.append(["赵六", "广州市测试路3号", None, 4, 1, "中国经营报", None, "赠阅", "监管", "", "月", "中通物流", ""])

    ancillary = wb.create_sheet("样报缴送清单（当月）-关联赵六")
    ancillary.append(["样报缴送清单"])
    ancillary.append(["接收单位", "某单位"])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _db():
    engine = create_engine(
        "sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)()


def test_parser_recognizes_packages_and_no_tracking_rows():
    rows = parse_waybill_workbook(_workbook_bytes())
    assert len(rows) == 3
    assert sum(row.quantity for row in rows) == 5
    assert sum(bool(row.tracking_no) for row in rows) == 2
    assert sum(row.no_tracking_required for row in rows) == 1


def test_parser_retains_unrecognized_candidate_with_raw_cells():
    rows = parse_waybill_workbook(_unrecognized_workbook_bytes())
    assert len(rows) == 1
    assert rows[0].parse_reason == "未能按当前工作表格式识别，请人工补充"
    assert rows[0].raw_values == ["王五", "此行未按已知列布局排列", "3"]


def test_parser_recognizes_high_speed_rail_sheet_columns():
    rows = parse_waybill_workbook(_high_speed_rail_workbook_bytes())
    assert len(rows) == 1
    assert rows[0].source_sheet == "巴出高铁110"
    assert rows[0].tracking_no == "73708644153509"
    assert rows[0].carrier == "中通"
    assert rows[0].recipient_name == "赵叶"
    assert rows[0].phone == "15810698235"
    assert rows[0].address == "北京市东城区北京站广场西侧商务专用通道 赵叶 15810698235"
    assert rows[0].quantity == 5
    assert rows[0].parse_reason is None


def test_parser_recognizes_compact_high_speed_rail_sheet_columns():
    rows = parse_waybill_workbook(_compact_high_speed_rail_workbook_bytes())
    assert len(rows) == 1
    assert rows[0].source_sheet == "已出高铁110"
    assert rows[0].tracking_no == "73708644153510"
    assert rows[0].recipient_name == "测试客户"
    assert rows[0].phone == "13800000000"
    assert rows[0].address == "北京市测试地址"
    assert rows[0].quantity == 5
    assert rows[0].parse_reason is None


def test_parser_recognizes_compact_postal_30_columns():
    rows = parse_waybill_workbook(_postal_30_workbook_bytes(compact=True))
    assert len(rows) == 1
    assert (rows[0].recipient_name, rows[0].quantity) == ("上犹县政协办", 9)
    assert rows[0].phone == "0797-8541235"
    assert rows[0].address == "江西省赣州市上犹县县政府大楼232室政协办"
    assert rows[0].tracking_no == "9442663534703"
    assert rows[0].carrier == "邮政"


def test_parser_keeps_spaced_postal_30_columns_compatible():
    rows = parse_waybill_workbook(_postal_30_workbook_bytes(compact=False))
    assert len(rows) == 1
    assert (rows[0].recipient_name, rows[0].quantity) == ("上犹县政协办", 9)
    assert rows[0].address == "江西省赣州市上犹县县政府大楼232室政协办"
    assert rows[0].tracking_no == "9407632598208"


def test_repair_compact_postal_30_preview_row_relinks_shipping_detail():
    db = _db()
    issue = Issue(issue_number=2635, publish_date=date(2026, 1, 5), status=IssueStatus.confirmed)
    detail = ShippingDetail(
        issue_number=2635,
        sheet_name="上犹",
        channel="赠阅",
        sub_channel="政府",
        transport="邮政物流",
        frequency="周",
        status="正常",
        name="上犹县政协办",
        phone="0797-8541235",
        address="江西省赣州市上犹县县政府大楼232室政协办",
        quantity=9,
    )
    db.add_all([issue, detail])
    db.flush()
    batch = ShippingWaybillImportBatch(
        issue_id=issue.id,
        issue_number=2635,
        filename="单号经营报1-5日.xlsx",
        file_hash="compact-postal-30",
        status=WaybillImportStatus.previewed.value,
        expected_quantity=9,
        parsed_quantity=0,
        matched_quantity=0,
        pending_quantity=9,
        matched_rows=0,
        unmatched_rows=1,
        warning_count=1,
    )
    db.add(batch)
    db.flush()
    row = ShippingWaybillImportRow(
        batch_id=batch.id,
        source_sheet="已出邮政30",
        source_row=2,
        carrier="邮政",
        tracking_no="9442663534703",
        recipient_name="9",
        phone="0797-8541235",
        address="上犹县政协办",
        quantity=0,
        raw_values=[
            "经营报1-5日", "上犹县政协办9", "9442663534703", "上犹县政协办9",
            "0797-8541235", "江西省赣州市上犹县县政府大楼232室政协办",
            "上犹县政协办", "9", "中国经营报",
        ],
        match_status=WaybillMatchStatus.invalid.value,
    )
    db.add(row)
    db.commit()

    result = repair_postal_30_preview_rows(db, issue_number=2635, username="test")

    db.refresh(row)
    db.refresh(batch)
    assert result.repaired_rows == 1
    assert result.repaired_quantity == 9
    assert row.recipient_name == "上犹县政协办"
    assert row.address == detail.address
    assert row.quantity == 9
    assert row.match_status == WaybillMatchStatus.matched.value
    assert row.shipping_detail_id == detail.id
    assert batch.parsed_quantity == 9
    assert batch.matched_quantity == 9
    assert batch.pending_quantity == 0


def test_parser_does_not_retain_total_row_as_unrecognized_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "299（备用74+社用225）"
    ws.append(["姓名", "地址", "电话", "份数"])
    ws.append(["库房", "中通库房", "13900000000", 299])
    ws.append([None, None, "合计", 299])
    out = BytesIO()
    wb.save(out)
    rows = parse_waybill_workbook(out.getvalue())
    assert len(rows) == 1
    assert rows[0].quantity == 299


def test_parser_preserves_original_zto_shipping_detail_without_tracking_numbers():
    rows = parse_waybill_workbook(_original_zto_shipping_detail_bytes())

    assert len(rows) == 4
    assert [(row.source_sheet, row.source_row) for row in rows] == [
        ("每周（对公）", 3),
        ("每周（读者）", 3),
        ("高铁展示", 4),
        ("月底-整月", 3),
    ]
    assert all(row.tracking_no is None for row in rows)
    assert all(row.parse_reason is None for row in rows)
    assert [(row.recipient_name, row.phone, row.address, row.quantity) for row in rows] == [
        ("张三", "13800000000", "北京市测试路1号", 2),
        ("李四", "13900000000", "上海市测试路2号", 1),
        ("王五", "13700000000", "北京站测试地址", 3),
        ("赵六", "", "广州市测试路3号", 1),
    ]
    assert not any(row.source_sheet in {"每周合计", "样报缴送清单（当月）-关联赵六"} for row in rows)


def test_preview_marks_shipping_detail_rows_as_missing_tracking_without_counting_shipment():
    db = _db()
    user = User(username="tester", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=2635, publish_date=date(2026, 1, 5), status=IssueStatus.confirmed)
    db.add_all([user, issue])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=7,
        shipping_total=7,
        delta=0,
        is_match=True,
    ))
    db.add_all([
        ShippingDetail(issue_number=2635, sheet_name="每周（对公）", channel="渠道订阅", transport="中通物流", frequency="周", status="正常", name="张三", phone="13800000000", address="北京市测试路1号", quantity=2),
        ShippingDetail(issue_number=2635, sheet_name="每周（读者）", channel="个人订阅", transport="中通物流", frequency="周", status="正常", name="李四", phone="13900000000", address="上海市测试路2号", quantity=1),
        ShippingDetail(issue_number=2635, sheet_name="月底-整月", channel="赠阅", transport="中通物流", frequency="月", status="正常", name="赵六", phone=None, address="广州市测试路3号", quantity=1),
        ShippingDetail(issue_number=2635, sheet_name="高铁展示", channel="赠阅", transport="中通物流", frequency="周", status="正常", name="王五", phone="13700000000", address="北京站测试地址", quantity=3),
    ])
    db.commit()

    batch = preview_import(
        db,
        issue.id,
        "2026年1月5日《中国经营报》中通快递发货明细（2635）.xlsx",
        _original_zto_shipping_detail_bytes(),
        user,
    )

    assert batch.parsed_quantity == 7
    assert batch.matched_quantity == 0
    assert batch.matched_rows == 0
    assert batch.unmatched_rows == 4
    assert all(row.match_status == "invalid" for row in batch.rows)
    assert all(row.match_reason == "缺少运单号" for row in batch.rows)
    assert {row.recipient_name for row in batch.rows} == {"张三", "李四", "王五", "赵六"}


def test_preview_confirm_keeps_one_copy_pending_and_supports_multiple_packages():
    db = _db()
    user = User(username="tester", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=2638, publish_date=date(2026, 1, 26), status=IssueStatus.confirmed)
    db.add_all([user, issue])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=6,
        shipping_total=6,
        delta=0,
        is_match=True,
    ))
    recipient = ShippingDetail(
        issue_number=2638,
        sheet_name="每周",
        channel="个人订阅",
        transport="中通物流",
        frequency="周",
        status="正常",
        name="张三",
        phone="13800000000",
        address="北京市测试路1号",
        quantity=3,
    )
    internal = ShippingDetail(
        issue_number=2638,
        sheet_name="留存",
        channel="库房留存",
        transport="库房留存",
        frequency="周",
        status="正常",
        name="库房",
        phone="13900000000",
        address="中通库房",
        quantity=2,
    )
    missing = ShippingDetail(
        issue_number=2638,
        sheet_name="每周",
        channel="个人订阅",
        transport="中通物流",
        frequency="周",
        status="正常",
        name="待补客户",
        phone="13700000000",
        address="北京市测试路2号",
        quantity=1,
    )
    db.add_all([recipient, internal, missing])
    db.commit()

    content = _workbook_bytes()
    preview = preview_import(db, issue.id, "单号.xlsx", content, user)
    assert preview.expected_quantity == 6
    assert preview.parsed_quantity == 5
    assert preview.matched_quantity == 5
    assert preview.pending_quantity == 1
    assert preview.matched_rows == 3
    assert preview.unmatched_rows == 0
    assert len(WaybillImportBatchOut.model_validate(preview).rows) == 3

    first_row = preview.rows[0]
    edited = update_import_row(
        db,
        preview.id,
        first_row.id,
        WaybillImportRowUpdate(recipient_name="人工确认张三", shipping_detail_id=recipient.id),
    )
    assert edited.rows[0].manual_reviewed is True
    assert edited.rows[0].recipient_name == "人工确认张三"
    assert edited.matched_quantity == 5

    same_preview = preview_import(db, issue.id, "重命名.xlsx", content, user)
    assert same_preview.id == preview.id
    assert same_preview.rows[0].recipient_name == "人工确认张三"
    assert get_draft_import(db, issue.id).id == preview.id
    assert db.query(ShippingWaybillImportBatch).count() == 1

    confirmed = confirm_import(db, preview.id, user)
    assert confirmed.status == "confirmed"
    summary = fulfillment_summary(db, issue.id)
    assert summary.expected_quantity == 6
    assert summary.handled_quantity == 5
    assert summary.tracked_quantity == 3
    assert summary.no_tracking_quantity == 2
    assert summary.pending_quantity == 1
    assert summary.package_count == 2
    assert summary.status == "partial"

    db.refresh(recipient)
    db.refresh(internal)
    assert recipient.package_count == 2
    assert recipient.handled_quantity == 3
    assert recipient.fulfillment_status == "shipped"
    assert ShippingDetailOut.model_validate(recipient).package_count == 2
    assert internal.fulfillment_status == "no_tracking_required"
    assert missing.fulfillment_status == "pending"

    supplement = preview_import(db, issue.id, "补单.xlsx", _supplement_workbook_bytes(missing.id), user)
    assert supplement.matched_quantity == 1
    assert supplement.pending_quantity == 0
    confirm_import(db, supplement.id, user)
    completed = fulfillment_summary(db, issue.id)
    assert completed.handled_quantity == 6
    assert completed.pending_quantity == 0
    assert completed.status == "shipped"


def test_manual_review_ignore_add_and_duplicate_recalculation():
    db = _db()
    user = User(username="reviewer", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=3001, publish_date=date(2026, 2, 2), status=IssueStatus.confirmed)
    db.add_all([user, issue])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=2,
        shipping_total=2,
        delta=0,
        is_match=True,
    ))
    first = ShippingDetail(
        issue_number=3001, sheet_name="每周", channel="个人订阅", transport="中通物流",
        frequency="周", status="正常", name="待补客户", phone="13700000000",
        address="北京市测试路2号", quantity=1,
    )
    second = ShippingDetail(
        issue_number=3001, sheet_name="每周", channel="个人订阅", transport="中通物流",
        frequency="周", status="正常", name="人工客户", phone="13600000000",
        address="北京市测试路3号", quantity=1,
    )
    db.add_all([first, second])
    db.commit()

    batch = preview_import(db, issue.id, "待核对.xlsx", _supplement_workbook_bytes(first.id), user)
    row = batch.rows[0]
    ignored = update_import_row(
        db, batch.id, row.id, WaybillImportRowUpdate(ignored=True, ignore_reason="测试忽略")
    )
    assert ignored.rows[0].match_status == "ignored"
    assert ignored.rows[0].match_reason == "已人工忽略：测试忽略"
    assert ignored.unmatched_rows == 0
    assert ignored.matched_quantity == 0

    restored = update_import_row(
        db, batch.id, row.id, WaybillImportRowUpdate(ignored=False, shipping_detail_id=first.id)
    )
    assert restored.rows[0].match_status == "matched"
    assert restored.matched_quantity == 1

    with_duplicate = add_import_row(db, batch.id, WaybillImportRowCreate(
        carrier="中通", tracking_no="73592817529999", quantity=1,
        recipient_name="人工客户", phone="13600000000", address="北京市测试路3号",
        shipping_detail_id=second.id,
    ))
    duplicate = db.query(ShippingWaybillImportRow).filter(
        ShippingWaybillImportRow.batch_id == batch.id,
        ShippingWaybillImportRow.source_sheet == "人工补充",
    ).one()
    assert duplicate.match_status == "duplicate"
    assert with_duplicate.unmatched_rows == 1
    assert with_duplicate.matched_quantity == 1

    fixed = update_import_row(
        db,
        batch.id,
        duplicate.id,
        WaybillImportRowUpdate(tracking_no="73592817530000", shipping_detail_id=second.id),
    )
    assert fixed.matched_quantity == 2
    assert fixed.pending_quantity == 0

    confirm_import(db, batch.id, user)
    try:
        update_import_row(db, batch.id, duplicate.id, WaybillImportRowUpdate(quantity=2))
        assert False, "confirmed batches must be immutable"
    except Exception as exc:
        assert getattr(exc, "status_code", None) == 409


def test_confirmed_unmatched_split_packages_can_be_bulk_linked_then_close_file_gap():
    db = _db()
    user = User(username="operator", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=2638, publish_date=date(2026, 1, 26), status=IssueStatus.confirmed)
    db.add_all([user, issue])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=1321,
        shipping_total=1321,
        delta=0,
        is_match=True,
    ))
    detail = ShippingDetail(
        issue_number=2638,
        sheet_name="每周",
        channel="成都杂志铺",
        transport="中通物流",
        frequency="周",
        status="正常",
        name="肖波",
        phone="15719468023/\n028－85312807",
        address="成都市双流文星镇通关路86号A1－A4杂志铺",
        quantity=365,
    )
    base_detail = ShippingDetail(
        issue_number=2638,
        sheet_name="确认版",
        channel="其他发货",
        transport="中通物流",
        frequency="周",
        status="正常",
        name="已核销基础明细",
        quantity=955,
        shipping_requirement="no_tracking_required",
    )
    db.add_all([detail, base_detail])
    db.commit()

    batch = preview_import(db, issue.id, "单号-经营报1-26日.xlsx", _split_chengdu_packages_bytes(), user)
    assert batch.parsed_quantity == 365
    assert batch.matched_quantity == 0
    assert batch.unmatched_rows == 4
    assert sum(row.quantity for row in batch.rows) == 365

    confirmed = confirm_import(db, batch.id, user)
    assert confirmed.status == "confirmed"
    assert confirmed.pending_quantity == 366
    assert get_draft_import(db, issue.id).id == confirmed.id

    linked = bulk_match_import_rows(
        db,
        confirmed.id,
        WaybillBulkMatchIn(row_ids=[row.id for row in confirmed.rows], shipping_detail_id=detail.id),
        user,
    )
    assert linked.matched_rows == 4
    assert linked.matched_quantity == 365
    assert linked.unmatched_rows == 0
    assert linked.pending_quantity == 1

    db.refresh(detail)
    assert detail.package_count == 4
    assert detail.handled_quantity == 365
    assert {package.quantity for package in detail.packages} == {65, 100}

    completed = create_fulfillment_adjustment(
        db,
        issue.id,
        FulfillmentAdjustmentIn(
            quantity=1,
            reason="客户要求暂停本期发货",
            shipping_detail_id=detail.id,
        ),
        user,
    )
    assert completed.expected_quantity == 1321
    assert completed.planned_quantity == 1320
    assert completed.tracked_quantity == 365
    assert completed.no_tracking_quantity == 955
    assert completed.adjustment_quantity == 1
    assert completed.attributed_adjustment_quantity == 1
    assert completed.unattributed_adjustment_quantity == 0
    assert completed.actual_shipped_quantity == 1320
    assert completed.handled_quantity == 1321
    assert completed.pending_quantity == 0
    assert completed.status == "shipped"
    assert completed.shipment_status == "partial"

    report = get_report(issue.id, db=db)
    assert report.confirmation_summary.confirmed_shipping_total == 1321
    assert report.confirmation_summary.current_shipping_total == 1320
    assert report.confirmation_summary.plan_attributed_quantity == 1
    assert report.confirmation_summary.plan_unexplained_delta == 0
    assert report.confirmation_summary.plan_is_reconciled is True

    try:
        create_fulfillment_adjustment(
            db,
            issue.id,
            FulfillmentAdjustmentIn(quantity=1, reason=" ", shipping_detail_id=detail.id),
            user,
        )
        assert False, "blank adjustment reasons must be rejected"
    except Exception as exc:
        assert getattr(exc, "status_code", None) == 400


def test_historical_unassigned_adjustment_can_be_attributed_without_changing_physical_shipment():
    db = _db()
    user = User(username="historian", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=4001, publish_date=date(2026, 1, 26), status=IssueStatus.confirmed)
    detail = ShippingDetail(
        issue_number=4001,
        sheet_name="确认版",
        channel="个人订阅",
        transport="中通物流",
        frequency="半月",
        status="停发",
        name="暂停寄送客户",
        quantity=1,
        shipping_requirement="no_tracking_required",
    )
    db.add_all([user, issue, detail])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=2,
        shipping_total=2,
        delta=0,
        is_match=True,
    ))
    legacy = ShippingFulfillmentAdjustment(
        issue_id=issue.id,
        issue_number=issue.issue_number,
        quantity=1,
        reason="每月两次合寄 · 暂停寄送",
        created_by=user.id,
    )
    db.add(legacy)
    db.commit()

    before = fulfillment_summary(db, issue.id)
    before_report = get_report(issue.id, db=db)
    assert before_report.confirmation_summary.plan_is_reconciled is False
    assert before_report.confirmation_summary.unattributed_adjustment_quantity == 1
    assert before.actual_shipped_quantity == 1
    assert before.adjustment_quantity == 1
    assert before.unattributed_adjustment_quantity == 1
    assert before.pending_quantity == 0
    assert before.status == "shipped"
    assert before.shipment_status == "partial"
    assert before.adjustments[0].is_attributed is False

    after = attribute_fulfillment_adjustment(
        db,
        legacy.id,
        FulfillmentAdjustmentAttributionIn(shipping_detail_id=detail.id),
        user,
    )
    assert after.attributed_adjustment_quantity == 1
    assert after.unattributed_adjustment_quantity == 0
    assert after.actual_shipped_quantity == 1
    assert after.pending_quantity == 0
    assert after.adjustments[0].detail_name_snapshot == "暂停寄送客户"

    after_report = get_report(issue.id, db=db)
    assert after_report.confirmation_summary.plan_attributed_quantity == 1
    assert after_report.confirmation_summary.plan_unexplained_delta == 0
    assert after_report.confirmation_summary.plan_is_reconciled is True
    assert after_report.confirmation_summary.unattributed_adjustment_quantity == 0

    db.refresh(detail)
    assert detail.fulfillment_status == "no_tracking_required"
    assert detail.handled_quantity == 1


def test_explicit_reparse_replaces_the_active_draft():
    db = _db()
    user = User(username="reparser", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=3002, publish_date=date(2026, 2, 9), status=IssueStatus.confirmed)
    detail = ShippingDetail(
        issue_number=3002, sheet_name="每周", channel="个人订阅", transport="中通物流",
        frequency="周", status="正常", name="待补客户", phone="13700000000",
        address="北京市测试路2号", quantity=1,
    )
    db.add_all([user, issue, detail])
    db.commit()

    original = preview_import(db, issue.id, "旧草稿.xlsx", _supplement_workbook_bytes(detail.id), user)
    update_import_row(
        db, original.id, original.rows[0].id,
        WaybillImportRowUpdate(recipient_name="旧的人工修正", shipping_detail_id=detail.id),
    )
    replacement = preview_import(
        db,
        issue.id,
        "新草稿.xlsx",
        _supplement_workbook_bytes(detail.id, "73592817531111"),
        user,
        reparse=True,
    )
    assert db.query(ShippingWaybillImportBatch).count() == 1
    assert db.query(ShippingWaybillImportRow).count() == 1
    assert replacement.filename == "新草稿.xlsx"
    assert replacement.rows[0].tracking_no == "73592817531111"
    assert replacement.rows[0].recipient_name == "待补客户"


def test_preview_bulk_inserts_waybill_rows_in_one_database_statement():
    db = _db()
    user = User(username="bulk-importer", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=3003, publish_date=date(2026, 2, 16), status=IssueStatus.confirmed)
    db.add_all([user, issue])
    db.commit()

    insert_calls: list[tuple[bool, int]] = []

    def capture_row_insert(_conn, _cursor, statement, parameters, _context, executemany):
        if "INSERT INTO shipping_waybill_import_rows" in statement:
            insert_calls.append((executemany, len(parameters) if executemany else 1))

    engine = db.get_bind()
    event.listen(engine, "before_cursor_execute", capture_row_insert)
    try:
        batch = preview_import(
            db,
            issue.id,
            "67条运单.xlsx",
            _bulk_waybill_workbook_bytes(),
            user,
        )
    finally:
        event.remove(engine, "before_cursor_execute", capture_row_insert)

    assert len(batch.rows) == 67
    assert insert_calls == [(True, 67)]


def test_month_end_deferrals_are_separate_from_unexplained_pending_quantity():
    db = _db()
    user = User(username="month-end", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=5001, publish_date=date(2026, 5, 18), status=IssueStatus.confirmed)
    normal = ShippingDetail(
        issue_number=5001, sheet_name="每周", channel="个人订阅", transport="中通物流",
        frequency="周", status="正常", name="本期客户", phone="13800000001",
        address="北京市本期路1号", quantity=1,
    )
    month_end = ShippingDetail(
        issue_number=5001, sheet_name="月底-整月", channel="个人订阅", transport="中通物流",
        frequency="月", status="正常", name="月底客户", phone="13800000002",
        address="北京市月底路2号", quantity=3,
    )
    db.add_all([user, issue, normal, month_end])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id, snapshot_type="confirm", report_total=4, shipping_total=4,
        delta=0, is_match=True,
    ))
    db.commit()

    batch = preview_import(
        db, issue.id, "本期运单.xlsx",
        _single_waybill_bytes(normal.name, normal.phone, normal.address), user,
    )
    before = fulfillment_summary(db, issue.id)
    gap = next(item for item in before.gap_details if item.shipping_detail_id == month_end.id)
    assert gap.remaining_quantity == 3
    assert gap.suggested_month_end is True

    deferred = create_shipping_deferrals(
        db,
        issue.id,
        ShippingDeferralBulkIn(
            reason="月底合寄 · 本期报纸随月底最后一期统一寄送",
            items=[ShippingDeferralItemIn(shipping_detail_id=month_end.id, quantity=3)],
        ),
        user,
    )
    assert deferred.deferred_quantity == 3
    assert deferred.unexplained_pending_quantity == 1
    gap = next(item for item in deferred.gap_details if item.shipping_detail_id == month_end.id)
    assert gap.deferred_quantity == 3
    assert gap.remaining_quantity == 0

    confirm_import(db, batch.id, user)
    confirmed = fulfillment_summary(db, issue.id)
    assert confirmed.actual_shipped_quantity == 1
    assert confirmed.pending_quantity == 3
    assert confirmed.deferred_quantity == 3
    assert confirmed.unexplained_pending_quantity == 0


def test_twice_monthly_deferrals_target_second_and_last_issue_of_five_issue_month():
    db = _db()
    user = User(username="twice-monthly", password_hash="x", role=UserRole.admin)
    publish_dates = [
        date(2026, 8, 3),
        date(2026, 8, 10),
        date(2026, 8, 17),
        date(2026, 8, 24),
        date(2026, 8, 31),
    ]
    db.add_all([
        PublicationSchedule(year=2026, issue_number=6001 + index, publish_date=publish_date)
        for index, publish_date in enumerate(publish_dates)
    ])
    db.add(user)
    db.commit()

    created_deferrals = []
    for index in (0, 2):
        issue = Issue(
            issue_number=6001 + index,
            publish_date=publish_dates[index],
            status=IssueStatus.confirmed,
        )
        normal = ShippingDetail(
            issue_number=issue.issue_number, sheet_name="每周", channel="个人订阅",
            transport="中通物流", frequency="周", status="正常",
            name=f"正常客户{index}", phone=f"1380000060{index}",
            address=f"北京市每周路{index}号", quantity=1,
        )
        consolidated = ShippingDetail(
            issue_number=issue.issue_number, sheet_name="每月两次合寄", channel="个人订阅",
            transport="中通物流", frequency="每月两次", status="正常",
            name="两次合寄客户", phone="13900000666",
            address="北京市合寄路6号", quantity=1,
        )
        db.add_all([issue, normal, consolidated])
        db.flush()
        db.add(IssueAuditSnapshot(
            issue_id=issue.id, snapshot_type="confirm", report_total=2,
            shipping_total=2, delta=0, is_match=True,
        ))
        db.commit()
        preview_import(
            db,
            issue.id,
            f"{issue.issue_number}.xlsx",
            _single_waybill_bytes(normal.name, normal.phone, normal.address),
            user,
        )
        summary = create_shipping_deferrals(
            db,
            issue.id,
            ShippingDeferralBulkIn(
                deferral_type="twice_monthly_consolidation",
                reason="每月两次合寄 · 前两期一批、当月剩余期次月底一批",
                items=[ShippingDeferralItemIn(shipping_detail_id=consolidated.id, quantity=1)],
            ),
            user,
        )
        assert summary.twice_monthly_deferred_quantity == 1
        assert summary.month_end_deferred_quantity == 0
        created_deferrals.append(summary.deferrals[0])

    assert created_deferrals[0].target_issue_number == 6002
    assert created_deferrals[0].target_publish_date == date(2026, 8, 10)
    assert created_deferrals[0].consolidation_batch == "2026-08-first"
    assert created_deferrals[1].target_issue_number == 6005
    assert created_deferrals[1].target_publish_date == date(2026, 8, 31)
    assert created_deferrals[1].consolidation_batch == "2026-08-second"


def test_twice_monthly_four_issue_month_uses_two_plus_two_batches():
    db = _db()
    publish_dates = [date(2026, 9, day) for day in (7, 14, 21, 28)]
    db.add_all([
        PublicationSchedule(year=2026, issue_number=6101 + index, publish_date=publish_date)
        for index, publish_date in enumerate(publish_dates)
    ])
    first_issue = Issue(issue_number=6101, publish_date=publish_dates[0], status=IssueStatus.confirmed)
    third_issue = Issue(issue_number=6103, publish_date=publish_dates[2], status=IssueStatus.confirmed)
    db.add_all([first_issue, third_issue])
    db.commit()

    assert _deferral_target(db, first_issue, "twice_monthly_consolidation") == (
        6102,
        date(2026, 9, 14),
        "2026-09-first",
    )
    assert _deferral_target(db, third_issue, "twice_monthly_consolidation") == (
        6104,
        date(2026, 9, 28),
        "2026-09-second",
    )


def test_one_consolidated_package_fulfills_same_recipient_across_issues():
    db = _db()
    user = User(username="consolidator", password_hash="x", role=UserRole.admin)
    deferrals = []
    issues = []
    for index, quantity in enumerate((1, 2), start=1):
        issue = Issue(
            issue_number=5100 + index,
            publish_date=date(2026, 5, 11 + index * 7),
            status=IssueStatus.confirmed,
        )
        normal = ShippingDetail(
            issue_number=issue.issue_number, sheet_name="每周", channel="个人订阅",
            transport="中通物流", frequency="周", status="正常",
            name=f"当期客户{index}", phone=f"1380000010{index}",
            address=f"北京市当期路{index}号", quantity=1,
        )
        month_end = ShippingDetail(
            issue_number=issue.issue_number, sheet_name="月底-整月", channel="个人订阅",
            transport="中通物流", frequency="月", status="正常",
            name="同一月底客户", phone="13900000000", address="北京市月底路8号",
            quantity=quantity,
        )
        db.add_all([issue, normal, month_end])
        db.flush()
        db.add(IssueAuditSnapshot(
            issue_id=issue.id, snapshot_type="confirm", report_total=quantity + 1,
            shipping_total=quantity + 1, delta=0, is_match=True,
        ))
        db.commit()
        batch = preview_import(
            db, issue.id, f"{issue.issue_number}.xlsx",
            _single_waybill_bytes(normal.name, normal.phone, normal.address), user,
        )
        confirm_import(db, batch.id, user)
        summary = create_shipping_deferrals(
            db,
            issue.id,
            ShippingDeferralBulkIn(
                reason="月底合寄 · 本期报纸随月底最后一期统一寄送",
                items=[ShippingDeferralItemIn(shipping_detail_id=month_end.id, quantity=quantity)],
            ),
            user,
        )
        deferrals.append(summary.deferrals[0])
        issues.append(issue)

    result = create_consolidated_package(
        db,
        ConsolidatedPackageIn(
            carrier="中通",
            tracking_no="73592817688888",
            deferrals=[ConsolidatedAllocationIn(deferral_id=item.id) for item in deferrals],
        ),
        user,
    )
    assert result.quantity == 3
    assert db.query(ShippingPackageAllocation).count() == 2
    assert db.query(ShippingDeferral).filter(ShippingDeferral.status == "pending").count() == 0
    for issue in issues:
        summary = fulfillment_summary(db, issue.id)
        assert summary.pending_quantity == 0
        assert summary.deferred_quantity == 0
        assert summary.unexplained_pending_quantity == 0
        assert summary.gap_details == []


def test_plan_quantity_transfer_keeps_issue_total_unchanged():
    db = _db()
    user = User(username="planner", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=5201, publish_date=date(2026, 5, 18), status=IssueStatus.confirmed)
    reserve = ShippingDetail(
        issue_number=5201, sheet_name="每周（对公）", channel="库房留存",
        transport="库房留存", frequency="周", status="正常", name="马飞",
        address="中通库房", quantity=69,
    )
    db.add_all([user, issue, reserve])
    db.commit()

    result = transfer_shipping_plan_quantity(
        db,
        issue.id,
        ShippingPlanTransferIn(
            source_detail_id=reserve.id,
            quantity=1,
            reason="备用报应为68份，转入缺少的读者明细",
            target_name="缺少的读者",
            target_phone="13800000999",
            target_address="北京市读者路9号",
        ),
        user,
    )
    assert result.source_quantity == 68
    assert result.target_quantity == 1
    assert result.planned_quantity == 69
    assert db.query(ShippingDetail).filter(ShippingDetail.issue_number == 5201).count() == 2


def test_mafei_warehouse_retention_requires_stock_in_adjustment():
    db = _db()
    user = User(username="warehouse", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=5202, publish_date=date(2026, 5, 25), status=IssueStatus.confirmed)
    reserve = ShippingDetail(
        issue_number=5202, sheet_name="每周（对公）", channel="库房留存",
        transport="库房留存", frequency="周", status="正常", name="马飞",
        address="中通库房", quantity=72,
    )
    db.add_all([user, issue, reserve])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=72,
        shipping_total=72,
        delta=0,
        is_match=True,
    ))
    db.commit()

    try:
        create_fulfillment_adjustment(
            db,
            issue.id,
            FulfillmentAdjustmentIn(
                adjustment_type="no_shipment_required",
                quantity=72,
                reason="无需发货",
                shipping_detail_id=reserve.id,
            ),
            user,
        )
        assert False, "Mafei warehouse retention must reject no-shipment adjustment"
    except Exception as exc:
        assert getattr(exc, "status_code", None) == 400

    completed = create_fulfillment_adjustment(
        db,
        issue.id,
        FulfillmentAdjustmentIn(
            adjustment_type="warehouse_stock_in",
            quantity=72,
            reason="转库留存 · 当期报纸入马飞中通库房备货",
            shipping_detail_id=reserve.id,
        ),
        user,
    )
    assert completed.adjustment_quantity == 72
    assert completed.no_shipment_quantity == 0
    assert completed.warehouse_stock_in_quantity == 72
    assert completed.actual_shipped_quantity == 0
    assert completed.handled_quantity == 72
    assert completed.pending_quantity == 0

    db.expire_all()
    stored = db.query(ShippingDetail).filter(ShippingDetail.id == reserve.id).one()
    assert stored.warehouse_stock_in_quantity == 72
    assert stored.no_shipment_quantity == 0
    assert stored.fulfillment_status == "warehouse_stock_in"


def test_mafei_no_tracking_import_can_be_converted_to_stock_in_atomically():
    db = _db()
    user = User(username="warehouse_import", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=5204, publish_date=date(2026, 6, 8), status=IssueStatus.confirmed)
    reserve = ShippingDetail(
        issue_number=5204,
        sheet_name="库房留存",
        channel="库房留存",
        transport="库房留存",
        frequency="周",
        status="正常",
        name="马飞",
        phone="13800000000",
        address="中通库房",
        quantity=71,
    )
    db.add_all([user, issue, reserve])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=71,
        shipping_total=71,
        delta=0,
        is_match=True,
    ))
    db.commit()

    batch = preview_import(db, issue.id, "单号.xlsx", _mafei_warehouse_workbook_bytes(), user)
    row = batch.rows[0]
    assert row.shipping_detail_id == reserve.id
    assert row.match_status == WaybillMatchStatus.invalid.value
    assert "转库留存/库存入库" in (row.match_reason or "")
    assert batch.matched_quantity == 0

    db.add(ShippingFulfillmentAdjustment(
        issue_id=issue.id,
        issue_number=issue.issue_number,
        shipping_detail_id=reserve.id,
        adjustment_type="no_shipment_required",
        quantity=10,
        reason="历史误标无需发货",
    ))
    db.commit()

    converted = convert_import_row_to_warehouse_stock_in(db, batch.id, row.id, user)
    converted_row = next(item for item in converted.rows if item.id == row.id)
    assert converted_row.match_status == WaybillMatchStatus.ignored.value
    assert converted_row.shipping_detail_id == reserve.id
    assert converted_row.manual_reviewed is True
    assert converted.matched_quantity == 0
    assert converted.pending_quantity == 0

    db.expire_all()
    adjustments = db.query(ShippingFulfillmentAdjustment).filter(
        ShippingFulfillmentAdjustment.shipping_detail_id == reserve.id
    ).all()
    assert sum(item.quantity for item in adjustments) == 71
    assert {item.adjustment_type for item in adjustments} == {"warehouse_stock_in"}
    assert {item.reason for item in adjustments} == {
        "转库留存 · 当期报纸入马飞中通库房备货"
    }
    stored_detail = db.query(ShippingDetail).filter(ShippingDetail.id == reserve.id).one()
    assert stored_detail.shipping_requirement == "tracking_required"

    confirmed = confirm_import(db, batch.id, user)
    assert confirmed.status == WaybillImportStatus.confirmed.value
    summary = fulfillment_summary(db, issue.id)
    assert summary.warehouse_stock_in_quantity == 71
    assert summary.actual_shipped_quantity == 0
    assert summary.pending_quantity == 0


def test_mafei_import_stock_in_conversion_rejects_existing_physical_shipment():
    db = _db()
    user = User(username="warehouse_conflict", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=5205, publish_date=date(2026, 6, 15), status=IssueStatus.confirmed)
    reserve = ShippingDetail(
        issue_number=5205,
        sheet_name="库房留存",
        channel="库房留存",
        transport="库房留存",
        frequency="周",
        status="正常",
        name="马飞",
        phone="13800000000",
        address="中通库房",
        quantity=71,
    )
    db.add_all([user, issue, reserve])
    db.flush()
    db.add(IssueAuditSnapshot(
        issue_id=issue.id,
        snapshot_type="confirm",
        report_total=71,
        shipping_total=71,
        delta=0,
        is_match=True,
    ))
    db.commit()

    batch = preview_import(db, issue.id, "单号.xlsx", _mafei_warehouse_workbook_bytes(), user)
    row = batch.rows[0]
    reserve.shipped_at = datetime(2026, 6, 15, 12, 0)
    reserve.shipped_quantity = 71
    reserve.tracking_no = "TEST-MAFEI-CONFLICT"
    db.commit()

    try:
        convert_import_row_to_warehouse_stock_in(db, batch.id, row.id, user)
        assert False, "physical shipment history must block stock-in conversion"
    except Exception as exc:
        assert getattr(exc, "status_code", None) == 409

    assert db.query(ShippingFulfillmentAdjustment).filter(
        ShippingFulfillmentAdjustment.shipping_detail_id == reserve.id
    ).count() == 0
    db.refresh(row)
    assert row.match_status == WaybillMatchStatus.invalid.value


def test_stock_in_adjustment_is_rejected_for_non_mafei_detail():
    db = _db()
    user = User(username="warehouse_guard", password_hash="x", role=UserRole.admin)
    issue = Issue(issue_number=5203, publish_date=date(2026, 6, 1), status=IssueStatus.confirmed)
    detail = ShippingDetail(
        issue_number=5203, sheet_name="每周", channel="个人订阅",
        transport="中通物流", frequency="周", status="正常", name="普通读者",
        address="北京市测试路", quantity=1,
    )
    db.add_all([user, issue, detail])
    db.commit()

    try:
        create_fulfillment_adjustment(
            db,
            issue.id,
            FulfillmentAdjustmentIn(
                adjustment_type="warehouse_stock_in",
                quantity=1,
                reason="错误库存入库",
                shipping_detail_id=detail.id,
            ),
            user,
        )
        assert False, "stock-in adjustment must be exclusive to Mafei warehouse retention"
    except Exception as exc:
        assert getattr(exc, "status_code", None) == 400
