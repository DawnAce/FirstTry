import os
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ.setdefault("MYSQL_HOST", "localhost")
os.environ.setdefault("MYSQL_USER", "test")
os.environ.setdefault("MYSQL_PASSWORD", "test")
os.environ.setdefault("MYSQL_DATABASE", "test")

import pytest
from fastapi import HTTPException
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from openpyxl import load_workbook

from app.models import (
    BillingType,
    DeliveryMethod,
    FulfillmentAllocation,
    FulfillmentTarget,
    FulfillmentType,
    Issue,
    IssueStatus,
    Order,
    OrderCommercialStatus,
    OrderEntryMethod,
    OrderEvent,
    OrderEventType,
    OrderItem,
    OrderStatus,
    Publication,
    PublicationFormat,
    PublicationSchedule,
    Refund,
    ShippingChannel,
    ShippingDetail,
    ShippingDetailSourceType,
    ShippingDetailSyncStatus,
    SubscriptionTerm,
)
from app.services.excel_service import export_shipping_excel
from app.services.order_service import cancel_order, refund_order, void_order
from app.services.order_shipping_sync_service import (
    apply_order_shipping_sync,
    preview_order_shipping_sync,
)


@pytest.fixture
def db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


def seed_issue(db, issue_number=2655, publish_date=date(2026, 6, 1), is_suspended=False):
    issue = Issue(
        issue_number=issue_number,
        publish_date=publish_date,
        status=IssueStatus.draft,
    )
    schedule = PublicationSchedule(
        year=publish_date.year,
        issue_number=None if is_suspended else issue_number,
        publish_date=publish_date,
        is_suspended=is_suspended,
    )
    db.add_all([issue, schedule])
    db.commit()
    return issue


def seed_active_subscription_order(
    db,
    *,
    issue_from=2650,
    issue_until=None,
    channel=ShippingChannel.zto_outsource,
    fulfillment_type=FulfillmentType.subscription,
    coverage_start_date=date(2026, 1, 5),
    coverage_end_date=date(2026, 6, 29),
):
    order = Order(
        order_code="ORD-2026-000001",
        order_date=date(2026, 1, 1),
        entry_method=OrderEntryMethod.excel_import,
        payer_name="吴娟",
        payer_contact="13800000000",
        status=OrderStatus.active,
        source_platform="微信小程序",
        source_store="CBJ+",
        total_amount=195,
        paid_amount=195,
    )
    db.add(order)
    db.flush()
    item = OrderItem(
        order_id=order.id,
        publication=Publication.cbj,
        publication_format=PublicationFormat.paper,
        fulfillment_type=fulfillment_type,
        billing_type=BillingType.paid,
        subscription_term=SubscriptionTerm.half_year,
        delivery_method=DeliveryMethod.zto_mf,
        term_start_month="2026-01",
        coverage_start_date=coverage_start_date,
        coverage_end_date=coverage_end_date,
        total_quantity=1,
        unit_price=195,
        subtotal=195,
    )
    db.add(item)
    db.flush()
    allocation = FulfillmentAllocation(
        order_item_id=item.id,
        version_no=1,
        effective_from_issue=issue_from,
        effective_until_issue=issue_until,
    )
    db.add(allocation)
    db.flush()
    target = FulfillmentTarget(
        order_item_id=item.id,
        allocation_id=allocation.id,
        recipient_name="张三",
        recipient_phone="13900000000",
        recipient_address="北京市朝阳区测试路 1 号",
        quantity=1,
        shipping_channel=channel,
    )
    db.add(target)
    db.commit()
    return order, item, allocation, target


def test_preview_creates_candidate_for_active_subscription_order(db):
    seed_issue(db)
    order, item, _, target = seed_active_subscription_order(db)

    preview = preview_order_shipping_sync(db, order.id, 2655)

    assert preview.summary.candidates == 1
    assert preview.summary.to_create == 1
    assert preview.summary.conflicts == 0
    row = preview.items[0]
    assert row.action == "create"
    assert row.order_id == order.id
    assert row.order_item_id == item.id
    assert row.fulfillment_target_id == target.id
    assert row.name == "张三"
    assert row.quantity == 1


def test_preview_skips_coverage_based_item_with_missing_coverage_dates(db):
    seed_issue(db)
    order, item, _, _ = seed_active_subscription_order(
        db,
        coverage_start_date=None,
        coverage_end_date=None,
    )

    preview = preview_order_shipping_sync(db, order.id, 2655)

    assert preview.summary.candidates == 0
    assert preview.summary.to_create == 0
    assert preview.summary.skipped == 1
    assert preview.items[0].action == "skip"
    assert preview.items[0].order_item_id == item.id
    assert preview.items[0].reason == "覆盖期缺失"


def test_shipping_detail_has_unique_order_target_issue_index():
    index = next(
        (
            idx
            for idx in ShippingDetail.__table__.indexes
            if idx.name == "uq_shipping_detail_order_target_issue"
        ),
        None,
    )

    assert index is not None
    assert index.unique is True
    assert [column.name for column in index.columns] == [
        "issue_number",
        "order_id",
        "order_item_id",
        "fulfillment_target_id",
    ]


def test_apply_creates_shipping_detail_and_order_event(db):
    seed_issue(db)
    order, item, _, target = seed_active_subscription_order(db)

    result = apply_order_shipping_sync(db, order.id, 2655, operator_id=7)

    assert result.summary.to_create == 0
    assert result.summary.skipped == 1
    detail = db.query(ShippingDetail).one()
    assert detail.issue_number == 2655
    assert detail.name == "张三"
    assert detail.order_id == order.id
    assert detail.order_item_id == item.id
    assert detail.fulfillment_target_id == target.id
    assert detail.source_type == ShippingDetailSourceType.order_generated
    assert detail.sync_status == ShippingDetailSyncStatus.synced
    event = db.query(OrderEvent).filter(OrderEvent.event_type == OrderEventType.synced_to_shipping).one()
    assert event.operator_id == 7
    assert event.payload_json["issue_number"] == 2655
    assert event.payload_json["created_count"] == 1


def test_apply_is_idempotent_and_updates_linked_synced_row(db):
    seed_issue(db)
    order, _, _, _ = seed_active_subscription_order(db)
    apply_order_shipping_sync(db, order.id, 2655, operator_id=7)
    db.query(FulfillmentTarget).one().recipient_phone = "13911111111"
    db.commit()

    result = apply_order_shipping_sync(db, order.id, 2655, operator_id=7)

    assert result.summary.to_update == 0
    assert result.summary.skipped == 1
    assert db.query(ShippingDetail).count() == 1
    assert db.query(ShippingDetail).one().phone == "13911111111"
    event = (
        db.query(OrderEvent)
        .filter(OrderEvent.event_type == OrderEventType.synced_to_shipping)
        .order_by(OrderEvent.id.desc())
        .first()
    )
    assert event.payload_json["updated_count"] == 1


def test_manually_modified_order_generated_row_blocks_apply(db):
    seed_issue(db)
    order, _, _, _ = seed_active_subscription_order(db)
    apply_order_shipping_sync(db, order.id, 2655, operator_id=7)
    detail = db.query(ShippingDetail).one()
    detail.sync_status = ShippingDetailSyncStatus.manually_modified
    detail.phone = "manual"
    db.commit()

    preview = preview_order_shipping_sync(db, order.id, 2655)
    assert preview.summary.conflicts == 1
    assert preview.items[0].action == "conflict"

    with pytest.raises(HTTPException) as ctx:
        apply_order_shipping_sync(db, order.id, 2655, operator_id=7)
    assert ctx.value.status_code == 409


def test_suspended_issue_returns_empty_preview(db):
    seed_issue(db, is_suspended=True)
    order, _, _, _ = seed_active_subscription_order(db)

    preview = preview_order_shipping_sync(db, order.id, 2655)

    assert preview.summary.candidates == 0
    assert preview.summary.to_create == 0
    assert preview.message == "目标期号为休刊期，不生成发货明细"


def test_non_zto_target_is_skipped(db):
    seed_issue(db)
    order, _, _, _ = seed_active_subscription_order(db, channel=ShippingChannel.post_office)

    preview = preview_order_shipping_sync(db, order.id, 2655)

    assert preview.summary.candidates == 0
    assert preview.summary.skipped == 1
    assert preview.items[0].action == "skip"


def test_refunded_order_is_not_shipped(db):
    # 退款单（commercial_status=refunded）即便 status=active 也不生成发货明细。
    seed_issue(db)
    order, item, _, _ = seed_active_subscription_order(db)
    order.commercial_status = OrderCommercialStatus.refunded
    db.commit()

    preview = preview_order_shipping_sync(db, order.id, 2655)
    assert preview.summary.candidates == 0
    assert preview.summary.to_create == 0
    assert preview.summary.skipped == 1
    assert preview.items[0].action == "skip"
    assert "退款" in preview.items[0].reason

    # apply 也不应建任何发货行
    apply_order_shipping_sync(db, order.id, 2655, operator_id=7)
    assert db.query(ShippingDetail).count() == 0


def test_cancelled_order_is_not_shipped(db):
    seed_issue(db)
    order, _, _, _ = seed_active_subscription_order(db)
    order.commercial_status = OrderCommercialStatus.cancelled
    db.commit()

    preview = preview_order_shipping_sync(db, order.id, 2655)
    assert preview.summary.candidates == 0
    assert "取消" in preview.items[0].reason


def test_void_orphans_generated_details_and_export_excludes_them(db):
    # 作废订单 → 已生成的 order_generated 发货行置 orphaned，且中通导出排除它们。
    issue = seed_issue(db)
    order, _, _, _ = seed_active_subscription_order(db)
    apply_order_shipping_sync(db, order.id, 2655, operator_id=7)
    assert db.query(ShippingDetail).count() == 1

    # 作废前：导出含该行（1 表头 + 1 数据行）
    before = load_workbook(export_shipping_excel(issue.id, db)).active
    assert before.max_row == 2

    void_order(db, order.id, reason="客户取消", operator_id=7)

    detail = db.query(ShippingDetail).one()
    assert detail.sync_status == ShippingDetailSyncStatus.orphaned

    # 作废后：导出仅剩表头，孤儿行被排除
    after = load_workbook(export_shipping_excel(issue.id, db)).active
    assert after.max_row == 1

    event = (
        db.query(OrderEvent)
        .filter(OrderEvent.event_type == OrderEventType.voided)
        .one()
    )
    assert event.payload_json["orphaned_shipping_details"] == 1


def _synced_order(db):
    """Active zto subscription order with one synced shipping_details row for 2655."""
    seed_issue(db)
    order, item, _, _ = seed_active_subscription_order(db)  # paid_amount = 195
    apply_order_shipping_sync(db, order.id, 2655, operator_id=7)
    assert db.query(ShippingDetail).count() == 1
    return order, item


def _detail(db):
    return db.query(ShippingDetail).one()


def test_full_refund_marks_refunded_and_orphans_all_shipping(db):
    order, _ = _synced_order(db)
    refund_order(db, order.id, amount=195, reason="全额退款", operator_id=7)

    db.refresh(order)
    assert order.commercial_status == OrderCommercialStatus.refunded
    assert order.refunded_amount == 195
    assert _detail(db).sync_status == ShippingDetailSyncStatus.orphaned


def test_partial_money_only_refund_keeps_delivery(db):
    order, _ = _synced_order(db)
    refund_order(db, order.id, amount=50, reason="退差价", operator_id=7)

    db.refresh(order)
    assert order.commercial_status == OrderCommercialStatus.partial_refund
    assert order.refunded_amount == 50
    # 纯退钱 → 履约不动，发货行仍 synced
    assert _detail(db).sync_status == ShippingDetailSyncStatus.synced


def test_item_scoped_refund_orphans_that_item(db):
    order, item = _synced_order(db)
    refund_order(db, order.id, amount=50, order_item_id=item.id, operator_id=7)

    db.refresh(order)
    assert order.commercial_status == OrderCommercialStatus.partial_refund
    assert _detail(db).sync_status == ShippingDetailSyncStatus.orphaned


def test_stop_from_issue_only_orphans_from_that_issue(db):
    order, _ = _synced_order(db)
    # 停发起点在该期之后 → 该期(2655)不受影响
    refund_order(db, order.id, amount=20, stop_from_issue=2700, operator_id=7)
    assert _detail(db).sync_status == ShippingDetailSyncStatus.synced

    # 停发起点 <= 该期 → 该期被 orphan
    refund_order(db, order.id, amount=20, stop_from_issue=2655, operator_id=7)
    assert _detail(db).sync_status == ShippingDetailSyncStatus.orphaned


def test_over_refund_rejected(db):
    order, _ = _synced_order(db)
    refund_order(db, order.id, amount=100, operator_id=7)
    with pytest.raises(HTTPException) as ctx:
        refund_order(db, order.id, amount=100, operator_id=7)  # 100+100 > 195
    assert ctx.value.status_code == 422


def test_refund_on_void_order_rejected(db):
    order, _ = _synced_order(db)
    void_order(db, order.id, reason="作废", operator_id=7)
    with pytest.raises(HTTPException) as ctx:
        refund_order(db, order.id, amount=10, operator_id=7)
    assert ctx.value.status_code == 409


def test_cancel_records_full_refund_and_orphans_all(db):
    order, _ = _synced_order(db)
    cancel_order(db, order.id, reason="客户取消", operator_id=7)

    db.refresh(order)
    assert order.commercial_status == OrderCommercialStatus.cancelled
    assert order.refunded_amount == 195  # 实付全额记为退款
    assert _detail(db).sync_status == ShippingDetailSyncStatus.orphaned
    refund_row = db.query(Refund).one()
    assert refund_row.amount == 195
    event = (
        db.query(OrderEvent)
        .filter(OrderEvent.event_type == OrderEventType.cancelled)
        .one()
    )
    assert event.payload_json["refund_amount"] == "195.00"


def test_cancel_after_partial_refunds_only_outstanding(db):
    order, _ = _synced_order(db)
    refund_order(db, order.id, amount=95, operator_id=7)  # 已退 95
    cancel_order(db, order.id, reason="取消", operator_id=7)

    db.refresh(order)
    assert order.commercial_status == OrderCommercialStatus.cancelled
    assert order.refunded_amount == 195  # 95 + 余额 100
    cancel_refund = (
        db.query(Refund).order_by(Refund.id.desc()).first()
    )
    assert cancel_refund.amount == 100  # 仅补退未退的 100


def test_cancel_twice_rejected(db):
    order, _ = _synced_order(db)
    cancel_order(db, order.id, reason="取消", operator_id=7)
    with pytest.raises(HTTPException) as ctx:
        cancel_order(db, order.id, reason="再取消", operator_id=7)
    assert ctx.value.status_code == 409
