"""邮局订报数据生成模块 · 流水线 + 黄金样本回归测试。

- 合成样本（无 PII）：仿真来源A/B 真实版式，走完整流水，断言合并/金额/地区/生成，CI 常驻。
- 黄金样本：桌面 7月/8月 真实样本存在时，逐项比对生成结果（本地校验，不入 CI、不提交 PII）。
"""

import io
import os
import glob
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models import (
    SubscriptionBatchStatus,
    SubscriptionGenerationRun,
    SubscriptionImportStatus,
    SubscriptionOutputArtifact,
    SubscriptionRunStatus,
)
from app.services import attachment_service as att
from app.services import subscription_generation_service as gen_svc
from app.services import subscription_import_service as import_svc
from app.services import subscription_service as batch_svc

A_HEADERS = ["地区", "姓名", "联系电话", "省", "市", "区", "详细地址", "邮编", "年度",
             "产品名称", "起月日", "止月日", "份数", "金额", "渠道", "汇款名称", "汇款日期"]
B_HEADERS = ["序号", "下单日期", "商品", "订阅类型", "姓名", "电话", "地址", "邮政编码",
             "份数", "金额", "起投日期", "终止日期", "订单号", "订单平台", "物流平台",
             "状态", "付款平台", "发票抬头及相关", "备注"]


@pytest.fixture(autouse=True)
def _isolate_uploads(tmp_path, monkeypatch):
    """把落盘根目录重定向到临时目录 —— 测试**绝不**写/删真实 backend/uploads。"""
    monkeypatch.setattr(att, "UPLOAD_ROOT", tmp_path / "uploads")


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    s = Session()
    try:
        yield s
    finally:
        s.close()


def _source_a(records):
    """records: list of (name, phone, full_address, postal)。仿真：完整地址写进「省」列、详细地址留空。"""
    wb = Workbook(); ws = wb.active
    ws.append(A_HEADERS)
    for name, phone, addr, postal in records:
        ws.append(["", name, phone, addr, "", "", "", postal, "2026年", "中国经营报",
                   "0801", "1231", 1, 100, "中经报有赞", "高汇通", "20260316到账100"])
    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


def _source_b_csv(rows):
    """rows: list of (name, phone, addr, postal, logistics, start_date)。带 UTF-8 BOM。"""
    lines = ["《中国经营报》8月订单统计表" + "," * 18]
    lines.append(",".join(B_HEADERS))
    for i, (name, phone, addr, postal, logi, start) in enumerate(rows, start=1):
        row = [str(i), "2025/1/1", "《中国经营报》-纸质版", "全年", name, phone, addr, postal,
               "1", "100", start, "2026/12/31", f"ORD{i}", "淘宝发行部", logi, "未投递", "支付宝", "", ""]
        lines.append(",".join(row))
    return ("﻿" + "\n".join(lines)).encode("utf-8")


def _batch(db, year=2026, month=8, unit_price=None):
    return batch_svc.create_batch(db, {"year": year, "start_month": month, "make_date": None,
                                       "unit_price": unit_price, "notes": None}, operator_id=None)


# --- 合成流水测试（CI 常驻） -------------------------------------------------

def test_merge_a_all_plus_b_postal_this_month(db):
    a = _source_a([
        ("张三", "13800138000", "北京市朝阳区建国路1号", "100000"),
        ("李四", "13911112222", "上海市嘉定区嘉好路325号", "201802"),
        ("王五", "13712345678", "广东省广州市天河区体育西路100号", "510620"),
    ])
    b = _source_b_csv([
        ("魏建", "13700508977", "山西省太原市小店区龙城北街8号", "030032", "邮局", "2026/8/1"),
        ("陈六", "13600000000", "浙江省杭州市西湖区文一路1号", "310012", "中通", "2026/8/1"),   # 非邮局→排除
        ("赵七", "13500000000", "江苏省南京市玄武区中山路1号", "210008", "邮局", "2026/9/1"),   # 非本月→排除
    ])
    batch = _batch(db)
    v = import_svc.create_version(db, batch, [("A", "订阅明细.xlsx", a), ("B", "读者统计.csv", b)], operator_id=None)

    assert v.status == SubscriptionImportStatus.validation_passed
    s = v.summary_json
    assert (s["from_a"], s["from_b"], s["total_count"]) == (3, 1, 4)   # A全量3 + B筛后1
    assert s["months"] == 5 and s["total_copies"] == 4
    assert s["total_amount"] == "400.00"                              # 4 × (1×5×20)
    assert s["region_count"] == 4                                     # 北京/上海/广东/山西


def test_month_multiplier_july(db):
    a = _source_a([("甲", "13000000001", "北京市西城区月坛北街1号", "100037")])
    batch = _batch(db, month=7)
    v = import_svc.create_version(db, batch, [("A", "a.xlsx", a)], operator_id=None)
    assert v.summary_json["months"] == 6                              # 13-7
    assert v.summary_json["total_amount"] == "120.00"                 # 1×6×20


def test_custom_complete_term_unit_price_applies_to_records_and_outputs(db):
    a = _source_a([
        ("甲", "13000000001", "北京市西城区月坛北街1号", "100037"),
    ])
    batch = _batch(db, unit_price=Decimal("95.50"))
    version = import_svc.create_version(
        db, batch, [("A", "a.xlsx", a)], operator_id=None
    )
    assert version.records[0].amount == Decimal("95.50")
    assert version.summary_json["total_amount"] == "95.50"

    batch_svc.activate_version(db, version.id)
    run = gen_svc.generate(db, batch, operator_id=None)
    workbook_artifact = next(
        artifact for artifact in run.artifacts
        if artifact.artifact_type.value == "workbook"
    )
    detail = load_workbook(att.resolve_path(workbook_artifact.stored_path))["北京-明细"]
    assert detail.cell(row=2, column=14).value == "=M2*95.50"

    summary_artifact = next(
        artifact for artifact in run.artifacts
        if artifact.artifact_type.value == "postal_summary"
    )
    summary = load_workbook(att.resolve_path(summary_artifact.stored_path))["汇总"]
    assert summary.cell(row=2, column=7).value == 95.5
    assert summary.cell(row=2, column=8).value == 95.5
    assert summary.cell(row=2, column=7).number_format == gen_svc.NF_YEN2


def test_generate_matches_structure(db):
    a = _source_a([
        ("张三", "13800138000", "北京市朝阳区建国路1号", "100000"),
        ("李四", "13911112222", "上海市嘉定区嘉好路325号", "201802"),
    ])
    batch = _batch(db)
    v = import_svc.create_version(db, batch, [("A", "a.xlsx", a)], operator_id=None)
    batch_svc.activate_version(db, v.id)
    run = gen_svc.generate(db, batch, operator_id=None)

    assert run.status == SubscriptionRunStatus.success
    types = [ar.artifact_type.value for ar in run.artifacts]
    assert types.count("region_detail") == 2                          # 北京 + 上海
    assert "workbook" in types and "postal_summary" in types and "zip" in types
    assert all(ar.sha256 for ar in run.artifacts)
    assert batch.status == SubscriptionBatchStatus.generated

    # 明细金额复刻活公式；汇总表款额=份数×单价。
    wbk = [ar for ar in run.artifacts if ar.artifact_type.value == "workbook"][0]
    wb = load_workbook(att.resolve_path(wbk.stored_path))
    det = wb["北京-明细"]
    assert det.cell(row=2, column=14).value == "=M2*5*20"
    ps = [ar for ar in run.artifacts if ar.artifact_type.value == "postal_summary"][0]
    sw = load_workbook(att.resolve_path(ps.stored_path))["汇总"]
    # 表头
    assert [c.value for c in sw[1]] == ["代码", "报刊名称", "订期", "省份", "条数", "份数", "单价", "款额"]


def test_generation_failure_keeps_previous_artifacts_current_and_cleans_partial_files(
    db, monkeypatch
):
    """新一轮落盘中途失败时，旧产物仍是当前版本，本轮半成品文件被清理。"""
    a = _source_a([
        ("张三", "13800138000", "北京市朝阳区建国路1号", "100000"),
    ])
    batch = _batch(db)
    version = import_svc.create_version(
        db, batch, [("A", "a.xlsx", a)], operator_id=None
    )
    batch_svc.activate_version(db, version.id)
    successful_run = gen_svc.generate(db, batch, operator_id=None)
    old_artifacts = list(successful_run.artifacts)
    old_ids = {artifact.id for artifact in old_artifacts}
    old_paths = [att.resolve_path(artifact.stored_path) for artifact in old_artifacts]

    original_store = att.store_file
    attempted_paths = []
    call_count = 0

    def fail_on_second_store(category, filename, content):
        nonlocal call_count
        call_count += 1
        if call_count == 2:
            raise OSError("模拟第二个产物落盘失败")
        stored_path = original_store(category, filename, content)
        attempted_paths.append(att.resolve_path(stored_path))
        return stored_path

    monkeypatch.setattr(att, "store_file", fail_on_second_store)

    from fastapi import HTTPException

    with pytest.raises(HTTPException) as exc_info:
        gen_svc.generate(db, batch, operator_id=None)
    assert "模拟第二个产物落盘失败" in exc_info.value.detail

    db.expire_all()
    current = (
        db.query(SubscriptionOutputArtifact)
        .filter(
            SubscriptionOutputArtifact.batch_id == batch.id,
            SubscriptionOutputArtifact.is_historical.is_(False),
        )
        .all()
    )
    assert {artifact.id for artifact in current} == old_ids
    assert all(path.exists() for path in old_paths)
    assert attempted_paths and all(not path.exists() for path in attempted_paths)

    failed_run = (
        db.query(SubscriptionGenerationRun)
        .filter(SubscriptionGenerationRun.status == SubscriptionRunStatus.failed)
        .one()
    )
    assert failed_run.artifacts == []
    assert "模拟第二个产物落盘失败" in failed_run.error
    db.refresh(batch)
    assert batch.status == SubscriptionBatchStatus.generated


def test_duplicate_blocks(db):
    a = _source_a([
        ("张三", "13800138000", "北京市朝阳区建国路1号", "100000"),
        ("张三", "13800138000", "北京市朝阳区建国路2号", "100000"),
    ])
    batch = _batch(db)
    v = import_svc.create_version(db, batch, [("A", "a.xlsx", a)], operator_id=None)
    assert v.status == SubscriptionImportStatus.validation_failed
    assert any(i.code == "duplicate_subscriber" for i in v.issues)


def test_csv_gbk_blocks(db):
    content = ("《中国经营报》\n" + ",".join(B_HEADERS)).encode("gbk")
    batch = _batch(db)
    a = _source_a([("甲", "13000000001", "北京市西城区月坛北街1号", "100037")])
    v = import_svc.create_version(db, batch, [("A", "a.xlsx", a), ("B", "b.csv", content)], operator_id=None)
    assert any(i.code == "encoding" and i.level.value == "block" for i in v.issues)


def test_generate_requires_active_version(db):
    from fastapi import HTTPException
    batch = _batch(db)
    with pytest.raises(HTTPException):
        gen_svc.generate(db, batch, operator_id=None)


# --- 黄金样本回归（本地：桌面样本存在才跑；不提交 PII、不进 CI） ------------

GOLDEN_ROOT = os.environ.get("SUBSCRIPTION_GOLDEN_ROOT", os.path.expanduser("~/Desktop"))


def _golden_paths(month_dir):
    root = os.path.join(GOLDEN_ROOT, month_dir)
    src = os.path.join(root, "来源")
    if not os.path.isdir(src):
        return None
    a = glob.glob(os.path.join(src, "*订阅明细*"))
    b = glob.glob(os.path.join(src, "*读者统计*"))
    summary = os.path.join(root, "输出", "北京邮局订报数据", "北京局订报汇总表.xlsx")
    region_dir = os.path.join(root, "输出", "北京邮局订报数据")
    if not (a and b and os.path.exists(summary)):
        return None
    return a[0], b[0], summary, region_dir


def _summary_map(xbytes):
    ws = load_workbook(io.BytesIO(xbytes))["汇总"]
    d = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[3] and r[3] != "合计" and r[0] != "合计":
            d[r[3]] = (r[4], r[5], r[6], r[7])   # 省份 -> (条数,份数,单价,款额)
    return d


@pytest.mark.skipif(_golden_paths("8月") is None, reason="缺 8月 黄金样本（桌面）")
def test_golden_august_exact(db):
    ap, bp, summary_path, region_dir = _golden_paths("8月")
    a = open(ap, "rb").read(); b = open(bp, "rb").read()
    batch = _batch(db, year=2026, month=8)
    v = import_svc.create_version(db, batch, [("A", os.path.basename(ap), a), ("B", os.path.basename(bp), b)], operator_id=None)
    assert v.status == SubscriptionImportStatus.validation_passed, [i.code for i in v.issues if i.level.value == "block"]
    batch_svc.activate_version(db, v.id)
    run = gen_svc.generate(db, batch, operator_id=None)

    mine = _summary_map(open(att.resolve_path([x for x in run.artifacts if x.artifact_type.value == "postal_summary"][0].stored_path), "rb").read())
    gold = _summary_map(open(summary_path, "rb").read())
    assert mine == gold, "北京局订报汇总表 per-region 不一致"

    my_files = sorted(x.filename for x in run.artifacts if x.artifact_type.value == "region_detail")
    gold_files = sorted(os.path.basename(f) for f in glob.glob(os.path.join(region_dir, "*集订分送表*.xlsx")))
    assert my_files == gold_files, "地区分送表文件清单不一致"

    # 版式 1:1：明细/汇总表数据区必须有细边框 + 列宽与黄金样本一致。
    def _wb(art_type, region=None):
        art = next(x for x in run.artifacts if x.artifact_type.value == art_type and (region is None or x.region_name == region))
        return load_workbook(att.resolve_path(art.stored_path))

    gwb = load_workbook(os.path.join(os.path.dirname(summary_path), "..", "北京-2026年8月中国经营报订阅汇总+明细+申请.xlsx"))
    mwb = _wb("workbook")

    def _has_box(c):
        b = c.border
        return all(getattr(b, s) and getattr(b, s).style for s in ("left", "right", "top", "bottom"))

    for sheet in ("北京-汇总", "北京-明细"):
        ms, gs = mwb[sheet], gwb[sheet]
        mw = {k: round(v.width, 1) for k, v in ms.column_dimensions.items() if v.width}
        gw = {k: round(v.width, 1) for k, v in gs.column_dimensions.items() if v.width}
        assert mw == gw, f"[{sheet}] 列宽与黄金样本不一致"
        # 凡黄金样本有内容且有四边框处，生成结果也必须有（数据表格区）。
        for row in gs.iter_rows():
            for gc in row:
                if gc.value is not None and _has_box(gc):
                    assert _has_box(ms.cell(gc.row, gc.column)), f"[{sheet}] {gc.coordinate} 缺边框"

    # 汇总表 + 区域文件：有内容单元格样式与黄金样本逐格一致（0 差异）。
    def _sig(c):
        bd = "".join(s[0] for s in ("left", "right", "top", "bottom") if getattr(c.border, s) and getattr(c.border, s).style)
        val = "F" if isinstance(c.value, str) and c.value.startswith("=") else c.value
        return (val, c.font.name, bool(c.font.bold), c.alignment.horizontal, c.alignment.vertical, bd, c.number_format)

    from openpyxl.utils import range_boundaries

    def _assert_same(mine, gold, name):
        c1, r1, c2, r2 = range_boundaries(gold.calculate_dimension())
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                g = gold.cell(rr, cc)
                if g.value is not None:
                    assert _sig(mine.cell(rr, cc)) == _sig(g), f"[{name}] {g.coordinate} 样式不一致"

    _assert_same(_wb("postal_summary")["汇总"], load_workbook(summary_path)["汇总"], "汇总表")
    sh_file = next(f for f in gold_files if "上海" in f)
    _assert_same(_wb("region_detail", "上海")["数据页"],
                 load_workbook(os.path.join(region_dir, sh_file))["数据页"], "区域-上海")


@pytest.mark.skipif(_golden_paths("7月") is None, reason="缺 7月 黄金样本（桌面）")
def test_golden_july_parses_xls_and_generates(db):
    """7月来源含运营手工剔除的行，无法自动复刻精确 64；此处验证 .xls 解析 + N=6 + 能出文件。"""
    ap, bp, _summary, _region = _golden_paths("7月")
    a = open(ap, "rb").read(); b = open(bp, "rb").read()
    batch = _batch(db, year=2026, month=7)
    v = import_svc.create_version(db, batch, [("A", os.path.basename(ap), a), ("B", os.path.basename(bp), b)], operator_id=None)
    assert v.summary_json["months"] == 6
    assert v.summary_json["from_a"] >= 40 and v.summary_json["total_count"] >= 60
